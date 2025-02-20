#!/usr/bin/env python

"""ssms_issue_tracker_automation_test.py: Oddly enough, the "test" version of ssms_issue_tracker_automation is actually the
go-forward version. Mostly from a lack of versioning and adding this app to the shared server and the webserver. So, going
forward that shouldn't be an issue.

This process is the main process for B2C requests. It pulls in the data from the Sales Issue Tracker and then supplements that
with data from CODS.BA, CODS.SOURCE_BA, CODS.CUST, CODS.CUST_ATTR, CRPL.SALESFORCE, etc. all kinds of data coming in here to help us
make a decision about the requests we are given. 

This process imports the bulk_mapping_process.py file which is where bulk tickets are scraped and evaluated. and then that data
is sent to this process for final evaluation and then sent to a file for team analysis. """

__author__ = "Dave Curtis"
__copyright__ = "Copyright 2023, Dave Curtis"
__credits__ = ["Dave Curtis"]
__license__ = "GPL"
__version__ = "1.0"
__maintainer__ = "Dave Curtist"
__email__ = "dave.curtis1@lumen.com"
__status__ = "Development" 

## -- IMPORTS ---- ##
#standard library

import csv
import cx_Oracle
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import os
import pandas as pd
import pyodbc
import re
import PySimpleGUI as sg
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment



## user-defined modules
import BULK_mapping_process as bmp
import dataGovComp as dgc

# 3rd party
import stat

# ======================================================================================================
# ======================================================================================================
#GLOBAL variables ## ---------------------------------------------

current_date = datetime.today()
        
today = datetime.now().strftime('%m%d%Y') ## strftime changed datetime to string
today_my = datetime.now().strftime('%Y%m')
# print(today_my)
# print(today)


stored_months = []
stored_last_3_months = []
stored_last_6_months = []
stored_last_9_months = []
stored_last_12_months = []
stored_last_13_months = []
stored_this_month = []


stored_fans = []
stored_customers = []

fans_sql_string = ''
to_cust_sql_string = ''

green_can_map_list = []
red_can_map_list = []
can_map_list = []
cannot_map_list = []
research_list = []
maybe_can_map_list = []
definitely_map_these_mamma_jammas_list = []

can_map_bans_list = []
cannot_map_bans_list = []
bans_with_rev_list = []

stored_rev_query_results = []
stored_ban_lob_n_age_query_results = []
stored_to_cust_lob_info_query_results = []
stored_requesting_employee_info_query_results = []

fed = ['563', '484', '485', '486', '487', '129']
internal = ['450', '460']
test_lob = '440' #technically internal
SBG = ['541', '328', '273', '542']
indirect = ['147','512','513','515','516','517','518','130']
enterprise = []

damages = '3-RH5QJP'
cpniIssuesCust = '3-8YT7ZMKQ0H'

allBanODSids = []
allFans = []

oracle = ['ORACLE2E', 'ORACLE2E_FED', 'ORACLREV']
oracleBillingAccounts = []
oracleBillingAccountsODSids = []
oracleBillingAccountsODSidsString = ''

kenan = ['IDC_KENAN', 'KENANFX']
kenanBillingAccounts = []
kenanBillingAccountsODSids = []
kenanBillingAccountsODSidsString = ''
kenanFanPattern = '^.*\-(A|CTL)$'

FANandOrderEntryDict = {}

analyzed_bulk_requests_dict = {}
bulk_ticket_num_list= []

path = r"L:Finance_Data_Governance\Sales Issue Tracker Automation\BAN Mapping\\" + today + " - CSVs for Dataloader"
LDrive_BULK_templates_path = r"L:Finance_Data_Governance\Sales Issue Tracker Automation\BAN Mapping\\" + today + r" - CSVs for Dataloader\BULK Templates"
green_BAN_to_DL_csv_path = f'{path}\{today}_GREEN_BILLING_ACCTS_TO_DATALOADER.csv'
red_BAN_csv_path = f'{path}\{today}_RED_BILLING_ACCTS_TO_MAP.csv'
complete_data_csv_path = f'{path}\{today}_COMPLETE_DATA.csv'
error_csv_path = f'{path}\{today}_ERROR_LIST.csv'
tickets_analysis_template_path = f'{path}\{today}_TICKETS_ANALYSIS_SHEET.csv'
tickets_analysis_template_xlsx_path = f'{path}\{today}_TICKETS_ANALYSIS_SHEET.xlsx'
mdm_analysis_template_path = f'{path}\{today}_MDM_ANALYSIS_SHEET.csv'
B2C_Quarterly_and_Annual_Rev_csv_path = f'{path}\{today}_B2C_QUARTERLY_AND_ANNUAL_REVIEW.csv'
B2C_Quarterly_and_Annual_Rev_xlsx_path = f'{path}\{today}_B2C_QUARTERLY_AND_ANNUAL_REVIEW.xlsx'

testing_csv_path = f'{path}\{today}_TESTING_CSV.csv'

bad_ticket_path = f'{path}\{today}_BAD_TICKET.csv'

voice_services = '.*[- ]*(CCS|EVT|CALL CENTER|PSAP|COTN[- ]LI|COTN[- ]LVT|911|NBLI|OFFNET|OFF[- ]NET|TOLL[- ]FREE|TOLLFREE|NON[- ]BILLABLE|NONBILLABLE|OUTBOUND|INBOUND|(.*[- ](LI|ELS|TF|VC|VOICE|SIP)[- ].*)|([- ](V|LI|ELS|TF|VC|VOICE|SIP)$)).*$'


salesforce_billing_systems = [  'BART', 'CABS', 'CRB', 'CRIS', 'CRISC', 'CRISE',
                                'CRISW', 'ENS', 'IABS', 'LATIS', 'LEXCIS', 'NIBS', 'PARALLELS', 'PPP', 'SAP', 'MBS',
                                'ORACLE2E', 'ORACLE2E_FED', 'ORACLREV', 'SALESFORCE_FIBER']

am_source_systems = ['ACCTMGT', 'ADC_KENAN', 'BRM', 'CDC_KENAN', 'IDC_KENAN', 'KENANFX', 'LEXM', 'SIEBEL',
                     'SIEBEL6_WILTEL', 'SIEBEL8_LATAM', 'VANTIVE', 'KENAN_LATA']


mdm_query_results = []

bac_team = [['00314977', 'Ashley', 'Aikin'], ['00309319', 'Andy', 'Lyons'], ['00306345', 'Lyndsay', 'Fritchell']]
bac_team_empIDs = ['00314977', '00309319', '00306345', '00307101', '00326822'] #00326822 is Dave's for testifying

big_dawg_for_voice_products = '00303166' # Ashley Ellis as of 1/30/2025
big_dawg_for_voice_products_name = 'Ashley Ellis'

mdm_fan_results = []
mdm_cust_results = []

mdm_FANs_list = []
mdm_Cust_list = []

mdm_FANs_dict = {}
mdm_Cust_dict = {}

mdm_fan_results_list = []
mdm_cust_results_list = []

billAcctActvtyStatusCd = ''

containsMSG = []

arDataResultsList = []

b2cQuarterlyDataResultsList = []

ticket_analysis_dict = {}

fanTier2ProductDict = {}

mass_markets_products = ['Copper Broadband', 'Fiber Broadband', 'Retail Video']
mass_markets_and_enterprise_products = ['Equipment', 'Other', 'Voice']
enterprise_products = ['CDN', 'Cloud Services', 'Colocation', 'Contact Center', 'Dark Fiber', 'Enterprise Broadband', \
                        'Ethernet', 'Ethernet SONET', 'Event Conferencing', 'IP', 'IT Solutions', 'Managed Hosting', \
                        'Managed Security', 'NG911', 'Private Line', 'Ready Access', 'SAP', 'SASE', 'UCC', 'UNE and Other',\
                        'VoIP', 'VPN Data Networks', 'Vyvx', 'Wavelengths']

## ==== LINKS === ###
howToFindCRISBans = 'https://centurylink.sharepoint.com/sites/FinCustomerData/SitePages/How-to-Find-CRIS-Billing-Accounts.aspx'

placeholder_customers = [
'3-XHYWBJW7Q7',
'3-A85231',
'3-TVXNBMKCGN',
'3-3RJVSP',
'3-4QJKVCWQPM',
'3-MWQKDS',
'3-C12FHZ',
'3-TJCCFM',
'3-XPYJKZ',
'3-TNHG1C',
'3-XYWRHR',
'3-QMMYX2',
'3-A85628',
'3-L456T6',
'3-J38XSJ',
'3-MLJCLL',
'3-2PVFVT',
'3-GBLDYK',
'3-JK4HW4',
'3-ZFDGVZ63GW',
'3-ZXPFQ7',
'3-975JKD',
'3-849928'
]

cris_to_ens_dict = {}
ens_BANs = []

evalStr = ''
dg_comment = ''

ERP_source_systems = ['CABS','KENANFX','PPP','ZUORA','MBS','SAP','ENS','LATIS']

allTickets = []
latestRqstrCommentDict = {}
latestTeamCommentDict = {}





def getLast13Months(date):
    global stored_last_3_months
    global stored_last_6_months
    global stored_last_9_months
    global stored_last_12_months
    global stored_last_13_months
    global stored_this_month
    global stored_months

    
    mon_num = 14
    for m in range(mon_num):
        past_date = current_date - relativedelta(months = m)
        stored_months.append(past_date.strftime('%Y%m'))
        
    x = 1
    while x < 4:
        stored_last_3_months.append(stored_months[x])
        #print("stored last 3: ", stored_last_3_months)
        x+=1
    x = 1
    while x < 7:
        stored_last_6_months.append(stored_months[x])
        #print("stored last 6: ", stored_last_6_months)
        x+=1
    x = 1
    while x < 10:
        stored_last_9_months.append(stored_months[x])
        #print("stored last 9: ", stored_last_9_months)
        x+=1
    x = 1
    while x < 13:
        stored_last_12_months.append(stored_months[x])
        #print("stored last 12: ", stored_last_12_months)
        x+=1
    x = 1
    while x < 14:
        stored_last_13_months.append(stored_months[x])
        #print("stored last 13: ", stored_last_13_months)
        x+=1
    stored_this_month = stored_months[0]
    
    print("here are the months: ", stored_last_3_months)



def lDriveTest():
    x = 0
    while x == 0:
        thisPath = os.path.exists(path) 
        try:
            if not thisPath:
                os.makedirs(path)
                #print("made a path")
                break
            elif thisPath:
                #print("We're up n' running, baby!")
                break
        except Exception as e:
            # lDrivePop()
            #print("nuh uh")
            lDriveError = ('Please connect to the L Drive.\nThen click "Check L Drive."\nClose this window, or click "Close" to end the program.')
            errorPopup= [
                            [sg.Text(lDriveError, size=(40, 1))],
                            [sg.Button('Check L Drive', key='-CHECK_L_DRIVE-'), sg.Button('Close')]
                        ]
            errorCPNIpopup = sg.Window('L Drive Error', errorPopup, modal=True)
            event, values = errorCPNIpopup.read(close=True)
            if event == sg.WIN_CLOSED or event == 'Close': # if user closes window or clicks cancel
                #print("closing the window")
                break
            elif re.match('^-CHECK_L_DRIVE-$', event):
                #print("made it to match check")
                x == 0
                #print(x)


query_SAVE_DONT_USE_07012022 = (f"""
    SELECT DISTINCT
        M2.[issue_id]
       ,M2.[create_employee_nbr]                     as  'RQSTR_EMP_NBR'
           ,U.firstname                                  as  'RQSTR_FIRSTNAME'                   
           ,U.lastname                                   as  'RQSTR_LASTNAME'
       ,M2.[issue_classification_id]                      
           ,IC.[name]                                    as  'ISSUE_NAME'
           ,IC.[description]                             as  'ISSUE_DESC'
       ,M2.[status_id]                                           
           ,S.[name]                                    as  'STATUS_NAME'
           ,isnull(M2.[source_cust_nbr],'')              as  'SOURCE_CUST_NBR'
           ,isnull(M2.[source_bill_account_nbr],'')      as  'SOURCE_BAN'
           ,isnull(M2.[target_cust_nbr],'')              as  'RQSTD_CUST_NBR'
           ,isnull(M2.[target_cust_name],'')              as  'RQSTD_CUST_NAME'
           ,isnull(M2.[target_cust_sales_chan_name],'')    as 'RQSTD_CUST_SALES_CHAN_NAME'
           ,isnull(M2.[source_finance_account_nbr],'')     as 'RQSTD_FAN'
           ,bd.sales_chan_name                             as 'CURRNT_CUST_SALES_CHAN_NAME'
  FROM salesissuetracker.dbo.issue_m2 M2
	LEFT JOIN salesissuetracker.dbo.issue_reason_codes IRC
		ON M2.issue_reason_code_id = IRC.issue_reason_code_id
	LEFT JOIN salesissuetracker.dbo.issue_classification IC
		ON M2.issue_classification_id = IC.issue_classification_id

  LEFT JOIN [salesissuetracker].[dbo].[user] U
  ON M2.[create_employee_nbr] = U.employee_number

  LEFT JOIN [salesissuetracker].[dbo].[user] U2
  ON M2.modify_employee_nbr = U2.employee_number

  LEFT JOIN [salesissuetracker].[dbo].[user] U3
  ON M2.assigned_to_employee_nbr = U3.employee_number

  LEFT JOIN [salesissuetracker].[dbo].[status] S
  ON M2.status_id = S.status_id

  LEFT JOIN [salesissuetracker].[dbo].[issue_comments] CO
  ON M2.issue_id = CO.[issue_id]

  LEFT JOIN [salesissuetracker].[dbo].busOrgs BD
  ON M2.source_cust_nbr = BD.cust_nbr


  WHERE issue_category_id = '9'  ----issue category code 9 is 'Update Billing Account to Customer Mapping'----
  AND S.status_id = '20'       ---- status id 24 is 'Pending Quarterly Move' / 20 is 'New'
  AND (ic.issue_classification_id = '144') -- issue 144 is single update request
                        """)

# -- > this query can pull in all the tickets (every single one) and only return the first comment in those ticket
#### ---> just comment out the "s.status_id" in the where clause (or change it to only include those that are in a status of new or requester comments)

query = (f"""
SELECT 
	M2.issue_id
	,M2.create_employee_nbr
		,U.firstname
		,U.lastname
	,IC.issue_id
		,comment
		,comment_ts
	,M2.status_id
		,'SINGLE UPDATE' as TICKET_TYPE -- previous was: S.[name]
		,isnull(M2.[source_cust_nbr],'')
		,isnull(M2.[source_bill_account_nbr],'')
		,isnull(M2.[target_cust_nbr],'')
		,isnull(M2.[target_cust_name],'')
		,isnull(M2.[target_cust_sales_chan_name],'')
		,isnull(M2.[source_finance_account_nbr],'')
	,BO.sales_chan_name
    ,M2.issue_root_cause_id
from
	[salesissuetracker].[dbo].[issue_comments] IC
	LEFT JOIN salesissuetracker.dbo.issue_m2 M2
		ON ic.issue_id = M2.issue_id
	LEFT JOIN salesissuetracker.dbo.issue_classification ICL
		ON M2.issue_classification_id = ICL.issue_classification_id
	LEFT JOIN salesissuetracker.dbo.[user] U
		ON M2.create_employee_nbr = U.employee_number
	LEFT JOIN salesissuetracker.dbo.[status] S
		ON M2.status_id = S.status_id
	LEFT JOIN salesissuetracker.dbo.busOrgs BO
		ON M2.source_cust_nbr = BO.cust_nbr
inner join 
	(select issue_id, min(comment_ts) as first_comment
	from [salesissuetracker].[dbo].[issue_comments]
	group by issue_id) cc
	on ic.issue_id = cc.issue_id
where 1=1 
    -- and ic.issue_id = '20036870' --- uncomment and replace ticket ID to run a super small batch
	and issue_category_id = '9'
	and S.status_id in ('20', '26')
	and ic.comment_ts = cc.first_comment
	and icl.issue_classification_id = '144'
	""")

cx_Oracle.init_oracle_client(lib_dir=r"C:\ORACLE\instantclient_19_18")


def querySSMS(query):
    bad_tickets_rows = []
    conn = pyodbc.connect("Driver={SQL Server Native Client 11.0};"
                                   "Server=USIDCVSQL0251;"
                                   "Database=salesissuetracker;"
                                   "UID=CUSTHIER;"
                                   "PWD=SchebangCUSTH!er0072024$")

    '''
    conn = pyodbc.connect("Driver={SQL Server Native Client 11.0};"
                                   "Server=USIDCVSQL0251;"
                                   "Database=salesissuetracker;"
                                   "Trusted_Connection=yes;")
    '''

    query_2 = "SELECT * FROM dbo.status_type"
    c = conn.cursor()
    results = c.execute(query).fetchall()
    c.close()
    ## if no FAN in the results from initial issuetracker query then they are pulled from
    ## full population and sent to the error file
    for result in results:
        #print("ticket no: ", result[0], "customer: ", result[14])
        if (result[14] is None or result[14] == '' or result[14] == ' ') or (result[10] is None or result[10] == '' or result[10] == ' '):
            bad_tickets_rows.append(result)
            #print("bad ticket: ", result[0])
            results.remove(result)
        else:
            continue
    if len(bad_tickets_rows) > 0:
        if os.path.isfile(bad_ticket_path):
            with open(bad_ticket_path, 'a+', newline='') as out:
                csv_out = csv.writer(out)
                for result in bad_tickets_rows:
                    csv_out.writerow(result)
                    #print("this is a bad ticket. Bad ticket. Bad!: ", result[0])
        else:
            with open(bad_ticket_path, 'a+', newline='') as out:
                csv_out = csv.writer(out)
                csv_out.writerow([
                                'CHIT_TICKET_NUM', 'CREATE_EMP_NUMBER', 'RQSTR_FIRST_NAME', 'RQSTR_LAST_NAME',
                                'ISSUE_CLASSIFICATION_ID',
                                'ISSUE_CLASS_NAME', 'ISSUE_CLASS_DESCRPTN',
                                'STATUS_ID',
                                'STATUS_NAME',
                                'SOURCE_CUST_NBR', 'SOURCE_BAN', 'SOURCE_CUST_NAME', 'RQSTD_CUST_NBR',
                                'RQSTD_CUST_NAME', 'RQSTD_CUST_SALES_CHAN_NAME', 'RQSTD_FAN',
                                'CURRNT_CUST_SALES_CHAN_NAME'
                                ])
                for result in bad_tickets_rows:
                    csv_out.writerow(result)
                    #print("this is a bad ticket. Bad ticket. Bad!: ", result[0])
    return(results) 


def queryMDM(qNum, data_string):
    global fans_sql_string
    global to_cust_sql_string
    dsn_tns = cx_Oracle.makedsn(r'RACORAP32-SCAN.CORP.INTRANET', '1521', service_name='SVC_IDG01P')
    conn = cx_Oracle.connect(user='CUSTHIER', password='CarmineH20#2024', dsn=dsn_tns)

    ### this query is not used - at the moment. Originally it was intended to pull the MDM recommendations into this process.
    ### --> now we'll run the billing accounts through MDM to see if any suggested mappings pop up
    query_unused = """
SELECT
	'MDM MATCHING' AS MDM_MATCHING,
	ACCT.ACCOUNT_NBR AS BAN,
	ACCOUNT_NAME AS BAN_Name,
	ACCOUNT_SRC_SYS_CD AS BAN_Source,
	ACCOUNT_STATUS_CD AS BAN_Status,
	ACCOUNT_ACTVTY_STATUS_CD AS BAN_Activity_Status,
	acct.CUST_NBR AS current_cust_nbr,
	REL_TYP AS MDM_matching_status,
	CONFIDENCE_SCORE,
	RULE_NAME,
	RULE_HITS,
	X_CUST_NBR AS REC_CUST_NBR,
	RULE_COUNT,
	NORMALIZED_SCORE,
	ACCT.FINANCE_ACCOUNT_NBR AS FINANCE_ACCT_NBR,
	AUTOMATED_SCORE
FROM cmxorsp.c_xr_account_prty_rel rel
JOIN cmxorsp.c_xo_account acct ON rel.account_id = acct.rowid_object
JOIN cmxorsp.c_bo_prty prty ON rel.prty_id = prty.rowid_object
WHERE rel.rule_name IS NOT NULL
--AND COUNT(ACCT.FINANCE_ACCOUNT_NBR) = 1
-- AND finance_account_nbr = '2087336888196-CRISC'
AND CUST_NBR IN ('3-3RJVSP',
'3-TVXNBMKCGN',
'3-XPYJKZ',
'3-XHYWBJW7Q7',
'3-A85231',
'3-A85628',
'3-MLJCLL',
'3-ZFDGVZ63GW',
'3-A85231',
'3-ZXPFQ7',
'3-XYWRHR',
'3-C12FHZ',
'3-L456T6',
'3-TNHG1C',
'3-QMMYX2',
'3-MWQKDS',
'3-975JKD',
'3-J38XSJ',
'3-JK4HW4',
'3-TJCCFM',
'3-849928',
'3-GBLDYK',
'3-4QJKVCWQPM',
'3-2PVFVT')
and rel.result in ('Single Rec (No Prev Cust)', 'Single Rec = Curr Cust', 'Single Rec <> Curr Cus')
"""
    #also unused (at least as of 11/7/2023... well.. actually... well before then, but this is the time that I'm noting it)
    query_unused = f"""
        SELECT
            RULE_NAME,
            AUTOMATED_SCORE, 
            ACCT.FINANCE_ACCOUNT_NBR,
            REL."RESULT",
            PRTY.X_CUST_NBR AS "RECOMMEND_CUST"-- # RECOMMENDED CUST 
        FROM cmxorsp.c_xr_account_prty_rel rel
        JOIN cmxorsp.c_xo_account acct ON rel.account_id = acct.rowid_object
        JOIN cmxorsp.c_bo_prty prty ON rel.prty_id = prty.rowid_object
        WHERE 1=1
            AND RULE_NAME IS NOT NULL
            AND acct.finance_account_nbr IN ({fans_sql_string})
"""

    query_FAN = f"""
        SELECT
            FINANCE_ACCOUNT_NBR
            ,CUST_NBR AS MDM_CUST_NBR
            ,SECURE_COMPANY_NBR 
        FROM cmxorsp.c_xo_account
        WHERE
            FINANCE_ACCOUNT_NBR in ({data_string})
"""
    
    query_Cust = f"""
        SELECT
            X_CUST_NBR 
            ,X_STS_CD 
            ,X_SECURE_COMPANY_NBR 
        FROM CMXORSP.C_BO_PRTY CUST
        WHERE
            CUST.X_CUST_NBR in ({data_string})
    """

    query_CRIStoENS = f"""
        SELECT
            CONCAT(ACCOUNT_NUMBER, '-ENS') as ENS_FINANCE_ACCOUNT_NBR,
            CONCAT(LEGACY_IDENTIFIER, CONCAT('-', SOURCE_SYSTEM)) as CRIS_FINANCE_ACCOUNT_NBR
        FROM MDMRULE.LKP_BAN_XREF_ENS
        WHERE ACCOUNT_NUMBER in('333146816','334167449')
    """

    c = conn.cursor()
    
    if qNum == 1:
        #print("made it to FAN info in MDM query. Here is the FANs string", data_string)
        results = c.execute(query_FAN).fetchall()
    elif qNum == 2:
        # print("made it to CUST info in MDM query. Here is the Cust string", data_string)
        results = c.execute(query_Cust).fetchall()
    elif qNum == 3:
        # print("made it to CRIS to ENS mapping in MDM query. Here is the Cust string", data_string)
        results = c.execute(query_CRIStoENS).fetchall()
    c.close()

    numresults = 0
    for result in results:
        numresults = numresults + 1
    #print("this is the number of MDM results: ", numresults)
    return results



def queryOracle(num, query='', data_string=''):
    #print("in the oracle query here is the num", num, data_string)
    conn = dgc.connectCDW()

    if num == 1:
        print("made it to query 1: revenue. Here is the FANs string", data_string)
        ## -- query updated on 5/14/2024 to use FRIDA instead of UDL_FI - Vishwas helped to streamline it
        DONOTUSEquery = f"""
SELECT
	FA.FINANCE_ACCOUNT_NBR
	,FA.GL_PERIOD_YM
	,FA.SUM_USD
FROM
	(SELECT DISTINCT
		F.FINANCE_ACCOUNT_NBR
		,TO_CHAR(F.GL_PERIOD_START_DT,'YYYYMM') AS GL_PERIOD_YM
		,ROUND(SUM(F.USD_BUDGET_AMT), 0) AS SUM_USD
        --,ROUND(SUM(USD_AMT), 0) AS SUM_USD
	FROM UDL_FI.V_SS_DSV_F_FI_REVENUE_01_FMC F
    LEFT JOIN UDL_FI.SS_DSV_D_DATE_MASTER DM
        ON F.FI_DATE_MASTER_ID = DM.FI_DATE_MASTER_ID
        -- AND DM.SUB_TARGET_DATA IN ('ACTUAL', 'MJE') -- Manual Journal Entry / Actual(s) == actual finances or actual revenue 
    LEFT JOIN UDL_FI.SS_DSV_D_ACCOUNT_CO A
        ON F.GL_REVENUE_CATEGORY_CD = A.AMOUNT_TYPE_CD
	WHERE 
		TO_CHAR(F.GL_PERIOD_START_DT,'YYYYMM') 
		IN (
            TO_CHAR(ADD_MONTHS(SYSDATE,-1),'YYYYMM')
			,TO_CHAR(ADD_MONTHS(SYSDATE,-2),'YYYYMM')
			,TO_CHAR(ADD_MONTHS(SYSDATE,-3),'YYYYMM')
			,TO_CHAR(ADD_MONTHS(SYSDATE,-4),'YYYYMM')
			,TO_CHAR(ADD_MONTHS(SYSDATE,-5),'YYYYMM')
			,TO_CHAR(ADD_MONTHS(SYSDATE,-6),'YYYYMM')
			,TO_CHAR(ADD_MONTHS(SYSDATE,-7),'YYYYMM')
			,TO_CHAR(ADD_MONTHS(SYSDATE,-8),'YYYYMM')
			,TO_CHAR(ADD_MONTHS(SYSDATE,-9),'YYYYMM')
			,TO_CHAR(ADD_MONTHS(SYSDATE,-10),'YYYYMM')
			,TO_CHAR(ADD_MONTHS(SYSDATE,-11),'YYYYMM')
			,TO_CHAR(ADD_MONTHS(SYSDATE,-12),'YYYYMM')
			,TO_CHAR(ADD_MONTHS(SYSDATE,-13),'YYYYMM') )
		AND 
			F.FINANCE_ACCOUNT_NBR IN ({data_string})
        AND dm.SUB_TARGET_DATA IN ('ACTUAL','MJE') -- Manual Journal Entry / Actual(s) == actual finances or actual revenue 
		GROUP BY
			F.FINANCE_ACCOUNT_NBR
			,F.GL_PERIOD_START_DT
	) FA
	LEFT JOIN CODS.BILLING_ACCOUNT BA
		ON FA.FINANCE_ACCOUNT_NBR = BA.FINANCE_ACCOUNT_NBR 
	ORDER BY  FA.GL_PERIOD_YM DESC
"""
        ### --- use this query instead... you don't have a choice since it is the one that is named correctly
        query = f"""
                    SELECT
                        FINANCE_ACCOUNT_NBR
                        ,TO_CHAR(GL_PERIOD_START_DT,'YYYYMM') AS GL_PERIOD_YM
                        ,ROUND(SUM(USD_CURM_AMT),2) AS SUM_USD
                    FROM 
                        DSL_FINANCE.F_REVENUE_DETAIL_ALL FRDA
                        INNER JOIN CODS_FINANCE.GL_ACCOUNT GA 
                            ON FRDA.GL_ACCOUNT_ODS_ID = GA.GL_ACCOUNT_ODS_ID
                            AND GA.GL_ACCOUNT_TYP = 'Revenues'
                    WHERE 
                        GL_PERIOD_START_DT >= add_months(trunc(sysdate,'mm'),-13)
                        AND FRDA.journal_source_cd in ('BR','BA')                           
                        AND FINANCE_ACCOUNT_NBR IN ({data_string})
                    GROUP BY FRDA.FINANCE_ACCOUNT_NBR, FRDA.GL_PERIOD_START_DT
                """
        c = conn.cursor()
        #print("this is the query: ", query)
        results = c.execute(query).fetchall()
        # for result in results:
        #     print("this is the result: ", result)
        return(results)
    
    elif num == 2:
        print("Made it to query 2 - get current BAN age and lob info and current cust info")
        # print("If thereis no FAN in the request, then this will not return data - use the BAN if no FAN found?")
        #print("here is the string of FANs - to get current BAN age and LOB info: ", data_string)
        # ------> L.ASSIGNMENT_GROUP__C on 5/13/2024 to build out the new process
        query = f"""
SELECT  
    BA.FINANCE_ACCOUNT_NBR
    ,BA.BILL_ACCOUNT_ODS_ID
    ,CASE
        WHEN BA.BILL_ACCOUNT_NAME IS NULL THEN 'UNKNOWN - NO NAME ASSIGNED'
        ELSE BA.BILL_ACCOUNT_NAME
    END AS BILL_ACCOUNT_NAME
    ,BA.DW_SOURCE_SYSTEM_CD
    ,BA.DW_CREATE_DT
    ,CASE 
    	WHEN (to_date(current_date) - to_date(BA.DW_CREATE_DT)) <= 90 
    		THEN 'Under 90'
        ELSE 'Over 90' 
    END AS DATE_CHECK
    ,C.CUST_NBR
    ,C.CUST_ODS_ID
    ,CA.LOB_ID
    ,CASE 
    	WHEN (CA.EXTERNAL_RPT_SALES_CHAN_NAME IN ('INTERNAL')) 
    		THEN 'Internal'
        WHEN (CA.EXTERNAL_RPT_SALES_CHAN_NAME IN ('MASS MARKETS'))
        	THEN 'SBG' 
        WHEN (CA.LOB_ID IN ('563', '484', '485', '486', '487', '129')) 
        	THEN 'Federal'
        WHEN (CA.LOB_ID IN ('523')) 
        	THEN 'COLO'
        ELSE 'Other' 
    END AS LOB_GROUP
    ,CA.EXTERNAL_RPT_SALES_CHAN_NAME AS CRRNT_CUST_EXTRNL_RPRT_SALES_CHANNEL
    ,CA.SALES_CHAN_NAME AS CRRNT_CUST_SALES_CHAN
    ,CA.SALES_SUB_CHAN_NAME AS CRRNT_CUST_SALES_SUB_CHAN
    ,C.CUST_NAME
    ,C.DW_SECURE_COMPANY_NBR
    ,C.CUST_STATUS_CD
    ,CSA.HOUSE_ACCOUNT__C
    ,CSU.ID
    ,CSU.EMAIL
    ,CSU.NAME
    ,CSU.COMPANYNAME
    ,CSU.DEPARTMENT
    ,CSU.TITLE
    ,BA.BILL_ACCT_ACTVTY_STATUS_CD
    ,CASE 
    	WHEN ba.dw_source_system_cd like 'CRIS%' and substr(ba.bill_account_nbr,1,1) not in ('0','1','2','3','4','5','6','7','8','9') 
        	THEN substr(replace( replace(ba.bill_account_nbr ,'-',''),' ',''),2,13)
        /* built 1-18-2022 to deal with the CRIS billing accounts with suffixes in Bill_account_nbr like "605 367-7324 B224   443" */
        WHEN ba.dw_source_system_cd LIKE 'CRIS%' AND REGEXP_LIKE(BA.BILL_ACCOUNT_NBR , '.*[A-Z]\d{3}[ ]+\d{3}$')
        	THEN SUBSTR(BILL_ACCOUNT_SYSTEM_ID, 1, 13)
    	WHEN ba.dw_source_system_cd like 'CRIS%' 
    		THEN substr(replace( replace(ba.bill_account_nbr ,'-',''),' ',''),1,13)
    	ELSE replace( replace(ba.bill_account_nbr ,'-',''),' ','') 
    END AS BILLING_ACCOUNT_ID__C
    ,CASE
        WHEN ba.finance_account_nbr LIKE '%-LATIS'
            THEN 'LAT' || ba.BILL_ACCOUNT_NBR
        WHEN (ba.dw_source_system_cd = 'ORACLE2E')
            THEN
            CASE
                WHEN ba.finance_account_nbr LIKE '%-P-%'
                    THEN 'QIA' || bill_account_nbr || '-P'
                ELSE 'QIA' || bill_account_nbr
            END
        WHEN ba.dw_source_system_cd LIKE 'CRIS%' AND SUBSTR(ba.bill_account_nbr,1,1) NOT IN ('0','1','2','3','4','5','6','7','8','9') 
            THEN ba.dw_source_system_cd||SUBSTR(REPLACE(REPLACE(ba.bill_account_nbr ,'-',''),' ',''),2,13) 
        WHEN ba.dw_source_system_cd LIKE 'CRIS%' AND REGEXP_LIKE(ba.BILL_ACCOUNT_SYSTEM_ID , '\s[A-Z]\d{3}$')
	    THEN ba.dw_source_system_cd || SUBSTR(ba.bill_account_system_id, 0, 13)
        WHEN ba.dw_source_system_cd LIKE 'CRIS%' 
            THEN ba.dw_source_system_cd||SUBSTR(REPLACE( REPLACE(ba.bill_account_nbr ,'-',''),' ',''),1,13)
        ELSE ba.dw_source_system_cd || SUBSTR (bill_account_nbr, 1, 13)
    END AS BILLING_UNIQUE_EXTERNAL_ID__C
    ,CASE
        WHEN ba.finance_account_nbr LIKE '%-LATIS'THEN 'LAT' 
        ELSE ba.dw_source_system_cd
    END AS BILLING_SYSTEM__C
    ,BA.BILL_LINE1_ADDR
    ,BA.BILL_LINE2_ADDR
    ,BA.BILL_CITY_NAME
    ,BA.BILL_STATE_CD
    ,BA.BILL_POSTAL_CD
    ,BILL_COUNTRY_ISO_3_CD
    ,BA.DW_SECURE_COMPANY_NBR
    ,CASE
        WHEN BA.FINANCE_ACCOUNT_NBR  LIKE '%-ORACLE%'
            THEN CASE
                    WHEN BA.FINANCE_ACCOUNT_NBR LIKE '%-P-ORACLE%' THEN '012F0000000yFVmIAM'
                    ELSE '012F0000000yEHtIAM'
                END
        ELSE NULL 
    END AS RECORDTYPEID
    ,CASE
        WHEN ba.dw_source_system_cd = 'ORACLE2E' THEN 'QIA'
        ELSE NULL
    END AS BILLING_INSTANCE__C
    ,BA.BILL_ACCT_ALT_SYS_ID
    ,BA.BILL_ACCOUNT_LEVEL_TYP
    ,BA.BILL_CYCLE_CD
    ,BA.BILL_ACCT_ACTIVE_STATUS_IND
    ,L.ASSIGNMENT_GROUP__C
    ,CSA.TRADING_PARTNER__C
    ,BILL_ACCT_CUST_TYP             ----- added this cust_type and cust_sub_type 11/15/20224
    ,BILL_ACCT_CUST_SUB_TYP
    --,BA.BILL_ACCOUNT_STATUS_CD      ----- status cd and parent bill_account added 12/15/2024 for changes requsted by Emily on 12/12/2024 (ERP rules for BAN logic)
    --,BA.PARENT_BILL_ACCOUNT_NBR
FROM CODS.BILLING_ACCOUNT BA
    LEFT JOIN CODS.CUSTOMER C 
  	    ON c.CUST_ODS_ID = BA.CUST_ODS_ID 
    LEFT JOIN CODS.CUSTOMER_ATTRIBUTION CA 
	    ON C.CUST_ODS_ID = CA.CUST_ODS_ID
    LEFT JOIN CRPL_SALESFORCE_CTL.ACCOUNT CSA
        ON C.CUST_NBR = CSA.ACCOUNT_NUMBER__C
    LEFT JOIN CRPL_SALESFORCE_CTL."USER" CSU
        ON CSA.OWNERID = CSU.ID 
    LEFT JOIN CRPL_SALESFORCE_CTL.LOB__C L 
    	ON CA.LOB_ID = L.NAME 
WHERE 
    --SBA.DW_SOURCE_SYSTEM_CD <> 'ORAFIN'
    BA.FINANCE_ACCOUNT_NBR IN ({data_string})
    """
        c = conn.cursor()
        #print("this is the query: ", query)
        results = c.execute(query).fetchall()
        # for result in results:
        #     print("this is the result: ", result)
        return(results)


    elif num == 3:
        print("made it to query 3 - get requested customer info")
        #print("get requested customer info")
        query = f"""
SELECT
    C.CUST_NBR
    ,C.CUST_ODS_ID
    ,C.DUNS_NBR
    ,C.NAICS_CD
    ,C.CUST_LINE1_ADDR
    ,C.CUST_CITY_NAME
    ,C.CUST_STATE_CD
    ,C.CUST_POSTAL_CD
    ,A.COUNTRY_ISO_CODE__C -- # want this country code so we get USA instead of just US
    ,CA.LOB_ID
    ,CA.EXTERNAL_RPT_SALES_CHAN_NAME
    ,CA.SALES_CHAN_NAME
    ,CA.SALES_SUB_CHAN_NAME
    ,C.CUST_NAME
    ,C.DW_SECURE_COMPANY_NBR
    ,C.CUST_STATUS_CD
    ,CSA.HOUSE_ACCOUNT__C
    ,CSU.ID
    ,CSU.EMAIL
    ,CSU.NAME
    ,CSU.COMPANYNAME
    ,CSU.DEPARTMENT
    ,CSU.TITLE
    ,A.ID
    ,CASE WHEN (ca.EXTERNAL_RPT_SALES_CHAN_NAME IN ('INTERNAL')) THEN 'Internal - can move' ----- =================== CHECK FOR REMOVAL? OR CHANGE?
        WHEN (c.cust_status_cd = 'Inactive') THEN 'To BusOrg is not approved - cannot move'
        WHEN (ca.EXTERNAL_RPT_SALES_CHAN_NAME IN ('MASS MARKETS')) THEN 'SBG - Must match From BusOrg' 
        WHEN (ca.LOB_ID IN ('563', '484', '485', '486', '487', '129')) THEN 'Federal - can move'
        WHEN (ca.LOB_ID IN ('523')) THEN 'Colo - must match From BusOrg'
        WHEN (c.cust_status_cd = 'Inactive') THEN 'To BusOrg is not approved - cannot move'
    ELSE 'Other Enterprise' END AS LOB_GROUP
    ,L.ASSIGNMENT_GROUP__C --- [90]
    ,A.TRADING_PARTNER__C -- - [91]
FROM CODS.CUSTOMER c 
    LEFT JOIN CODS.CUSTOMER_ATTRIBUTION CA 
        ON C.CUST_ODS_ID = CA.CUST_ODS_ID 
    JOIN crpl_salesforce_ctl.ACCOUNT a 
	    ON C.CUST_NBR = A.ACCOUNT_NUMBER__C
    LEFT JOIN CRPL_SALESFORCE_CTL.ACCOUNT CSA
        ON C.CUST_NBR = CSA.ACCOUNT_NUMBER__C
    LEFT JOIN CRPL_SALESFORCE_CTL."USER" CSU
        ON CSA.OWNERID = CSU.ID
    LEFT JOIN CRPL_SALESFORCE_CTL.LOB__C L 
    	ON CA.LOB_ID = L.NAME 
WHERE C.CUST_NBR IN ({data_string})

"""

        c = conn.cursor()
        #print("this is the query: ", query)
        results = c.execute(query).fetchall()
        return(results)

    elif num == 4:
        print("made it to query 4 - requester emp info")
        query = f"""
SELECT
    E.EMPLOYEE_NBR
    ,E.EMP_JOB_FAMILY_NAME
    ,E.EMP_JOB_TITLE_NAME
    ,E.EMP_MANAGER_ID
    ,E_2.EMP_FULL_NAME
    ,E_2.EMP_JOB_TITLE_NAME
FROM
    CODS.EMPLOYEE E
    LEFT JOIN CODS.EMPLOYEE E_2
        ON E.EMP_MANAGER_ID = E_2.EMPLOYEE_NBR
WHERE
    E.EMPLOYEE_NBR IN ({data_string})
    """
        c = conn.cursor()
        results = c.execute(query).fetchall()
        return(results)


    elif query is not None and data_string is None:
        c = conn.cursor()
        results = c.execute(query).fetchall()
        return(results)


    elif num == 5:
        print("in query 5")
        query = f""" -- original query by Vishwas - changed to incorporate BA and the specific data I'm looking for (in subquery)
SELECT
	DISTINCT T1.FINANCE_ACCOUNT_NBR
	,T1.BILL_ACCOUNT_ODS_ID
	--,T1.BILL_CYCLE_CD
	,T1.SERV_INACTIVE_DT
FROM
	(
	SELECT --+ PARALLEL(16)
            BA.BILL_ACCOUNT_ODS_ID BILL_ACCOUNT_ODS_ID
            ,BA.FINANCE_ACCOUNT_NBR FINANCE_ACCOUNT_NBR
            ,BA.BILL_CYCLE_CD 
            ,SERV_INACTIVE_DT
	FROM CODS_BILLING.BILLING_PRODUCT_COMPNT BPC
	JOIN CODS.BILLING_ACCOUNT BA 
		ON BA.BILL_ACCOUNT_ODS_ID = BPC.BILL_ACCOUNT_ODS_ID 
	WHERE 1=1
		-- BA.BILL_ACCOUNT_ODS_ID IN ()
		AND BPC.DW_SOURCE_SYSTEM_CD = 'IDC KENAN'
		AND SERV_INACTIVE_DT IS NULL
	) T1 
WHERE
	T1.BILL_ACCOUNT_ODS_ID IN ({data_string})
    """
        c = conn.cursor()
        results = c.execute(query).fetchall()
        return(results)

    elif num == 6:
        print("in query 6")
        query = f"""SELECT 
                        UNIQUE cop.DW_SOURCE_SYSTEM_CD
                        ,ba.BILL_ACCOUNT_ODS_ID 
                        ,ba.FINANCE_ACCOUNT_NBR
                    FROM 
                        CODS.BILLING_ACCOUNT ba
                    LEFT JOIN CODS_BILLING.BILLING_PRODUCT_COMPNT bpc
                        ON ba.BILL_ACCOUNT_ODS_ID = bpc.BILL_ACCOUNT_ODS_ID
                    LEFT JOIN CODS.CUSTOMER_ORDER_PRODUCT cop
                        ON bpc.PRODUCT_INST_ID = cop.Product_Inst_Id
                    WHERE 
                        ba.BILL_ACCOUNT_ODS_ID IN ({data_string})
                        AND cop.DW_SOURCE_SYSTEM_CD IN 
                                (
                                    'SWIFT'
                                    ,'PIPELINE'
                                    ,'CLARIFY'
                                    ,'LUMEN_SERVICENOW' -- added 11/18/2024 per BAC team (CFS) after Tiger team discussion
                                    ,'BLUEMARBLE'
                                )
                """
        
        # print(query)
        c = conn.cursor()
        results = c.execute(query).fetchall()
        print("after query 6")
        return(results)
    
    elif num == 7:
        print("in query 7")

    c.close()
        


### ADD THIS IN 2/8/2023
""" -- built by Vishwas
SELECT --+ PARALLEL(16)
DISTINCT DW_SOURCE_SYSTEM_CD
FROM CODS_BILLING.BILLING_PRODUCT_COMPNT BPC
WHERE 1=1
AND BPC.DW_SOURCE_SYSTEM_CD = 'IDC KENAN'
AND SERV_INACTIVE_DT IS NULL
;

--BILL_ACCT_ACTIVE_STATUS_IND 
"""

#### ----> pull the data for each kenan account by ODS ID



def queryOracle_PLSQL():
    dsn_tns = cx_Oracle.makedsn(r'dbaracplb35-vip.twtelecom.com', '1521', service_name='ACCTMGTP1')
    conn = cx_Oracle.connect(user='AC79386', password='ORA#ctli_123', dsn=dsn_tns)
    query2 = f"""
SELECT
    AB.*,
    BA.ACCOUNT_NAME
FROM
    ACCTMGT_DBA.ALTERNATE_BILLING_ACCOUNT ab
    LEFT JOIN ACCTMGT_DBA.BILLING_ACCOUNT BA
         ON AB.BILLING_ACCOUNT_ID = BA.BILLING_ACCOUNT_ID
WHERE 
      AB.billing_account_source = 'Finance Account Number'
      and AB.ALTERNATE_BILLING_ACCOUNT_ID IN ('324532-TW')
"""

    c = conn.cursor()
    results = c.execute(query).fetchall()
    for row in results:
        print("\n\nhere is the plsql result " , row)
    return(results)



def getFANsString(stored_fans):
    fans_string = []
    fans_string = str([str(x) for x in stored_fans]).strip("[]") # peel off the brckets so it is stored in a format SQL can parse
    #print("\n\nHere is the converted tuple to string (inside getFANsString()): \n", fans_string)
    return fans_string



def getMDMdata(ssms_results):
    global fans_sql_string
    global to_cust_sql_string
    global stored_customers
    global stored_fans


    global mdm_fan_results
    global mdm_cust_results

    global mdm_FANs_list
    global mdm_Cust_list

    global mdm_fan_results_list
    global mdm_cust_results_list

    # mdm_fan_results = []
    # mdm_cust_results = []
    num_1k_blocs = 0

    print(len(ssms_results))

    #cust_nbr = result[11]
    for result in ssms_results:
        stored_fans.append(result[14])
        stored_customers.append(result[11])

    ### --> NEED TO ADDRESS THIS: MAKE SURE IT WORKS FOR NUMBERS LESS THAN 1K
    if len(ssms_results) > 999:
        num_1k_blocs = len(ssms_results) / 999
        print("this is the total over 1k: " ,num_1k_blocs)
        #print("and here is the length: " , len(ssms_results))

    #test_num = 513
    num_1k_blocs = len(ssms_results) // 999
    remainder = len(ssms_results) % 999
    max_number_divisible = len(ssms_results) - remainder
    #print(remainder)
    #print(num_1k_blocs)
    #print("this is max number: ", max_number_divisible)
    start_chunk = 0
    num = 0
    rev_query_results = []


    if num_1k_blocs:
        while num < num_1k_blocs:
        ##for chunk in range(num_1k_blocs):
            #print("this is the start_chunk: ", start_chunk)
            end_chunk = start_chunk + 999
            ranged_save_FANs = []
            ranged_save_Customers = []
            for number in range(start_chunk, end_chunk):
                ranged_save_FANs.append(ssms_results[number][14])
                ranged_save_Customers.append(ssms_results[number][11])
                #print(start_chunk, end_chunk)
                #print(ranged_save_FANs)
                #print("this is the number: ", number)
            fans_string = getFANsString(ranged_save_FANs)
            fans_sql_string = fans_sql_string + fans_string
            ##print(query_3)

            cust_string = getFANsString(ranged_save_Customers)
            to_cust_sql_string = to_cust_sql_string + cust_string

            
            mdm_fan_results.append(queryMDM(1, fans_string))
            mdm_cust_results.append(queryMDM(2, cust_string))
            start_chunk = start_chunk + 999
            #print("this is the new start_chunk: ", start_chunk)
            if start_chunk == max_number_divisible:
                ranged_save_FANs = []
                ranged_save_Customers = []
                end_chunk = start_chunk + remainder

                for number in range(start_chunk, end_chunk):
                    ranged_save_FANs.append(ssms_results[number][14])
                    ranged_save_Customers.append(ssms_results[number][11])
                fans_string = getFANsString(ranged_save_FANs)
                fans_sql_string = fans_sql_string + fans_string # append fans to the fans_sql_string so we can use that global variable later

                cust_string = getFANsString(ranged_save_Customers)
                to_cust_sql_string = to_cust_sql_string + cust_string

                mdm_fan_results.append(queryMDM(1, fans_string))
                mdm_cust_results.append(queryMDM(2, cust_string))
            num += 1
            #print("this is the num: ", num)
            #print("chunk: ", start_chunk)

    else: #else, if there are less than 1k fans to research
        ranged_save_FANs = []
        ranged_save_Customers = []
        for entry in ssms_results:
            ranged_save_FANs.append(entry[14])
            ranged_save_Customers.append(entry[11])
        #fans_sql_string here to be sure that we have identified and updated the Global variable
        fans_string = getFANsString(ranged_save_FANs)
        cust_string = getFANsString(ranged_save_Customers)
        #print("here is the fans sql string", fans_sql_string)
        mdm_fan_results.append(queryMDM(1, fans_string))
        mdm_cust_results.append(queryMDM(2, cust_string))

        print(mdm_fan_results)

    mdm_fan_results_list = [list(found_fan) for item in mdm_fan_results for found_fan in item]
    mdm_cust_results_list = [list(found_cust) for item in mdm_cust_results for found_cust in item]

    #print("\n\n Nebkidnezzer: found the mdm fan results: ", mdm_fan_results_list)
    #print("\n\n Faffilwight: found the mdm cust results: ", mdm_cust_results_list)

    for item in mdm_fan_results_list:
        # print("\n\Madmadigan: found the mdm fan in the fan results: ", item[0], item)
        mdm_FANs_list.append(item[0])
        mdm_FANs_dict[item[0]] = [item[1]]
        # print("\n\n HERE's the mdm FANs dict", mdm_FANs_dict)
        mdm_FANs_dict[item[0]].append(item[2])
    for item in mdm_cust_results_list:
        # print("\n\nBankofamerica: found the mdm cust in the cust results: ", item[0], item)
        mdm_Cust_list.append(item[0])
        mdm_Cust_dict[item[0]] = [item[1]]
        mdm_Cust_dict[item[0]].append(item[2])



#going to take one list of the same data and split it into a dictionary of lists that are numbered
# so, {1, 2, 3, 4, 5, 6, 7, 8, 9, 10} becomes {1: [1, 2, 3], 2: [4, 5, 6], 3: [7, 8, 9], 4: [10]} essentially (but the nubmers will go to 999)
def get1kBlocksFromList(list, divider):
    num_1k_blocks = len(list) // divider
    remainder = len(list) % divider
    max_number_divisible = len(list) - remainder
    
    # print("heres the complete list: ", list)    
    # print("len of list: ", len(list))
    # print("remainder after dividing by divider: ", remainder)
    # print("total number of 1k blocks to work with: ", num_1k_blocks)
    # print("this is max number: ", max_number_divisible)

    start_chunk = 0
    num = 0
    ranged_save_data =  {}
    
    # if there actually is an amount of fans over 1k
    if num_1k_blocks:
        while num < num_1k_blocks:
            end_chunk = start_chunk + divider 
            ranged_save_data[num] = list[start_chunk:end_chunk]

            num += 1
            start_chunk = end_chunk

            if start_chunk == max_number_divisible:
                ranged_save_data[num] = list[start_chunk:]
    else:
        ranged_save_data[0] = list

    # print("all the data for chunked data: ", ranged_save_data)
    return ranged_save_data



def getBanRevenue(ssms_results):
    last_3_months_date_list = []
    global fans_sql_string
    global bans_with_rev_list
    global stored_last_3_months
    global stored_last_6_months
    global stored_last_9_months
    global stored_last_12_months
    global stored_last_13_months
    global stored_this_month
    global stored_months
    global stored_rev_query_results
    global to_cust_sql_string

    global stored_customers
    global stored_fans

    fans_string = []

    #print("\n\n\n here is the ssms_results:", ssms_results)

    '''
    cust_nbr = result[11]
    for result in ssms_results:
        this_fan = result[14]
        stored_fans.append(result[14])
        stored_customers.append(cust_nbr)
    '''
              
    query_3 = f"""SELECT
                    FINANCE_ACCOUNT_NBR
                    ,TO_CHAR(GL_PERIOD_START_DT,'YYYYMM') AS GL_PERIOD_YM
                    ,ROUND(SUM(USD_CURM_AMT),2) AS SUM_USD
                FROM 
                    DSL_FINANCE.F_REVENUE_DETAIL_ALL FRDA
                    INNER JOIN CODS_FINANCE.GL_ACCOUNT GA ON FRDA.GL_ACCOUNT_ODS_ID = GA.GL_ACCOUNT_ODS_ID
                        AND GA.GL_ACCOUNT_TYP = 'Revenues'
                WHERE 1=1
                    AND GL_PERIOD_START_DT >= add_months(trunc(sysdate,'mm'),-13)
                    AND FRDA.journal_source_cd in ('BR','BA')                           
                    AND FINANCE_ACCOUNT_NBR IN ('5-BGJXL13J')
                GROUP BY FRDA.FINANCE_ACCOUNT_NBR, FRDA.GL_PERIOD_START_DT
                    order by  FRDA.FINANCE_ACCOUNT_NBR, FRDA.GL_PERIOD_START_DT desc
"""
    
    query_results_storage = []
    num_1k_blocs = 0

    #print(len(ssms_results))

    ### --> NEED TO ADDRESS THIS: MAKE SURE IT WORKS FOR NUMBERS LESS THAN 1K
    if len(ssms_results) > 999:
        num_1k_blocs = len(ssms_results) / 999
        print("this is the total over 1k: " ,num_1k_blocs)
        #print("and here is the length: " , len(ssms_results))

    #test_num = 513
    num_1k_blocs = len(ssms_results) // 999
    remainder = len(ssms_results) % 999
    max_number_divisible = len(ssms_results) - remainder
    #print(remainder)
    #print(num_1k_blocs)
    #print("this is max number: ", max_number_divisible)
    start_chunk = 0
    num = 0
    rev_query_results = []


    if num_1k_blocs:
        while num < num_1k_blocs:
        ##for chunk in range(num_1k_blocs):
            #print("this is the start_chunk: ", start_chunk)
            end_chunk = start_chunk + 999
            ranged_save_FANs = []
            ranged_save_Customers = []
            for number in range(start_chunk, end_chunk):
                ranged_save_FANs.append(ssms_results[number][14])
                ranged_save_Customers.append(ssms_results[number][11])
                #print(start_chunk, end_chunk)
                #print(ranged_save_FANs)
                #print("this is the number: ", number)
            fans_string = getFANsString(ranged_save_FANs)
            fans_sql_string = fans_sql_string + fans_string
            ##print(query_3)

            cust_string = getFANsString(ranged_save_Customers)
            to_cust_sql_string = to_cust_sql_string + cust_string

            
            rev_query_results.append(queryOracle(1, query_3, fans_string))
            start_chunk = start_chunk + 999
            #print("this is the new start_chunk: ", start_chunk)
            if start_chunk == max_number_divisible:
                ranged_save_FANs = []
                end_chunk = start_chunk + remainder
                for number in range(start_chunk, end_chunk):
                    ranged_save_FANs.append(ssms_results[number][14])
                fans_string = getFANsString(ranged_save_FANs)
                fans_sql_string = fans_sql_string + fans_string # append fans to the fans_sql_string so we can use that global variable later
                rev_query_results.append(queryOracle(1, query_3, fans_string))
            num += 1
            #print("this is the num: ", num)
            #print("chunk: ", start_chunk)

    else: #else, if there are less than 1k fans to research
        ranged_save_FANs = []
        for entry in ssms_results:
            ranged_save_FANs.append(entry[14])
        #fans_sql_string here to be sure that we have identified and updated the Global variable
        fans_sql_string = getFANsString(ranged_save_FANs)
        #print("here is the fans sql string", fans_sql_string)
        rev_query_results.append(queryOracle(1, query_3, fans_sql_string))
        
        
    

    rev_query_holder = []
    for item in rev_query_results:
        for thingy in item:
            rev_query_holder.append(list(thingy))
            #print(thingy)
            #print(rev_query_holder)

    # rev_query_results_list = [list(item) for item in rev_query_results]
    ssms_results_list = [list(item) for item in ssms_results] ## turn tuple into list
    
    #print("\n\nhere is the ssms results list to list: \n", ssms_results_list)
    # --> get a list of FANs back prepped for SQL entry
    if len(rev_query_results) > 0:
        #print("here are the results in BAN Revenue query: ", rev_query_results[0][0])
        
        #print("rev results list exists and is greater than 0")
        for rev_result in rev_query_holder:
            #print(rev_result)
            #print(rev_result)
            #print("this should be a FAN: ", rev_result[0])
            revenue_results_container = []
            
            revenue_results_container.append(rev_result[0])
            revenue_results_container.append(rev_result[1])
            revenue_results_container.append(rev_result[2])

            stored_rev_query_results.append(revenue_results_container)
            
            if rev_result[0] not in bans_with_rev_list:
                bans_with_rev_list.append(rev_result[0])

        for ssms_result in ssms_results_list:
            if ssms_result[14] in bans_with_rev_list:
                #print("here is the ssms_result from BAN rev func: ", ssms_result[14])
                ssms_result.append('Yes')
            else:
                ssms_result.append('No')

                
        stored_FAN = ''
        finished_rev_FAN = []
        fallout_list = [0,0,0,0,0,0]
        x = 0

        #print("\n\nhere comes the BOOOOOOM")
        #print("reviewing the revenue")
        while x < len(stored_rev_query_results):
            revenue_container = []
            this_months_revenue = 0
            _3_months_revenue = 0 
            _6_months_revenue = 0
            _9_months_revenue = 0
            _12_months_revenue = 0
            _13_months_revenue = 0
            stored_FAN = stored_rev_query_results[x][0]
            for entry in stored_rev_query_results:
                if entry[0] == stored_FAN and entry[1] in stored_this_month and entry[0] not in finished_rev_FAN:
                    #print("\n\n\nSOOOOO, here is the date: ", entry[1], "\nand here is the amount: ", entry[2])
                    this_months_revenue = float(entry[2])
                if entry[0] == stored_FAN and entry[1] in stored_last_3_months and entry[0] not in finished_rev_FAN:
                    #print("\n\n\nSOOOOO, here is the date: ", entry[1], "\nand here is the amount: ", entry[2])
                    _3_months_revenue += float(entry[2])
                if entry[0] == stored_FAN and entry[1] in stored_last_6_months and entry[0] not in finished_rev_FAN:
                    #print("\n\n\nSOOOOO, here is the date: ", entry[1], "\nand here is the amount: ", entry[2])
                    _6_months_revenue += float(entry[2])
                if entry[0] == stored_FAN and entry[1] in stored_last_9_months and entry[0] not in finished_rev_FAN:
                    #print("\n\n\nSOOOOO, here is the date: ", entry[1], "\nand here is the amount: ", entry[2])
                    _9_months_revenue += float(entry[2])
                if entry[0] == stored_FAN and entry[1] in stored_last_12_months and entry[0] not in finished_rev_FAN:
                    #print("\n\n\nSOOOOO, here is the date: ", entry[1], "\nand here is the amount: ", entry[2])
                    _12_months_revenue += float(entry[2])
                if entry[0] == stored_FAN and entry[1] in stored_last_13_months and entry[0] not in finished_rev_FAN:
                    #print("\n\n\nSOOOOO, here is the date: ", entry[1], "\nand here is the amount: ", entry[2])
                    _13_months_revenue += float(entry[2])

            # Averages
            _3_months_revenue = _3_months_revenue / 3 if _3_months_revenue != 0 else 0
            _6_months_revenue = _6_months_revenue / 6 if _6_months_revenue != 0 else 0
            _9_months_revenue = _9_months_revenue / 9 if _9_months_revenue != 0 else 0
            _12_months_revenue = _12_months_revenue / 12 if _12_months_revenue != 0 else 0
            _13_months_revenue = _13_months_revenue / 13 if _13_months_revenue != 0 else 0

            #print(f"\n\n\nand now here is the averaged revenue for {stored_FAN}: ",
                  #_3_months_revenue, _6_months_revenue, _9_months_revenue, _12_months_revenue, _13_months_revenue)
            revenue_container.append(this_months_revenue)
            revenue_container.append(_3_months_revenue)
            revenue_container.append(_6_months_revenue)
            revenue_container.append(_9_months_revenue)
            revenue_container.append(_12_months_revenue)
            revenue_container.append(_13_months_revenue)

            #print("\n\nhere is the container of revenue: ", revenue_container)

                    
            for ssms_result in ssms_results_list:
                #print("\n\nhere is the ssms_result: ", ssms_result)
                if stored_FAN == ssms_result[14] and stored_FAN not in finished_rev_FAN:
                    #print("\n\nhere is the fan from ssms_result: ", ssms_result[14])
                    for revenue in revenue_container:
                        #print("\n\nhere is the revenue from revenue_container: ", revenue)
                        ssms_result.append(revenue)
                        #print("\n\nand here is the appended result from ssms results: ", ssms_result)
            finished_rev_FAN.append(stored_FAN)
            
            x+=1
            
        #print(stored_rev_query_results)
        #print(bans_with_rev_list)
            
        #print("\n\nHere are the SSMS results list: \n", ssms_results_list)
        for ssms_result in ssms_results_list:
            if ssms_result[14] not in finished_rev_FAN:
                #print("not in finished_rev")
                for x in fallout_list:
                    ssms_result.append(x)
        """if os.path.isfile(testing_csv_path):
            with open(testing_csv_path,'a+', newline='') as out: ## w+ to read and write (just in case) a+ to append (read and write)
                csv_out=csv.writer(out)
                ##for result in results:
                for result in ssms_results_list:
                    csv_out.writerow(result)
                print("in the with loop to print ERRORS to csv")
        else:
            with open(testing_csv_path,'a+', newline='') as out: ## w+ to read and write (just in case) a+ to append (read and write)
                csv_out=csv.writer(out)
                ##for result in results:
                for result in ssms_results_list:
                    csv_out.writerow(result)
                print("in the with loop to print ERRORS to csv")"""
        return ssms_results_list
    else:
        for result in ssms_results_list:
            iter_num = 0
            while iter_num < 6:
                result.append('0')
                #print("\n\n HERE IS THE RESULT: ", result)
                iter_num = iter_num + 1
            #print("\n\n HERE IS THE RESULT: ", result)
        return ssms_results_list



def getBanAgeAndLOB(revenue_results):
    global fans_sql_string
    # for result in revenue_results:  
    #     print("\n\nhere are rev results: ", result)


    #print(query_4)
    query_results_storage = []
    num_1k_blocs = 0

    #print(len(revenue_results))

    ### --> NEED TO ADDRESS THIS: MAKE SURE IT WORKS FOR NUMBERS LESS THAN 1K
    if len(revenue_results) > 999:
        num_1k_blocs = len(revenue_results) / 999
        #print("this is the total over 1k: " ,num_1k_blocs)
        #print("and here is the length: " , len(revenue_results))

    #test_num = 513
    num_1k_blocs = len(revenue_results) // 999
    remainder = len(revenue_results) % 999
    max_number_divisible = len(revenue_results) - remainder
    #print(remainder)
    #print(num_1k_blocs)
    #print("this is max number: ", max_number_divisible)
    start_chunk = 0
    num = 0
    ban_age_and_LOB_results = []


    if num_1k_blocs:
        while num < num_1k_blocs:
            ##for chunk in range(num_1k_blocs):
            #print("this is the start_chunk: ", start_chunk)
            end_chunk = start_chunk + 999
            ranged_save_FANs = []
            for number in range(start_chunk, end_chunk):
                ranged_save_FANs.append(revenue_results[number][14])
                #print(start_chunk, end_chunk)
                #print(ranged_save_FANs)
                #print("this is the number: ", number)
            fans_string = getFANsString(ranged_save_FANs)
            ##print(query_3)
            ban_age_and_LOB_results.append(queryOracle(2, '', fans_string))
            start_chunk = start_chunk + 999
            #print("this is the new start_chunk: ", start_chunk)
            if start_chunk == max_number_divisible:
                ranged_save_FANs = []
                end_chunk = start_chunk + remainder
                for number in range(start_chunk, end_chunk):
                    ranged_save_FANs.append(revenue_results[number][14])
                fans_string = getFANsString(ranged_save_FANs)
                ban_age_and_LOB_results.append(queryOracle(2, '', fans_string))
            num += 1
            #print("this is the num: ", num)
            #print("chunk: ", start_chunk)
        
            #print("here are the results in BAN Revenue query: ", ban_age_and_LOB_results[0][0])
    else:
        ban_age_and_LOB_results.append(queryOracle(2, '', fans_sql_string))



    ban_age_and_LOB_query_holder = []
    for item in ban_age_and_LOB_results:
        for thingy in item:
            ban_age_and_LOB_query_holder.append(list(thingy))
            #print("\nhere's the thingy:", thingy)
            #print("\nhere's the item:", item)
    #print("this is the length of sql string", len(fans_sql_string))

    #BANlob_query_results = queryOracle(2, '', fans_sql_string)
    for item in ban_age_and_LOB_query_holder:
        if item[0] == 'PPB03647678-PPP':
            print('found that fuckin BAN: ', item)
    BANlob_query_results_list = [list(item) for item in ban_age_and_LOB_results]

    #print("here is the BAN age and lob result: ", ban_age_and_LOB_results)
    #print("and here is the list: ", ban_age_and_LOB_results)

    if len(ban_age_and_LOB_results) > 0:
        for ban_result in ban_age_and_LOB_query_holder:
            stored_ban_lob_n_age_query_results.append(ban_result)
            # print("here is the ban resutl: ", ban_result)
            for rev_result in revenue_results:
                # print("here are the rev results: ", rev_result)
                # print("here is the length of rev_result: ", len(rev_result))
                if rev_result[14] == ban_result[0] and len(rev_result) == 24: ### had to update this amount since I added data to the first query !!!! IMPORTANT!!!!
                    # print("here is a match: ", rev_result[14], ban_result[0])
                    i = 0
                    while i < len(ban_result):
                        rev_result.append(ban_result[i])
                        i+=1
                    # print("here is the new rev_result row: " , rev_result)
                    #item.append(str(str(x) for x in ban_result))
                    #print("\nban result: ", ban_result[0])
                    #print("\nreve result: ", item[14])
                else:
                    continue
        #print("here is the changed revenue results: " ,revenue_results)
        """
        if os.path.isfile(testing_csv_path):
            with open(testing_csv_path,'a+', newline='') as out: ## w+ to read and write (just in case) a+ to append (read and write)
                csv_out=csv.writer(out)
                ##for result in results:
                for result in revenue_results:
                    csv_out.writerow(result)
                print("in the with loop to print ERRORS to csv")
        else:
            with open(testing_csv_path,'a+', newline='') as out: ## w+ to read and write (just in case) a+ to append (read and write)
                csv_out=csv.writer(out)
                ##for result in results:
                for result in revenue_results:
                    csv_out.writerow(result)
                print("in the with loop to print ERRORS to csv")"""
        return revenue_results
    else:
        return revenue_results
        



def getRevFANsFromSSMSList(revenue_fans, ssms_list):
    global can_map_list
    global cannot_map_list

    for row in ssms_list:
        fan = row[14]
        if fan in revenue_fans:
            can_map_list.append(row)
            #print("this is the can_map_list", can_map_list)
        elif fan not in revenue_fans:
            cannot_map_list.append(row)
    #print("can map 'em", "\n\n", can_map_list)
    #print("cannot map 'em", "\n\n", cannot_map_list)
            



def getToBusOrgLOBInfo(ban_and_ticket_info):
    print("Made it to the getToBusOrgLOBInfo function")
    #for row in ban_and_ticket_info:
        #print("\n\nhere is the ban and ticket info row: ", row)
    to_customer_list = []
    existing_cust_list = []
    
    for cust in ban_and_ticket_info:
        # print("this is the row length: ", len(cust))
        rqstd_cust = cust[11]
        if rqstd_cust not in existing_cust_list:
            existing_cust_list.append(rqstd_cust)
            to_customer_list.append(rqstd_cust)
            
    to_customer_list_string = str([str(x) for x in to_customer_list]).strip("[]")

    #print("\n\nhere is the to_customer_list: ", to_customer_list)

    query_results_storage = []
    num_1k_blocs = 0

    ### --> NEED TO ADDRESS THIS: MAKE SURE IT WORKS FOR NUMBERS LESS THAN 1K
    if len(to_customer_list) > 999:
        num_1k_blocs = len(to_customer_list) / 999
        #print("this is the total over 1k custs: " ,num_1k_blocs)
        #print("and here is the length of that list of custs: " , len(to_customer_list))

    #test_num = 513
    num_1k_blocs = len(to_customer_list) // 999
    remainder = len(to_customer_list) % 999
    max_number_divisible = len(to_customer_list) - remainder
    #print(remainder)
    #print(num_1k_blocs)
    #print("this is max number of custs: ", max_number_divisible)
    start_chunk = 0
    num = 0
    toBusOrg_query_results = []


    if num_1k_blocs:
        while num < num_1k_blocs:
        ##for chunk in range(num_1k_blocs):
            #print("this is the start_chunk: ", start_chunk)
            end_chunk = start_chunk + 999
            ranged_save_CUSTs = []
            #print("here is the start:", start_chunk, "and here is the end: " ,end_chunk)
            for number in range(start_chunk, end_chunk):
                ranged_save_CUSTs.append(to_customer_list[number])
                if to_customer_list[number] == '3-A76052':
                    print(f"i have found the customer: " , to_customer_list[number], "at number {number}")
                #print(ranged_save_CUSTs)
                #print("this is the number: ", number)
            cust_sql_string = getFANsString(ranged_save_CUSTs)
            #print("here is the cust sql string: ", cust_sql_string)
            ##print(query_3)
            toBusOrg_query_results.append(queryOracle(3, '', cust_sql_string))

            #print("and the tobusorg query results", toBusOrg_query_results)
            start_chunk = start_chunk + 999
            #print("here is the new start:", start_chunk)
            #print("this is the new start_chunk: ", start_chunk)
            if start_chunk == max_number_divisible:
                ranged_save_CUSTs = []
                end_chunk = start_chunk + remainder
                for number in range(start_chunk, end_chunk):
                    ranged_save_CUSTs.append(to_customer_list[number])
                cust_sql_string = getFANsString(ranged_save_CUSTs)
                toBusOrg_query_results.append(queryOracle(3, '', cust_sql_string))
            print(num)            
            num += 1

        '''
        i = 0
        x = 0
        while i < 1:
            while x < len(toBusOrg_query_results[i]):
                #print("here is the item" , toBusOrg_query_results[i][x])
                if toBusOrg_query_results[i][x][0] == '3-A76052': ## item[0] == to-customer
                    print(f'Yes, I found that busorg! Number {x}: ', toBusOrg_query_results[i][x])
                    x+=1
                else:
                    x+=1
            i+=1
        '''
            #print("this is the num: ", num)
            #print("chunk: ", start_chunk)
        
    else:
        toBusOrg_query_results.append(queryOracle(3, '', to_customer_list_string))
    #print("here are the results of the ToBusOrg query: ", toBusOrg_query_results)

    toBusOrg_query_holder = []
    i = 0
    for item in toBusOrg_query_results:
        for thingy in item:
            toBusOrg_query_holder.append(list(thingy))
            if thingy[0] == '3-A76052':
                print(f"I found the Busorg! and the number is {i} " , thingy)
            i += 1
            #print(rev_query_holder)
    i = 0
    x = 0
    while i < 1:
        while x < len(toBusOrg_query_results[i]):
            #print("here is the item" , toBusOrg_query_results[i][x])
            if toBusOrg_query_results[i][x][0] == '3-A76052': ## item[0] == to-customer
                #print('Yes, I found that motherfucker: ', toBusOrg_query_results[i][x])
                x+=1
            else:
                #print('Nope, did not find it this time')
                x+=1
        i+=1
    # rev_query_results_list = [list(item) for item in rev_query_results]
    # ssms_results_list = [list(item) for item in ssms_results] ## turn tuple into list

    #print("\nhere is the customer list: \n\n", to_customer_list_string)
    #for item in ban_and_ticket_info:
       # print("Here's the item: ", item)

    #print("here is the query holder", toBusOrg_query_holder)
    # DW_SECURE_COMPANY_NBR --> 1 = Lumen, 2 = LATAM, 4 = Brightspeed
    
    # ----> busorg_query_results = queryOracle(3, '', to_customer_list_string)
    # busorg_query_results_list = str([str(x) for x in busorg_query_results]).strip("[]")
    toBusOrg_query_holder_list = str([str(x) for x in toBusOrg_query_holder]).strip("[]")

    #print("\n\nHere is the query result for To-BusOrg info: \n", toBusOrg_query_holder)

    for item in toBusOrg_query_holder:
        if item[0] == '3-A76052': ## item[0] == to-customer
            print('found that motherfucker: ', item)
        else:
            continue
    
    if len(toBusOrg_query_holder) > 0:
        #print("wooot!!!!!!!!!!!!!!!!!!!!!!")
        for cust_result in toBusOrg_query_holder:
            stored_to_cust_lob_info_query_results.append(cust_result)
            for item in ban_and_ticket_info:
                if item[11] == cust_result[0]:
                    #print("\n\n\nhere is the cust_result", cust_result)
                    #print(item[11])
                    #print(cust_result[0])
                    i = 0
                    while i < len(cust_result):
                        item.append(cust_result[i])
                        #print("the printed cust result:", cust_result[i])
                        i+=1
                    #item.append(str(str(x) for x in ban_result))
                    #print("\nto-cust result: ", cust_result[0])
                    #if item[14] == '300890983-ENS':
                        #print("\n\n\nfull sheet result: ", item)
        #print("\n\nHere is BAN and ticket info: \n", ban_and_ticket_info)
        if os.path.isfile(testing_csv_path):
            with open(testing_csv_path,'a+', newline='') as out: ## w+ to read and write (just in case) a+ to append (read and write)
                csv_out=csv.writer(out)
                ##for result in results:
                for cust in toBusOrg_query_results:
                    csv_out.writerow(cust)
                #print("in the with loop to print testing to csv")
        else:
            with open(testing_csv_path,'a+', newline='') as out: ## w+ to read and write (just in case) a+ to append (read and write)
                csv_out=csv.writer(out)
                ##for result in results:
                for cust in toBusOrg_query_results:
                    csv_out.writerow(cust)
                #print("in the with loop to print ERRORS to csv")
        #print("\n\nban and ticket info", ban_and_ticket_info)
        #for item in ban_and_ticket_info:
            #print("\n\n\n this is the length of the item in ban_and_ticket_info:\n", len(item),  item)
        return ban_and_ticket_info
    else:
        return ban_and_ticket_info



def compareLOBs(data_list):
    # this function is called before the evaluation function
    # so any values in here are the same as they should show in the Complete file
    # in other words, the 'SUGGESTION' field is unpopulated then, so the list is still zero indexed
    for row in data_list:
        # print("\n\nprinting row inside LOB compare:" , row)
        # print(len(row))
        #LOB compare
        if len(row) < 77:
            print("\n\nthis row is missing data: ", row)
        fan = row[14]
        crrnt_cust_lob, rqstd_cust_lob = row[32], row[77]
        crrnt_cust_BU_or_ExternalSalesChannel, rqstd_cust_BU_or_ExternalSalesChannel = row[34], row[78]
        crrnt_cust_SalesChannel, rqstd_cust_SalesChannel = row[35], row[79]
        # print("\n\nfan: ", fan, "\ncrrnt_cust_lob: ", crrnt_cust_SalesChannel, "        rqstd_cust_lob: ", rqstd_cust_SalesChannel, "\n")
        # print("crrnt_cust_SalesChannel: ", crrnt_cust_SalesChannel, "        rqstd_cust_SalesChannel: ", rqstd_cust_SalesChannel, "\n")
        # print("crrnt_cust_BU_or_ExternalSalesChannel: ", crrnt_cust_BU_or_ExternalSalesChannel, "        rqstd_cust_BU_or_ExternalSalesChannel: ", rqstd_cust_BU_or_ExternalSalesChannel, "\n\n")

        if crrnt_cust_lob == rqstd_cust_lob:
            row.append('N')
        elif crrnt_cust_lob != rqstd_cust_lob:
            row.append('Y')
        else:
            row.append('?')

        ## Sales Channel compare
        if crrnt_cust_SalesChannel == rqstd_cust_SalesChannel:
            row.append('N')
        elif crrnt_cust_SalesChannel != rqstd_cust_SalesChannel:
            row.append('Y')
        else:
            row.append('?')

        
        ##BU Compare
        if crrnt_cust_BU_or_ExternalSalesChannel == rqstd_cust_BU_or_ExternalSalesChannel:
            row.append('N')
        elif crrnt_cust_BU_or_ExternalSalesChannel != rqstd_cust_BU_or_ExternalSalesChannel:
            row.append('Y')
        else:
            row.append('?')

    return data_list



def compareDWSecure(data_list):
    # compare dw secure BAN to customer    

    for row in data_list:
        ban_dw_secure, rqstd_cust_dw_secure = row[57], row[82]

        if ban_dw_secure == rqstd_cust_dw_secure:
            row.append('Y')
        else:
            row.append('N')
    return data_list

        

def getRequesterInfo(ban_ticket_andToCustomer_info):
    requesting_emp_list = []
    existing_emp_list = []
    global stored_requesting_employee_info_query_results

    for emp in ban_ticket_andToCustomer_info:
        employee_number = emp[1]
        #print("\n\nhere is the emp #: " , employee_number)
        if employee_number not in existing_emp_list:
            existing_emp_list.append(employee_number)
            requesting_emp_list.append(employee_number)
            
    requesting_emp_list_string = str([str(x) for x in requesting_emp_list]).strip("[]")
    #print("\n\nhere's the requesting employee string: ", requesting_emp_list_string)

    #print("\n\nhere is the to_customer_list: ", to_customer_list)

    query_results_storage = []
    num_1k_blocs = 0

    ### --> NEED TO ADDRESS THIS: MAKE SURE IT WORKS FOR NUMBERS LESS THAN 1K
    if len(requesting_emp_list) > 999:
        num_1k_blocs = len(requesting_emp_list) / 999
        #print("this is the total over 1k custs: " ,num_1k_blocs)
        #print("and here is the length of that list of custs: " , len(requesting_emp_list))

    #test_num = 513
    num_1k_blocs = len(requesting_emp_list) // 999
    remainder = len(requesting_emp_list) % 999
    max_number_divisible = len(requesting_emp_list) - remainder
    #print(remainder)
    #print(num_1k_blocs)
    #print("this is max number of custs: ", max_number_divisible)
    start_chunk = 0
    num = 0
    requesting_emp_query_results = []


    if num_1k_blocs:
        while num < num_1k_blocs:
        ##for chunk in range(num_1k_blocs):
            #print("this is the start_chunk: ", start_chunk)
            end_chunk = start_chunk + 999
            ranged_save_EMPs = []
            #print("here is the start:", start_chunk, "and here is the end: " ,end_chunk)
            for number in range(start_chunk, end_chunk):
                ranged_save_EMPs.append(requesting_emp_list[number])
                if requesting_emp_list[number] == '00067416':
                    print(f"i have found the employee: " , requesting_emp_list[number], "at number {number}")
                #print(ranged_save_CUSTs)
                #print("this is the number: ", number)
            emp_sql_string = getFANsString(ranged_save_EMPs)
            #print("here is the cust sql string: ", cust_sql_string)
            ##print(query_3)
            requesting_emp_query_results.append(queryOracle(4, '', emp_sql_string))

            #print("and the tobusorg query results", toBusOrg_query_results)
            start_chunk = start_chunk + 999
            #print("here is the new start:", start_chunk)
            #print("this is the new start_chunk: ", start_chunk)
            if start_chunk == max_number_divisible:
                ranged_save_EMPs = []
                end_chunk = start_chunk + remainder
                for number in range(start_chunk, end_chunk):
                    ranged_save_EMPs.append(requesting_emp_list[number])
                emp_sql_string = getFANsString(ranged_save_EMPs)
                requesting_emp_query_results.append(queryOracle(4, '', emp_sql_string))
            #print(num)            
            num += 1

        '''
        i = 0
        x = 0
        while i < 1:
            while x < len(toBusOrg_query_results[i]):
                #print("here is the item" , toBusOrg_query_results[i][x])
                if toBusOrg_query_results[i][x][0] == '3-A76052': ## item[0] == to-customer
                    print(f'Yes, I found that busorg! Number {x}: ', toBusOrg_query_results[i][x])
                    x+=1
                else:
                    x+=1
            i+=1
        '''
            #print("this is the num: ", num)
            #print("chunk: ", start_chunk)
        
    else:
        requesting_emp_query_results.append(queryOracle(4, '', requesting_emp_list_string))

    requesting_emp_query_holder = []
    i = 0
    for item in requesting_emp_query_results:
        for thingy in item:
            requesting_emp_query_holder.append(list(thingy))

    requesting_emp_query_holder_list = str([str(x) for x in requesting_emp_query_holder]).strip("[]")

    if len(requesting_emp_query_holder) > 0:
        #print("wooot!!!!!!!!!!!!!!!!!!!!!!")
        for emp_result in requesting_emp_query_holder:
            stored_requesting_employee_info_query_results.append(emp_result)
            for item in ban_ticket_andToCustomer_info:
                if item[1] == emp_result[0]:
                    #print("\n\n\nhere is the cust_result", cust_result)
                    #print(item[11])
                    #print(cust_result[0])
                    i = 0
                    while i < len(emp_result):
                        item.append(emp_result[i])
                        #print("the printed cust result:", cust_result[i])
                        i+=1
                    #item.append(str(str(x) for x in ban_result))
                    #print("\nto-cust result: ", cust_result[0])
                    #if item[14] == '300890983-ENS':
                        #print("\n\n\nfull sheet result: ", item)
        return ban_ticket_andToCustomer_info
    else:
        return ban_ticket_andToCustomer_info



def getKenanSrvcStatusAndOrderSystemsAndVoiceNameCheck(data_list):
    # this functions checks Kenan billing account service status, and Order Entry systems
    global kenanBillingAccountsODSids
    global kenanBillingAccounts
    global kenanBillingAccountsODSidsString

    global FANandOrderEntryDict

    global allBanODSids
    
    kenan_ods_list = []
    allBANsOESystems = []


    
    for row in data_list:
        billingSystem, BANodsID, financeAcctNbr = row[27], row[25], row[14]
        BANhasRevenue = row[17]

        #find dem kenan accounts and add 'em to their lists
        # if any pure Kenan accounts have active services, but no revenue then they should be flagged as possible voice accounts
        # --> the code below will return active services or not
        if re.search(kenanFanPattern, financeAcctNbr):
            #print("there is a match on '-A' for Kenan accounts: ", financeAcctNbr, BANodsID)
            kenan_book = []
            kenan_book.append(financeAcctNbr)
            kenan_book.append(BANodsID)
            kenanBillingAccounts.append(kenan_book)
            kenan_ods_list.append(BANodsID)

        # appending all billing account ODS ids to the global list
        allBanODSids.append(row[25])
 

    #print("\nhere is the kenan ODS list:", kenan_ods_list)

    kenanAcctsSrvcInactiveDate = ''

    if kenan_ods_list:
        for ods in kenan_ods_list:
            kenanBillingAccountsODSids.append(str(ods))
            kenanBillingAccountsODSidsString = getFANsString(kenanBillingAccountsODSids)
            # print("\n\n", kenanBillingAccountsODSidsString, " --> here it is, the Kenan ods ids \n\n")

    
        
        kenanAcctsSrvcInactiveDate = queryOracle(5, '', kenanBillingAccountsODSidsString)
        kenanAcctsSrvcInactiveDateODSids = []
     ## --> need to get the oDS ids that are coming from the query - i.e., the ones that are active with no billing

    if kenanAcctsSrvcInactiveDate != '' and kenanAcctsSrvcInactiveDate is not None:
        for item in kenanAcctsSrvcInactiveDate:
            kenanAcctsSrvcInactiveDateODSids.append(item[1])
        
        
    print(f"\n\n\nhere si the kenan data:", kenanAcctsSrvcInactiveDate)

    ## if this is a kenan account with no revenue and active services then mark the column Y and add the 
    if kenanAcctsSrvcInactiveDate != '' and kenanAcctsSrvcInactiveDate is not None:
        for row in data_list:
            banODSid = row[25]
            revenueCheck = row[17]
            if banODSid in kenanAcctsSrvcInactiveDateODSids:
                for item in kenanAcctsSrvcInactiveDate:
                    kenanBanOdsID = item[1] 
                    #print("\nand the other data: ", row[24], row[16], item[1])
                    if kenanBanOdsID == banODSid:
                        if revenueCheck == 'No':
                            # print("\nand item 1 and row 24, and row 16,", item[1], banODSid, revenueCheck)
                            #row.append('hapsburgers')
                            row.append('Y')
                        elif revenueCheck == 'Yes': 
                            # print("\nmade it here -- and item 1 and row 24, and row 16,", item[1], banODSid, revenueCheck)
                            #row.append('shplurgenstein')
                            row.append('N')
                    elif kenanBanOdsID != banODSid:
                        #print("\nmade it to the pass")
                        # print("and here's my schwag -- and item 1 and row 24, and row 16,", item[1], banODSid, revenueCheck)
                        pass
                    else:
                        # print("\nnot sure what happened")
                        # print("here's the show - and item 1 and row 24, and row 16,", item[1], banODSid, revenueCheck)
                        #row.append('fantasm')
                        row.append('UNK')
                    #print("\nand here's the data we're checking", item[1], banODSid, revenueCheck)
            else:
                #print("made it to the last else")
                #row.append('')
                row.append('')
            # print("\nand after the build, here is the row, ", row, "and the length", len(row))
    else:
        for row in data_list:
            row.append('')




 
    if len(allBanODSids) > 999:
        # print("\n\ncust is long")
        cust_data = []
        num_1k_blocs = len(allBanODSids) / 999
        # print("this is the total over 1k: " ,num_1k_blocs)
        #print("and here is the length: " , len(ssms_results))

        #test_num = 513
        num_1k_blocs = len(allBanODSids) // 999
        remainder = len(allBanODSids) % 999
        max_number_divisible = len(allBanODSids) - remainder
        #print(remainder)
        #print(num_1k_blocs)
        #print("this is max number: ", max_number_divisible)
        start_chunk = 0
        num = 0
        rev_query_results = []


        if num_1k_blocs:
            while num < num_1k_blocs:
            ##for chunk in range(num_1k_blocs):
                #print("this is the start_chunk: ", start_chunk)
                end_chunk = start_chunk + 999
                ranged_save_ods_ids = []
                for number in range(start_chunk, end_chunk):
                    ranged_save_ods_ids.append(allBanODSids[number])
                    #print(start_chunk, end_chunk)
                    #print(ranged_save_FANs)
                    #print("this is the number: ", number)
                    
                # print("here are the saved ODS IDs", ranged_save_ods_ids)
                allBANODSidsString = getFANsString(ranged_save_ods_ids)
                # print("here are the strings", allBANODSidsString)
                ##print(query_3)
                allBANsOESystems.append(queryOracle(6, '', allBANODSidsString))
                start_chunk = start_chunk + 999

                print("here is the oe systesm:", allBANsOESystems)
                #print("this is the new start_chunk: ", start_chunk)
                if start_chunk == max_number_divisible:
                    ranged_save_ods_ids = []
                    end_chunk = start_chunk + remainder
                    for number in range(start_chunk, end_chunk):
                        ranged_save_ods_ids.append(allBanODSids[number])
                    allBANODSidsString = getFANsString(ranged_save_ods_ids)
                    #fans_sql_string = fans_sql_string + fans_string # append fans to the fans_sql_string so we can use that global variable later
                    allBANsOESystems.append(queryOracle(6, '', allBANODSidsString))
                num += 1
                #print("this is the num: ", num)
                #print("chunk: ", start_chunk)
            #cust_data = [cust for sublist in cust_data for item in sublist for cust in item]

        # print("\nZAMBAZOO - here are the OE system info ", allBANsOESystems)
        allBANsOESystems = [list(item) for sublist in allBANsOESystems for item in sublist if item] #"if item" makes sure not to include the blank entries (if, by chance, some chunk has fans with OE systems attached and the other chunks do not)

    else:
        allBANODSidsString = getFANsString(allBanODSids)
        allBANsOESystems = queryOracle(6, '', allBANODSidsString)
        #print("\n\n here are all the BAN ODS IDs: " , allBANODSidsString) 

    
    #print("here's all the stuff for the OE Systems", allBANsOESystems)



    listBANodsIdsWithFlaggedOESystems = []
    inList = []
        

    if allBANsOESystems != '' and allBANsOESystems is not None:
        for item in allBANsOESystems:
            fan = item[2]
            OEcode = item[0]
            ODSID = item[1]
            if fan in FANandOrderEntryDict:
                FANandOrderEntryDict[item[2]].append(OEcode)
            else:
                FANandOrderEntryDict[item[2]] = [OEcode]
            if ODSID in inList: # since there are multiple entries for each billing account - we don't need dupes in the list
                continue
            else:
                listBANodsIdsWithFlaggedOESystems.append(ODSID)
                inList.append(ODSID)

        print("\n\nall BAN ODS IDS flagged for OE systems: ", listBANodsIdsWithFlaggedOESystems)

        for row in data_list:
            billingSystem, BANodsID, financeAcctNbr = row[27], row[25], row[14]
            BANhasRevenue = row[17]
            if BANodsID in listBANodsIdsWithFlaggedOESystems:
                row.append('Y')
            else:
                row.append('N')
            


    
                
            
        
        
        

    # if any billing accounts a
    

def createErrorAndMappingCSVs(data_list):
    global can_map_list
    global cannot_map_list
    global can_map_bans_list
    global cannot_map_bans_list

    global research_list
    global maybe_can_map_list

    global definitely_map_these_mamma_jammas_list

    global oracleBillingAccountsODSids
    global oracleBillingAccounts
    global oracleBillingAccountsODSidsString

    global bac_team
    global bac_team_empIDs

    global voice_services

    global billAcctActvtyStatusCd

    global allFans

    global ens_BANs

    global allTickets

    
    ## field_nums = [0, 14, 11, 18, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 40 ]
    green_field_nums =[0, 14, 56, 24, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 77]
    red_field_nums = [0, 14, 56]
    error_field_nums = [0, 14, 16, 25, 28, 29, 36, 54, 61, 63, 67, 68, 69]

    punctuation = ["'", ",", ".", "-", "_", "#"]


    consumer100 = '3-YCQXZX'
    damages = '3-RH5QJP'
    oracle_ods_list = []
    
    ####
    # don't move anything but LUMEN (1) to anything but LUMEN (1)
    #### 
    for row in data_list:
        # print("\n\nhere is the row: ", row)
        ### --> salesforce systems
        ### --> can combine both the red and green systems below with less text than these are built with....
        ### --> won't include ORACLE accounts here -- instead I'll build the whole file with the parent info and other
        ##### --> green accounts hopefully should ignore the blank field
        # row 28 == secure company code (1 == Lumen) -- if it is anything but Lumen, then we probably don't want to move too or from

        # if the billing account is a Green biller, the requested customer is Approved, and it is not a House account \
        ### --> and it is a Lumen secure BAN and a Lumen secure Requested customer
        # --> all conditions resolve to true
        

        requestedCustDWSecure, custDWSecure, banDWSecure = row[82], row[38], row[57]
        billingSystem, BANodsID, billAcctNbr, financeAcctNbr = row[27], row[25], row[10], row[14]
        requestedCust, rqstd_cust_status = row[11], row[83]
        billAcctActvtyStatusCd = row[47]
        currntCustChannel, currntCustLOB = row[35], row[32]
        crrnt_cust_bu_or_ext_sales_channel = row[34]
        full_billing_address_list = [row[51] , row[53] , row[54] , row[55] , row[56]]
        full_requestedCust_address_list = [row[72] , row[73] , row[74] , row[75] , row[76]]
        currentCust, rqstdCust = row[30], row[11]
        billAcctLevelType = row[61]
        billCycleCode = row[62]
        billStatusIndForKenan = row[63]
        billAcctName = row[26].upper()
        banAge = row[29]
        ban_3_month_avg = row[19]
        kenanActvSrvcsAndRevenueCheck = row[105]
        badOESystemCheck = row[106]
        empID = row[1]
        currentHouseAcctIdentifier = row[40]
        rqstr_emp_nbr = row[1]

        banRevFlag = row[17]

        crrnt_cust_nbr, rqstd_cust_nbr = row[30], row[11]

        rqstd_cust_house_accnt_cd, crrnt_cust_house_accnt_cd = row[84], row[40]
        rqstd_cust_lob, crrnt_cust_lob = row[77], row[32]
        rqstd_cust_name = row[81].upper()
        rqstd_cust_bu_or_ext_sales_channel = row[78]   

        ticket = row[0] 


        # append all fans to the allFans list
        allFans.append(financeAcctNbr)

        allTickets.append(ticket)

        # print("\n\nwe's a huntin this row now to start evaluation: ", financeAcctNbr, "\n", row)
        #print("here is FAN, ODSid, and Billing system and BadOE system check: ", financeAcctNbr, BANodsID, billingSystem, badOESystemCheck)
        # print("here is the bill status indicator for kenan (mainly): ", billStatusIndForKenan)

        # print("here is the billAcctLvelType ", billAcctLevelType, financeAcctNbr)
        
        #find dem oracle accounts and add them to a list of lists
        if billingSystem in oracle:
            oracle_book = []
            oracle_book.append(financeAcctNbr) # fan
            oracle_book.append(BANodsID) # ods id
            oracleBillingAccounts.append(oracle_book)
            oracle_ods_list.append(BANodsID)
        elif billingSystem == 'ENS':
            ens_BANs.append(billAcctNbr)

        
        ''' ## TEST TEST    
        if ((row[31] in SBG and row[72] in indirect) == False) and (row[11] != consumer100 and row[29] != consumer100) or (row[11] == damages and row[29] != damages)\
           and billAcctLevelType != 'FALLOUT' and not re.search(voice_services, billAcctName):
            print("\n\nguess it found a match on voice services?")
            print("here is the bill account level type", billAcctLevelType)
        '''
        
        if (billingSystem in salesforce_billing_systems or (billingSystem in ['IDC_KENAN', 'KENANFX'] and re.match(r'.*\-LATIS$', financeAcctNbr))) \
            and ((custDWSecure == 1 or custDWSecure is None or custDWSecure == '') and requestedCustDWSecure == 1 and banDWSecure == 1) \
            and (crrnt_cust_nbr != rqstd_cust_nbr) and (rqstd_cust_house_accnt_cd == 0 or (rqstd_cust_house_accnt_cd == 1 and crrnt_cust_house_accnt_cd == 1)) \
            and rqstd_cust_status == 'Approved' and ((currntCustLOB in SBG and rqstd_cust_lob in indirect) == False) and ((rqstd_cust_nbr != consumer100 and crrnt_cust_nbr != consumer100) \
            or (rqstd_cust_nbr == damages and crrnt_cust_nbr != damages)) and ((badOESystemCheck == 'Y' and currentCust in placeholder_customers) \
            or (badOESystemCheck == 'Y' and empID in bac_team_empIDs) \
            or (badOESystemCheck == 'Y' and re.match(r'.*\-LATIS$', financeAcctNbr)) \
            or (badOESystemCheck == 'Y' and re.match(r'.*\-GC$', financeAcctNbr)) or badOESystemCheck == 'N') \
            and billAcctLevelType != 'FALLOUT' and not re.search(voice_services, billAcctName) and (requestedCust != cpniIssuesCust):
            #and billAcctActvtyStatusCd == 'Active':
            #and ((row[63] not in indirect and row[31] not in SBG) is True): requested cust in SBG moving to Indirect is an issue
            if financeAcctNbr == '333973800-ENS':
                print("\n\nthis one made it into the green sheet so far: " , financeAcctNbr, 
                      "\ncurrent cust nbr: ", crrnt_cust_nbr,
                      "\nrequestedCustDWSecure: ", requestedCustDWSecure, 
                      "\ncustDWSecure: ", custDWSecure,
                      "\nrequested cust house account code: ", rqstd_cust_house_accnt_cd,
                      "\ncurrent house account identifier: ", currentHouseAcctIdentifier,
                      "\nbill acct name: ", billAcctName,
                      "\nrequested cust name: ", rqstd_cust_name,
                      "\nbanAge: ", banAge,
                      "\nban rev flag: ", banRevFlag)
            create_a_list = []# clears the list for each row in data_list
            #separate green from red inside of each if statement
            #print("here is the uppercase ban name: ", row[25], "\nhere is the uppercase rqstd cust name: ", row[76].upper())
            #print("yes they match") if row[25] == row[76] else print("nope they dno't match")

            # if cust unassigned or requsted cust is legit cust and crrnt cust is house
            if ((crrnt_cust_nbr == '' or crrnt_cust_nbr is None) or (rqstd_cust_house_accnt_cd == 0 and currentHouseAcctIdentifier == 1)) \
                and ((billAcctName is not None and rqstd_cust_name is not None) and (billAcctName != '' and rqstd_cust_name !='')): 
                
                if financeAcctNbr == '333973800-ENS':
                    print("made it t othis section, bruh: ", financeAcctNbr)
                if (crrnt_cust_nbr == '' or crrnt_cust_nbr is None) and ((billAcctName == rqstd_cust_name) or (re.sub(r'[^a-zA-Z0-9]', '', billAcctName)) == re.sub(r'[^a-zA-Z0-9]', '', rqstd_cust_name) ) : #r'[^\w\s]'
                    #print("made it here somehow, this BAN: ", financeAcctNbr)
                    # print("here is the uppercase ban name: ", billAcctName, "\nhere is the uppercase rqstd cust name: ", rqstd_cust_name)
                    #print("here is row29, ", row[29])
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"{financeAcctNbr} ain't done been assigned-ified. Go ahead and map 'im")
                    row.append(f"Billing account name and Requested customer name are perfect match. Plus, the billing account is currently unmapped and heading to a legit customer. This one makes all kinds of sense.")
                    row.append("Mapped")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(financeAcctNbr)
                    row.insert(0, "Map it!")
                #if current cust has House designation (1) or if the customer is in the Placeholders list and name is match
                elif (crrnt_cust_house_accnt_cd == 1 or crrnt_cust_nbr in placeholder_customers) and ((billAcctName == rqstd_cust_name) or \
                     (re.sub(r'\.', '', billAcctName) == re.sub(r'\.', '', rqstd_cust_name) ) ):
                    
                    #print("name is a match and currnt attached to placeholder cust - green: ", financeAcctNbr)
                    if crrnt_cust_house_accnt_cd == 1 and crrnt_cust_nbr not in placeholder_customers:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"{financeAcctNbr} ain't done been assigned-ified. Go ahead and map 'im")
                        row.append(f"Billing account name and Requested customer name are perfect match, plus, the billing account is mapped to a House customer and heading to a legit customer. Map 'im!")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                    elif crrnt_cust_nbr in placeholder_customers:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"{financeAcctNbr} ain't done been assigned-ified. Go ahead and map 'im")
                        row.append(f"Billing account name and Requested customer name are perfect match. Plus, the billing account is mapped to a Placeholder customer and heading to a legit customer. Do it.")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                    #if ((row[66] == row[52]) and (row[67] == row[53])) and ((row[31] == row[70]) or (row[31] is None or row[31] == ''))): #if city and state match between BAN and RQSTD / and if the movement is not cross lob / and if requested cust is not Indirect or crrnt and rqstd are both Indirect
                else:
                    if financeAcctNbr == '333973800-ENS':
                        print("\n\nshucks, they don't match: ", financeAcctNbr)
                    if currentHouseAcctIdentifier == 1 and crrnt_cust_nbr not in placeholder_customers:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"{financeAcctNbr} ain't done been assigned-ified. Go ahead and map 'im")
                        row.append(f"Billing account name and Requested customer name do not match... but the billing account is mapped to a House customer and heading to a legit customer.")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                    elif crrnt_cust_nbr in placeholder_customers:
                        if financeAcctNbr == '333973800-ENS':
                            print("\ncurrent cust = placeholder: ", financeAcctNbr)
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"{financeAcctNbr} ain't done been assigned-ified. Go ahead and map 'im")
                        row.append(f"Billing account name and Requested customer name do not match... but the billing account is mapped to a Placeholder customer and heading to a legit customer.")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                    elif (crrnt_cust_nbr == '' or crrnt_cust_nbr is None):
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"{financeAcctNbr} ain't done been assigned-ified. Go ahead and map 'im")
                        row.append(f"Billing account name and Requested customer name do not match... but the billing account is currently unmapped and heading to a legit customer.")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                    elif (crrnt_cust_nbr != '' or crrnt_cust_nbr is not None) and (rqstd_cust_house_accnt_cd == 0 and currentHouseAcctIdentifier == 0):
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"{financeAcctNbr} ain't done been assigned-ified. Go ahead and map 'im")
                        row.append(f"Billing account name and Requested customer name do not match... but the billing account is currently mapped to a legit customer and heading to a legit customer.")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                '''           
            elif row[29] in placeholder_customers or row[68] == 1:
                if (row[25].upper() == row[67].upper()) or (re.sub(r'\.', '', row[25].upper()) == re.sub(r'\.', '', row[67].upper()) ): #r'[^\w\s]'
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                        create_a_list.append(f"{row[14]} ain't done been assigned-ified. Go ahead and map 'im")
                        row.append(f"BAN is currently unmapped | BAN name and Requested Customer name are perfect match")
                        green_can_map_list.append(create_a_list)
                        maybe_can_map_list.append(row[14])
                else:
                    print("\n\n shucks, they don't match")
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                        create_a_list.append(f"{row[14]} ain't done been assigned-ified. Go ahead and map 'im")
                        row.append(f"BAN is currently unmapped -- can't tell if the BAN name and Requested Customer names match or not")
                        green_can_map_list.append(create_a_list)
                        maybe_can_map_list.append(row[14])
                        
                '''
            elif banAge == 'Under 90':
                #create_a_list = [] # clears the list for each row in data_list
                # fan, fan name, ban_age, ban_revenue, ban_lob, ban_channel, to_cust_nbr, to_cust_status, to_cust_lob, to_cust_channel
                # appending multiple indices to the same row is proving to be a challenge, so, I've
                #--> decided to create a row, and then append that. It seems to work like a charm.
                if financeAcctNbr == '333973800-ENS':
                    print("ban is under 90 and green BABAY!")
                if (crrnt_cust_lob in SBG) and (rqstd_cust_lob not in SBG):
                    if rqstr_emp_nbr == '00096184': #00302684 - matty t # jacquie 00096184
                        # print("at 6 green" , row[14])
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{financeAcctNbr} is under 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is under 90 days old, and the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel} - but the request comes from Jacquie, so it is likely okay")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                    elif rqstr_emp_nbr == '00302684':
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{financeAcctNbr} is under 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is under 90 days old, and the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel} and the request comes from Matty T, so it is likely okay")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                    elif currentCust in placeholder_customers and rqstdCust not in placeholder_customers:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{row[14]} is under 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is under 90 days old, and the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel}. And it is moving FROM a Placeholder to a legit customer. Probably a go.")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(row[14])
                        row.insert(0, "Map it!")
                    elif currentCust not in placeholder_customers and rqstdCust not in placeholder_customers:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{row[14]} is under 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is under 90 days old, and the movement is cross-BU (from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel}) - should be good to go")
                        row.append(f"Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(row[14])
                        row.insert(0, "Map it!")
                    elif currentCust not in placeholder_customers and rqstdCust in placeholder_customers:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{row[14]} is under 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is under 90 days old, but the movement is cross-BU to a Placeholder customer. Nah. ")
                        row.append(f"Feedback Required: The movement is Cross-BU, from {crrnt_cust_bu_or_ext_sales_channel} to a Placeholder in ({rqstd_cust_bu_or_ext_sales_channel}). Please work with Sales, Sales Ops, your peers, or leadership to find or create the desired customer, then update this ticket.")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(row[14])
                        row.insert(0, "Feedback Required")
                    elif currentCust in placeholder_customers and rqstdCust in placeholder_customers:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{row[14]} is under 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is under 90 days old, there is revenue attached, and the movement is cross-BU from a Placeholder to a Placeholder customer. Nah. ")
                        row.append(f"REVIEW == Mapped | Feedback Required: The requested movement is Cross-BU: from an {crrnt_cust_bu_or_ext_sales_channel} Placeholder to a Placeholder in ({rqstd_cust_bu_or_ext_sales_channel}). Please work with Sales, Sales Ops, your peers, or leadership to find or create the desired customer, then update this ticket.")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(row[14])
                        row.insert(0, "Review")
                    else:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{row[14]} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                        row.append(f"Not sure why this request made it here. Dave needs to review")
                        row.append("REVIEW == Mapped | Feedback Required: The requested movement needs more data.")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(row[14])
                        row.insert(0, "Review")
                elif badOESystemCheck == 'Y' and empID in bac_team_empIDs:
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{row[14]} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                    row.append(f"This billing account is associated to an Order Entry (OE) system that cannot be disassociated from its current customer - But the request comes from the BAC team, so hopefully they are already working on it. 'sokay")
                    row.append("Mapped - you are on our exception list, you lucky duck, you.")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(row[14])
                    row.insert(0, "Map it!")
                elif (rqstdCust == damages and (currentCust != damages or currentCust is None or currentCust == '')): #row[11] is requested cust. so if rqstd cust is damages and crrnt cust is not
                    # print("at 2 green" , row[14])
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{row[14]} is over 90 days old, LOB of current customer and requested is {row[28]}, to-customer is Approved')
                    if re.search('^(BF|BH).*(-BART|-MBS)$', financeAcctNbr):
                        if (currentCust is None or currentCust == ''):
                            row.append(f"The billing account is a BART/MBS account that begins in 'BF' or 'BH' and is currently unmapped -- does it belong in Damages? Yes it do! Map it.") 
                            row.append("Mapped")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(row[14])
                            row.insert(0, "Map it!")
                        elif (currentCust is not None and currentCust != damages):
                            row.append(f"The billing account is a BART/MBS account that begins in 'BF' or 'BH'. It belongs in Damages and isn't mapped there right now. Let's map 'er. ") 
                            row.append("Mapped")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(row[14])
                            row.insert(0, "Map it!")
                    elif re.search('^.*(-BART|-MBS)$', financeAcctNbr):
                        if empID == '00096221': # steve isaac
                            row.append(f"The billing account is a BART/MBS account and the request comes from Steve Isaac. There's a good chance that the request to map to Damages is accurate. ") 
                            row.append("Mapped")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(row[14])
                            row.insert(0, "Map it!")
                        else:
                            row.append(f"The billing account is a BART/MBS account. There's a chance that the request to map to Damages is accurate, but we need more info first. ") 
                            row.append("REVIEW == Mapped | Feedback Required")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(row[14])
                            row.insert(0, "Review")
                    else:
                        row.append(f"The billing account isn't MBS/BART -- it shouldn't be in Damages") 
                        row.append("Feedback Required: this billing account is not BART or MBS - it shouldn't be mapped to Damages. Please work with Sales or Sales Ops to find or create the correct customer and then update this ticket.")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(row[14])
                        row.insert(0, "Review")
                else:
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{row[14]} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                    row.append(f"The billing account is under 90 days old. Barring anything unforeseen, we SHOULD be able to map this one.")
                    row.append("REVIEW == Mapped | Feedback Required")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(row[14])
                    row.insert(0, "Review")
            elif banAge == 'Over 90':
                if financeAcctNbr == '333973800-ENS':
                    print("\nmade it to Over 90 selection for Greens: ", row[14])
                if banRevFlag == 'No':
                    # print("at 1 green" , row[14])
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{row[14]} is over 90 days old, no revenue, to-customer is Approved')
                    row.append(f"Billing account is over 90 days old but there is no revenue associated with it over the last 13 months, i.e., there is no financial impact")
                    row.append("Mapped")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(row[14])
                    row.insert(0, "Map it!")
                elif currntCustLOB == rqstd_cust_lob:
                    # print("at 2 green" , row[14])
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{row[14]} is over 90 days old, LOB of current customer and requested is {currntCustLOB}, to-customer is Approved')
                    row.append(f"Billing account is over 90 days old, but there is no financial impact (Current and Requested customer LOBs are the same)") 
                    row.append("Mapped")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(row[14])
                    row.insert(0, "Map it!")
                elif (currntCustLOB in fed) or (rqstd_cust_lob in fed):
                    # print("at 3 green" , row[14])
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{row[14]} is over 90 days old, to-customer LOB is FED, to-customer is Approved')
                    if currntCustLOB in fed:
                        row.append(f"Billing account is over 90 days old but the Current customer is in Federal channel")
                    elif rqstd_cust_lob in fed:
                        row.append(f"Billing account is over 90 days old but the Requested customer is in Federal channel")
                    row.append("Mapped")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(row[14])
                    row.insert(0, "Map it!")
                elif (currntCustLOB in internal) or (rqstd_cust_lob in internal):
                    # print("at 4 green" , row[14])
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{row[14]} is over 90 days old, to-customer LOB is FED, to-customer is Approved')
                    if currntCustLOB in internal:
                        row.append(f"Billing account is over 90 days old, but it is in Internal channel -- it can move")
                    elif rqstd_cust_lob in internal:
                        row.append(f"Billing account is over 90 days old, but the requested Customer is in Internal channel -- we can move this")
                    row.append("Mapped")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(row[14])
                    row.insert(0, "Map it!")
                elif (rqstd_cust_nbr == damages and (crrnt_cust_nbr != damages or crrnt_cust_nbr is None or crrnt_cust_nbr == '')): #row[11] is requested cust. so if rqstd cust is damages and crrnt cust is not
                    # print("at 2 green" , row[14])
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{row[14]} is over 90 days old, LOB of current customer and requested is {currntCustLOB}, to-customer is Approved')
                    if re.search('^(BF|BH).*(-BART|-MBS)$', financeAcctNbr):
                        if (crrnt_cust_nbr is None or crrnt_cust_nbr == ''):
                            row.append(f"The billing account is a BART/MBS account that begins in 'BF' or 'BH' and is currently unmapped -- does it belong in Damages? Yes it do! Map it.") 
                            row.append("Mapped")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(row[14])
                            row.insert(0, "Map it!")
                        elif (crrnt_cust_nbr is not None and crrnt_cust_nbr != damages):
                            row.append(f"The billing account is a BART/MBS account that begins in 'BF' or 'BH'. It belongs in Damages and isn't mapped there right now. Let's map 'er. ") 
                            row.append("Mapped")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(row[14])
                            row.insert(0, "Map it!")
                    elif re.search('^.*(-BART|-MBS)$', financeAcctNbr):
                        if empID == '00096221': # steve isaac
                            row.append(f"The billing account is a BART/MBS account and the request comes from Steve Isaac. There's a good chance that the request to map to Damages is accurate. ") 
                            row.append("Mapped")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(row[14])
                            row.insert(0, "Map it!")
                        else:
                            row.append(f"The billing account is a BART/MBS account. There's a chance that the request to map to Damages is accurate, but we need more info first. ") 
                            row.append("REVIEW == Mapped | Feedback Required")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(row[14])
                            row.insert(0, "Review")
                    else:
                        row.append(f"The billing account isn't MBS/BART -- it shouldn't be in Damages") 
                        row.append("Feedback Required: this billing account is not BART or MBS - it shouldn't be mapped to Damages. Please work with Sales or Sales Ops to find or create the correct customer and then update this ticket.")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(row[14])
                        row.insert(0, "Review")
                elif (currntCustLOB not in SBG) and (rqstd_cust_lob not in SBG): #SBG = ['541', '328', '273', '542']
                    # print("at 5 green" , row[14])
                    if crrnt_cust_bu_or_ext_sales_channel == rqstd_cust_bu_or_ext_sales_channel: #originally compared sales channel, but I think BU is the better comparison
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{row[14]} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Current and Requested customer LOBs are both outside SBG (and under the same Channel); i.e., there is no financial impact to MASS MARKETS (SBG) and 0 financial impact overall")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(row[14])
                        row.insert(0, "Map it!")
                    else:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{row[14]} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Current and Requested customer LOBs are both outside SBG; i.e., there is no financial impact to MASS MARKETS (SBG)")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(row[14])
                        row.insert(0, "Map it!")
                elif (crrnt_cust_lob in SBG) and (rqstd_cust_lob not in SBG) and (banRevFlag == 'Yes'):
                    if row[1] == '00096184': #00302684 - matty t # jacquie 00096184
                        # print("at 6 green" , row[14])
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{row[14]} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is over 90 days old, there is revenue attached, and the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel} - but the request comes from Jacquie, so it is likely okay")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(row[14])
                        row.insert(0, "Map it!")
                    elif row[1] == '00302684':
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{row[14]} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is over 90 days old, and there is revenue attached. But the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel} and the request comes from Matty T, so it is likely okay")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(row[14])
                        row.insert(0, "Map it!")
                    elif int(ban_3_month_avg) < 2000: 
                        # print("at 7 green" , row[14])
                        if crrnt_cust_nbr in placeholder_customers and rqstd_cust_nbr not in placeholder_customers:
                            for num in green_field_nums:
                                create_a_list.append(row[int(num)])
                            create_a_list.append(f'{row[14]} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                            row.append(f"Billing account is over 90 days old, there is revenue attached, and the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel} - but the revenue is less than $2k. And it is moving FROM a Placeholder to a legit customer. Warrants a look.")
                            row.append("Mapped")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(row[14])
                            row.insert(0, "Map it!")
                        elif crrnt_cust_nbr not in placeholder_customers and rqstd_cust_nbr not in placeholder_customers:
                            for num in green_field_nums:
                                create_a_list.append(row[int(num)])
                            create_a_list.append(f'{row[14]} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                            row.append(f"Billing account is over 90 days old, there is revenue attached, and the movement is cross-BU (from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel}) - but the revenue is less than $2k. Warrants a look.")
                            row.append(f"REVIEW == Mapped | Feedback Required: The movement is Cross-BU - out of {crrnt_cust_bu_or_ext_sales_channel} and heading to {rqstd_cust_bu_or_ext_sales_channel} and there is revenue attached -- why is this movement necessary?")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(row[14])
                            row.insert(0, "Review")
                        elif crrnt_cust_nbr not in placeholder_customers and rqstd_cust_nbr in placeholder_customers:
                            for num in green_field_nums:
                                create_a_list.append(row[int(num)])
                            create_a_list.append(f'{row[14]} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                            row.append(f"Billing account is over 90 days old, there is revenue attached, and the movement is cross-BU to a Placeholder customer. Nah. ")
                            row.append(f"Feedback Required: The movement is Cross-BU - from {crrnt_cust_bu_or_ext_sales_channel} to a Placeholder customer in ({rqstd_cust_bu_or_ext_sales_channel}). Please work with Sales or Sales Ops to find or create the desired customer, then update this ticket.")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(row[14])
                            row.insert(0, "Review")
                        elif crrnt_cust_nbr in placeholder_customers and rqstd_cust_nbr in placeholder_customers:
                            for num in green_field_nums:
                                create_a_list.append(row[int(num)])
                            create_a_list.append(f'{row[14]} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                            row.append(f"Billing account is over 90 days old, there is revenue attached, and the movement is cross-BU from a Placeholder to a Placeholder customer. Nah. ")
                            row.append(f"Feedback Required: The movement is Cross-BU - from an {crrnt_cust_bu_or_ext_sales_channel} Placeholder to a Placeholder customer ({rqstd_cust_bu_or_ext_sales_channel}). Please work with Sales or Sales Ops to find or create the desired customer, then update this ticket.")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(row[14])
                            row.insert(0, "Review")
                        else:
                            for num in green_field_nums:
                                create_a_list.append(row[int(num)])
                            create_a_list.append(f'{row[14]} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                            row.append(f"Not sure why this request made it here. Dave needs to review")
                            row.append("REVIEW == Mapped | Feedback Required")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(row[14])
                            row.insert(0, "Review")
                    elif int(ban_3_month_avg) > 2000:
                        # print("at 8 green" , row[14])
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{row[14]} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is over 90 days old, there is revenue attached, the movement is from SBG to {rqstd_cust_bu_or_ext_sales_channel}, and the revenue is over $2k. Maybe not okay?")
                        row.append("REVIEW == Mapped | Feedback Required: The requested mapping is cross-BU and there is over $2k in revenue that will be moving - why is this movement desired?")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(row[14])
                        row.insert(0, "Review")
                        ### NEED PLACEHOLDER NOTIFICATIONS HERE ###
                    elif int(ban_3_month_avg) < 2000:
                        print("at 9 in reds " , row[14])
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{row[14]} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is over 90 days old, there is revenue attached, and the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {crrnt_cust_bu_or_ext_sales_channel} - but the revenue is less than $2k. Warrants a look.")
                        row.append("REVIEW == Mapped | Feedback Required: The movement is cross-BU, the billing account is over 90 days old and there is revenue attached. Why is this movement desired?")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(row[14])
                        row.insert(0, "Review")
                elif (currntCustLOB not in SBG) and (rqstd_cust_lob in SBG) and (banRevFlag == 'Yes'):
                    if row[1] == '00096184': #00302684 - matty t # jacquie 00096184
                        # print("at 8 green" , row[14])
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{row[14]} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is over 90 days old, there is revenue attached, and the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel} - but the request is from Jackie, so it is likely okay")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(row[14])
                        row.insert(0, "Map it!")
                    elif int(ban_3_month_avg) < 2000:
                        # print("at 9 green" , row[14])
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{row[14]} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is over 90 days old, there is revenue attached, and the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel} - but the revenue is less than $2k. Warrants a look.")
                        row.append("REVIEW == Mapped | Feedback Required")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(row[14])
                        row.insert(0, "Review")
                    elif int(ban_3_month_avg) > 2000:
                        # print("at 8 green" , row[14])
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is over 90 days old, there is revenue attached, the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel}, and the revenue is over $2k. Maybe not okay?")
                        row.append("REVIEW == Mapped | Feedback Required")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Review")
                elif (currntCustLOB in SBG) and (rqstd_cust_lob in SBG):
                    # print("at 9 green" , row[14])
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                    row.append(f"Billing account is only moving internally, within SBG, so hey, no biggie here - let's do it.")
                    row.append("Mapped")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(financeAcctNbr)
                    row.insert(0, "Map it!")
                elif badOESystemCheck == 'Y' and empID in bac_team_empIDs:
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                    row.append(f"This billing account is associated to an Order Entry (OE) system that cannot be disassociated from its current customer - But the request comes from the BAC team, so hopefully they are already working on it. 'sokay")
                    row.append("Mapped - you are on our exception list, you lucky duck, you.")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(financeAcctNbr)
                    row.insert(0, "Map it!")
                    
                        
                # print("\nmade it out of GREEN Over 90 selection: ", row[16])
            
            elif (rqstd_cust_house_accnt_cd == 1 and crrnt_cust_house_accnt_cd == 1): # current and requested customer are house accounts
                for num in green_field_nums:
                    create_a_list.append(row[int(num)])
                create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                row.append(f"Current customer and Requested customer are both House acounts - could be okay, but let's review the movement")
                row.append("REVIEW == Mapped | Feedback Required: please work with Sales, Sales Ops, or your peers to either find or create a customer, then update this ticket.")
                green_can_map_list.append(create_a_list)
                can_map_bans_list.append(financeAcctNbr)
                row.insert(0, "Review") 

                # print("to green error list")


            ###### ============= GREEN ERROR SHEET =================== ###############
            ## --> #handle the errors in green bans   
            ######## =================================================#################
            else:
                # print("\nlooks like this billing account made it to the GREEN ERROR SHEET: \n\n", financeAcctNbr)
                create_a_list = []
                error_book = []
                # append to row and change to elif statements to change back (if it is fucked)
                ## --> and this " cannot_map_list.append(create_a_list) "
                if rqstd_cust_status != 'Approved' or rqstd_cust_status == '':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"Customer # {rqstd_cust_nbr} is {rqstd_cust_status} so {financeAcctNbr} can't be mapped to it" )
                    row.append(f"Requested customer is not Approved")#
                    row.append("Feedback Required: the Requested customer is not Approved")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Feedback Required: please work with Sales, Sales Ops, or your peers to identify or create the correct customer and then update this ticket.")
                elif rqstd_cust_nbr == crrnt_cust_nbr: # if rqstd customer == currnt customer
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])#
                    create_a_list.append(f"The desired customer's Secure Company Number is not 1 (Lumen)... so... there's that")
                    row.append(f"Current customer and Requested customer are the same")
                    row.append("Validated - No Action Required: The billing account is already mapped to the desired customer")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Mark ticket: Closed - Resolved, Validated...")
                elif rqstd_cust_nbr in placeholder_customers:
                    if crrnt_cust_nbr == '' or rqstd_cust_nbr is None:
                        for num in error_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"Requested customer is a Placeholder" )
                        row.append(f"Requested customer is a Placeholder account, but the billing account is currently unmapped")
                        row.append("Mapped")
                        cannot_map_list.append(create_a_list)
                        row.insert(0, "Map it!")
                    else:
                        for num in error_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"Requested customer is a Placeholder" )
                        row.append(f"Requested customer is a Placeholder account")
                        row.append("Feedback Required: The Requested customer is a Placeholder account. Please work with Sales, Sales Ops, or your peers to find or create the right custome for this billing account and then update this ticket")
                        cannot_map_list.append(create_a_list)
                        row.insert(0, "Review")
                    ''' # removing Cross-BU problem
                if (row[31] in SBG and row[63] not in SBG) or (row[31] not in SBG and row[63] in SBG) and row[16] == 'Yes':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    #print(f"{row[14]} done made it to the SBG identfier in Error file")
                    create_a_list.append(f"BAN and desired customer have cross-BU LOB's" )
                    row.append(f"Movement is cross-BU")
                    cannot_map_list.append(create_a_list)
                    '''
                elif billingSystem == 'ORAFIN':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"{financeAcctNbr} is an Orafin build. It's a fallout build and can't be mapped for that reason" )
                    row.append(f"BAN is an Orafin billing account - we do not govern Orafin billing accounts.")
                    row.append("Feedback Required: Customer Hierarchy does not govern ORAFIN billing accounts. If you've identified the wrong billing account in this request, please visit this link for help with finding the right one: https://rb.gy/kxu12 -- then update this ticket.")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif billingSystem == 'BTN':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"""{financeAcctNbr} is a Fallout billing account. It should have a CRIS brother somewhere..
                                            It isn't a real billing account and cannot be mapped for that reason""" )
                    row.append(f"BAN is a BTN billing account - we do not govern BTN billing accounts.")
                    row.append(f"Feedback Required: Customer Hierarchy does not govern BTN billing accounts. Please visit this link for more information: {howToFindCRISBans}, then update this ticket once you've identified the correct, valid CRIS Finance Account Number")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif billingSystem == 'NFCN':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"""{financeAcctNbr} is a Fallout billing account. It should have a CRIS brother somewhere..
                                            It isn't a real billing account and cannot be mapped for that reason""" )
                    row.append(f"BAN is an NFCN billing account - we do not govern NFCN billing accounts.")
                    row.append(f"Feedback Required: Customer Hierarchy does not govern NFCN billing accounts. Please visit this link for more information: {howToFindCRISBans}, then update this ticket once you've identified the correct, valid CRIS Finance Account Number")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif billingSystem == 'ZUORA':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"""{financeAcctNbr} is a Fallout billing account. It should have a CRIS brother somewhere..
                                            It isn't a real billing account and cannot be mapped for that reason""" )
                    row.append(f"BAN is a ZUORA billing account -- we don't govern these. Traci Graham might")
                    row.append("Feedback Required: Customer Hierarchy does not govern ZUORA billing accounts. If this billing account was identified by mistake, please update this ticket with the actual, desired billing account")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                    '''
                elif row[37] != 1:
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The current customer's DW Secure Company Number is not 1 (Lumen)... so... there's that")
                    row.append("Current customer DW Secure Company Number is not 1")
                    '''
                elif banDWSecure != 1 or requestedCustDWSecure != 1: # --> DW Secure company code is not Lumen for either the BAN or Customer
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"{financeAcctNbr}'s DW Secure Company Number is not 1 (Lumen)... so... there's that")
                    if banDWSecure != 1 and requestedCustDWSecure == 1:
                        if banDWSecure == 4:
                            row.append("Billing account is Brightspeed, Requested customer is Lumen")
                            row.append("Denied - We do not govern Brightspeed billing accounts")
                        elif banDWSecure == 2:
                            row.append("Billing account is Cirion, Requested customer is Lumen")
                            row.append("Denied - We do not govern Cirion billing accounts")
                        elif banDWSecure == 6:
                            row.append("Billing account is Colt, Requested customer is Lumen")
                            row.append("Denied - We do not govern Colt billing accounts")
                    elif banDWSecure == 1 and requestedCustDWSecure != 1:
                        if requestedCustDWSecure == 4:
                            row.append("Requested customer is Brightspeed, billing account is Lumen")
                            row.append("Denied - We do not map Lumen billing accounts to Brightspeed customers")
                        elif requestedCustDWSecure == 2:
                            row.append("Requested customer is Cirion, billing account is Lumen")
                            row.append("Denied - We do not map Lumen billing accounts to Cirion customers")
                        elif requestedCustDWSecure == 6:
                            row.append("Requested Customer is Colt, billing account is Lumen")
                            row.append("Denied - We do not map Lumen billing accounts to Colt customers")
                    elif banDWSecure != 1 and requestedCustDWSecure != 1:
                        if banDWSecure == 4:
                            if requestedCustDWSecure == 4:
                                row.append("Billing account and Requested customer are Brightspeed")
                                row.append("Denied - we do not govern Brightspeed accounts")
                            elif requestedCustDWSecure == 2:
                                row.append("Billing account is Brightspeed, Requested customer is Cirion")
                                row.append("Denied - We do not govern Brightspeed billing accounts")
                            elif requestedCustDWSecure == 6:
                                row.append("Billing account is Brightspeed, Requested customer is Colt")
                                row.append("Denied - We do not govern Brightspeed billing accounts")
                        elif banDWSecure == 2:
                            if requestedCustDWSecure == 2:
                                row.append("Billing account and Requested customer are Cirion")
                                row.append("Denied - we do not govern Cirion accounts")
                            elif requestedCustDWSecure == 4:
                                row.append("Billing account is Cirion, Requested customer is Brightspeed")
                                row.append("Denied - We do not govern Cirion billing accounts")
                            elif requestedCustDWSecure == 6:
                                row.append("Billing account is Cirion, Requested customer is Colt")
                                row.append("Denied - We do not govern Cirion billing accounts")
                        elif banDWSecure == 6:
                            if requestedCustDWSecure == 2:
                                row.append("Billing account is Colt and Requested customer is Cirion")
                                row.append("Denied - we do not govern Colt billing accounts")
                            elif requestedCustDWSecure == 4:
                                row.append("Billing account is Colt, Requested customer is Brightspeed")
                                row.append("Denied - We do not govern Colt billing accounts")
                            elif requestedCustDWSecure == 6:
                                row.append("Billing account and Requested customer are Colt")
                                row.append("Denied - We do not govern Colt accounts")
                    else:
                        row.append("Either the BAN or the custoer is not Lumen -- needs to be Reviewed to fix")
                        row.append("Customer Hierarchy does not govern Cirion, Brightspeed, or Colt billing accounts (or customers). Please work with the Cirion, Brightspeed, or Colt teams instead. ")
                    cannot_map_list.append(financeAcctNbr)
                    row.insert(0, "Denied")
                elif rqstd_cust_house_accnt_cd == 1 and currentHouseAcctIdentifier == 0: #if rqstd cust is placeholder but crrnt cust is not
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The desired customer's Secure Company Number is not 1 (Lumen)... so... there's that")
                    row.append(f"Current customer is a valid, Enterprise customer. Requested customer is a House or Placeholder")
                    row.append("Feedback Required: Please work with Sales, Sales Ops, or your peers to identify an non-Placeholder, Enterprise customer to map this billing account to")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif requestedCust == cpniIssuesCust:
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The desired customer's Secure Company Number is not 1 (Lumen)... so... there's that")
                    row.append(f"Requested customer is the CPNI Issues customer. We have to find the correct customer.")
                    row.append("We are reviewing this request to find the correct customer")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif (currntCustLOB in SBG and rqstd_cust_lob in indirect):
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The current cust is in SBG and moving to Indirect - raises eyebrow")
                    row.append(f"Current customer in SBG and Requested customer is Indirect - please review the movement")
                    row.append("REVIEW == Mapped | Feedback Required: We need more information in order to proceed - why is this movement requested? And, how are you sure that the billing account is Indirect vs. Enterprise?")
                    research_list.append(financeAcctNbr)
                    row.insert(0, "Review")
                elif (billAcctLevelType == 'FALLOUT'):
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The current cust is in SBG and moving to Indirect - raises eyebrow")
                    row.append(f"Billing account is listed as 'Fallout' which means it is no longer being pulled into the Source Billing account table; i.e., it is no longer a valid billing account and cannot be mapped")
                    row.append("The billing account's BILL_ACCOUNT_LEVEL_TYP = 'FALLOUT' which means the billing account is no longer a valid billing account. We cannot govern this request.")
                    research_list.append(financeAcctNbr)
                    row.insert(0, "Denied")
                elif re.search(voice_services, billAcctName):
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"This billing account may have Voice Services on it. It can't be moved")
                    row.append(f"This billing account likely has Voice Services attached (according to its name) don't touch it.")
                    row.append("Feedback Required: This billing account may have Voice products attached. Billing accounts with Voice products cannot be mapped from their current customer. Please reach out to Ashley Ellis for more information")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif badOESystemCheck == 'Y':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"This billing account may have Voice Services on it. It can't be moved")
                    row.append(f"This billing account is associated to an Order Entry (OE) system that cannot be disassociated from its current customer.")
                    row.append("Feedback Required: This billing account has products which have ties to the billing account's current customer at an individual service level. Please visit this link for more information: https://bit.ly/4eT7NdH")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                    
                else:
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"not sure 'bout this one -- somethin' ain't right" )
                    row.append(f"We must decide whether requests of this type should be mapped or not - Green loop")
                    row.append("REVIEW")
                    research_list.append(financeAcctNbr)
                    row.insert(0, "Review")
            # print("out of green error list")
                    

        ###### ########################################################## 
        ###### --> am systems
        ###### ##########################################################
        
            """    
        elif row[26] in am_source_systems and ((custDWSecure == 1 or custDWSecure is None or custDWSecure == '') and requestedCustDWSecure == 1 and banDWSecure == 1) \
            and (row[79] == 0 or (row[79] == 1 and row[39] == 1)) and row[78] == 'Approved' and (currentCust != rqstdCust) and ((row[31] in SBG and row[72] in indirect) == False) \
            and ((row[11] != consumer100 and row[29] != consumer100) or (row[11] == damages and row[29] != damages)) and billAcctLevelType != 'FALLOUT' \
            and ((row[26] == 'IDC_KENAN' and billCycleCode == 'K86') == False) and (not re.search(voice_services, billAcctName)) \
            and ((badOESystemCheck == 'Y' and currentCust in placeholder_customers) or (badOESystemCheck == 'Y' and empID in bac_team_empIDs) or badOESystemCheck == 'N') \
            and billStatusIndForKenan == 'Y' and (kenanActvSrvcsAndRevenueCheck == 'N' or kenanActvSrvcsAndRevenueCheck == ''):
            print("\n\nWOOOOOOOOTMUHFUGGA!!!!!", financeAcctNbr + "\n\n")
            """
            

        elif banDWSecure in am_source_systems and ((custDWSecure == 1 or custDWSecure is None or custDWSecure == '') and requestedCustDWSecure == 1 and banDWSecure == 1)\
            and (rqstd_cust_house_accnt_cd == 0 or (rqstd_cust_house_accnt_cd == 1 and currentHouseAcctIdentifier == 1)) and rqstd_cust_status == 'Approved' and (currentCust != rqstdCust) and ((row[31] in SBG and row[76] in indirect) == False) \
            and ((rqstd_cust_nbr != consumer100 and crrnt_cust_nbr != consumer100) or (rqstd_cust_nbr == damages and crrnt_cust_nbr != damages)) and billAcctLevelType != 'FALLOUT' \
            and ((banDWSecure in ['IDC_KENAN', 'KENANFX'] and billCycleCode == 'K86') == False) and (not re.search(voice_services, billAcctName)) \
            and ((badOESystemCheck == 'Y' and currentCust in placeholder_customers) or (badOESystemCheck == 'Y' and empID in bac_team_empIDs) or badOESystemCheck == 'N')\
            and billStatusIndForKenan == 'Y' and (kenanActvSrvcsAndRevenueCheck == 'N' or kenanActvSrvcsAndRevenueCheck == '') and requestedCust != cpniIssuesCust:
            print("\n\nWOOOOOOOOTMUHFUGGA!!!!!", financeAcctNbr + "\n\n")
            # print(f"\n\nBAN in am source systems: {row[14]} and the status is {billStatusIndForKenan} \n\n")
            create_a_list = []
            #begin red list creation
            if ((crrnt_cust_nbr == '' or crrnt_cust_nbr is None) or (rqstd_cust_house_accnt_cd == 0 and crrnt_cust_house_accnt_cd == 1)) and ((billAcctName is not None and rqstd_cust_name is not None) \
                and (billAcctName != '' and rqstd_cust_name !='')): # if cust unassigned or requsted cust is legit cust and crrnt cust is house
                # print("row 29: ", row[29], "row 70", row[70], "row 25", row[25], "row 67", row[67])
                print("mde it to red  name check")
                if (crrnt_cust_nbr == '' or crrnt_cust_nbr is None) and ((billAcctName.upper() == rqstd_cust_name.upper()) or (re.sub(r'[^a-zA-Z0-9]', '', billAcctName.upper()) == re.sub(r'[^a-zA-Z0-9]', '', rqstd_cust_name.upper()) ) ): #r'[^\w\s]'
                    # print("made it here somehow, this BAN: ", row[14])
                    # print("here is row29, ", row[29])
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"{row[14]} ain't done been assigned-ified. Go ahead and map 'im")
                    row.append(f"Billing account is currently unmapped. Billing account name and Requested customer name are perfect match")
                    row.append("Mapped")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(financeAcctNbr)
                    row.insert(0, "Map it!")
                #if current cust has House designation (1) or if the customer is in the Placeholders list and name is match
                elif (currentHouseAcctIdentifier == 1 or crrnt_cust_nbr in placeholder_customers) and ( (billAcctName.upper() == rqstd_cust_name.upper()) or (re.sub(r'\.', '', billAcctName.upper()) == re.sub(r'\.', '', rqstd_cust_name.upper()) ) ):
                    if currentHouseAcctIdentifier == 1 and crrnt_cust_nbr not in placeholder_customers:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"{financeAcctNbr} ain't done been assigned-ified. Go ahead and map 'im")
                        row.append(f"Billing account is mapped to a House customer. Billing account name and Requested customer name are perfect match")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                    elif crrnt_cust_nbr in placeholder_customers:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"{financeAcctNbr} ain't done been assigned-ified. Go ahead and map 'im")
                        row.append(f"Billing account is mapped to a Placeholder customer. Billing account name and Requested customer name are perfect match")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                else:
                    print("\n\n shucks, name of ban and desired cust don't match")
                    if currentHouseAcctIdentifier == 1 and crrnt_cust_nbr not in placeholder_customers:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"{financeAcctNbr} ain't done been assigned-ified. Go ahead and map 'im")
                        row.append(f"Billing account is mapped to a House customer. Billing account name and Requested customer name do not match")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                    elif crrnt_cust_nbr in placeholder_customers:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"{financeAcctNbr} ain't done been assigned-ified. Go ahead and map 'im")
                        row.append(f"Billing account is mapped to a Placeholder customer. Billing account name and Requested customer name do not match")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                    elif (crrnt_cust_nbr == '' or crrnt_cust_nbr is None):
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"{financeAcctNbr} ain't done been assigned-ified. Go ahead and map 'im")
                        row.append(f"Billing account is currently unmapped. Billing account name and Requested customer name do not match")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                '''           
            elif row[29] in placeholder_customers or row[68] == 1:
                if (row[25].upper() == row[67].upper()) or (re.sub(r'\.', '', row[25].upper()) == re.sub(r'\.', '', row[67].upper()) ): #r'[^\w\s]'
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                        create_a_list.append(f"{row[14]} ain't done been assigned-ified. Go ahead and map 'im")
                        row.append(f"BAN is currently unmapped | BAN name and Requested Customer name are perfect match")
                        green_can_map_list.append(create_a_list)
                        maybe_can_map_list.append(row[14])
                else:
                    print("\n\n shucks, they don't match")
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                        create_a_list.append(f"{row[14]} ain't done been assigned-ified. Go ahead and map 'im")
                        row.append(f"BAN is currently unmapped -- can't tell if the BAN name and Requested Customer names match or not")
                        green_can_map_list.append(create_a_list)
                        maybe_can_map_list.append(row[14])
                        
                '''
            elif banAge == 'Under 90':
                print("in the red under 90")
                #create_a_list = [] # clears the list for each row in data_list
                # fan, fan name, ban_age, ban_revenue, ban_lob, ban_channel, to_cust_nbr, to_cust_status, to_cust_lob, to_cust_channel
                # appending multiple indices to the same row is proving to be a challenge, so, I've
                #--> decided to create a row, and then append that. It seems to work like a charm.
                if badOESystemCheck == 'Y' and empID in bac_team_empIDs:
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                    row.append(f"This billing account is associated to an Order Entry (OE) system that cannot be disassociated from its current customer - But the request comes from the BAC team, so hopefully they are already working on it. 'sokay")
                    row.append("Mapped - you are on our exception list, you lucky duck, you.")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(financeAcctNbr)
                    row.insert(0, "Map it!")
                
                else:
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                    row.append(f"The billing account is under 90 days old. Barring anything unforeseen, we should be able to map this one.")
                    row.append("Mapped | Feedback Required")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(financeAcctNbr)
                    row.insert(0, "Review")
                
            elif row[28] == 'Over 90':
                print("\nmade it to Over 90 selection in RED billers: ", financeAcctNbr)
                if banRevFlag == 'No':
                    print("at 1 in reds" , financeAcctNbr)
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{financeAcctNbr} is over 90 days old, no revenue, to-customer is Approved')
                    row.append(f"Billing account is over 90 days old, but it there is no revenue associated with it in the last 13 months, i.e., there is no financial impact")
                    row.append("Mapped")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(financeAcctNbr)
                    row.insert(0, "Map it!")
                elif crrnt_cust_lob == rqstd_cust_lob:
                    print("at 2 in reds" , financeAcctNbr)
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{financeAcctNbr} is over 90 days old, LOB of current customer and requested is {rqstd_cust_lob}, to-customer is Approved')
                    row.append(f"Billing account is over 90 days old, but there is no financial impact (Current and Requested customer LOBs are the same)") 
                    row.append("Mapped")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(financeAcctNbr)
                    row.insert(0, "Map it!")
                elif (crrnt_cust_lob in fed) or (rqstd_cust_lob in fed):
                    print("at 3 in reds" , financeAcctNbr)
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{financeAcctNbr} is over 90 days old, to-customer LOB is FED, to-customer is Approved')
                    if crrnt_cust_lob in fed:
                        row.append(f"Billing account is over 90 days old, but the Current customer is in the Federal channel")
                    elif rqstd_cust_lob in fed:
                        row.append(f"Billing account is over 90 days old, but the Requested Customer is in the Federal channel")
                    row.append("Mapped")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(financeAcctNbr)
                    row.insert(0, "Map it!")
                elif (crrnt_cust_lob in internal) or (rqstd_cust_lob in internal):
                    print("at 4 in reds" , financeAcctNbr)
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{financeAcctNbr} is over 90 days old, to-customer LOB is FED, to-customer is Approved')
                    if crrnt_cust_lob in internal:
                        row.append(f"BAN is in Internal channel")
                    elif rqstd_cust_lob in internal:
                        row.append(f"Billing account is over 90 days old, but the Requested customer is in the Internal channel")
                    row.append("Mapped")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(financeAcctNbr)
                    row.insert(0, "Map it!")
                elif (rqstd_cust_nbr == damages and (crrnt_cust_nbr != damages or crrnt_cust_nbr is None or crrnt_cust_nbr == '')):
                    print("at 5 in reds" , financeAcctNbr)
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{financeAcctNbr} is over 90 days old, LOB of current customer and requested is {rqstd_cust_lob}, to-customer is Approved')
                    row.append(f"The billing account is moving from a legitimate customer to Damages -- why?") 
                    row.append("Feedback Required: Red billing systems cannot be mapped to Damages. Only BART/MBS billing accounts. Why is this move desired?")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(financeAcctNbr)
                    row.insert(0, "Review")
                elif (crrnt_cust_lob not in SBG) and (rqstd_cust_lob not in SBG):
                    print("at 6  in reds" , financeAcctNbr)
                    if crrnt_cust_bu_or_ext_sales_channel == rqstd_cust_bu_or_ext_sales_channel:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Current and Requested customer LOBs are both outside SBG; i.e., there is no financial impact to MASS MARKETS (SBG) and they are both under the same Channel, so there is 0 financial impact overall")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                    else:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Current and Requested customer LOBs are both outside SBG; i.e., there is no financial impact to MASS MARKETS (SBG)")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                elif ((crrnt_cust_lob in SBG) and (rqstd_cust_lob not in SBG)) and (banRevFlag == 'Yes'):
                    print("at sbg to outside sbg in reds" , financeAcctNbr)
                    if row[1] == '00096184': #00302684 - matty t # jacquie 00096184
                        print("at 7 in reds" , financeAcctNbr)
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is over 90 days old, there is revenue attached, and the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel} - but the request comes from Jacquie, so it is likely okay")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                    elif row[1] == '00302684':
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is over 90 days old, and there is revenue attached. But the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel} and the request comes from Matty T, so it is likely okay")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                    elif int(ban_3_month_avg) < 2000:
                        print("at 8 in reds" , financeAcctNbr)
                        if crrnt_cust_nbr in placeholder_customers and rqstd_cust_nbr not in placeholder_customers:
                            for num in green_field_nums:
                                create_a_list.append(row[int(num)])
                            create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                            row.append(f"Billing account is over 90 days old, there is revenue attached, and the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel} - but the revenue is less than $2k. And it is moving FROM a Placeholder to a legit customer. Warrants a look.")
                            row.append("Mapped")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(financeAcctNbr)
                            row.insert(0, "Map it!")
                        elif crrnt_cust_nbr not in placeholder_customers and rqstd_cust_nbr not in placeholder_customers:
                            for num in green_field_nums:
                                create_a_list.append(row[int(num)])
                            create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                            row.append(f"Billing account is over 90 days old, there is revenue attached, and the movement is cross-BU (from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel}) - but the revenue is less than $2k. Warrants a look.")
                            row.append(f"REVIEW == Mapped | Feedback Required: The movement is Cross-BU - out of {crrnt_cust_bu_or_ext_sales_channel} and heading to {rqstd_cust_bu_or_ext_sales_channel} and there is revenue attached -- why is this movement necessary?")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(financeAcctNbr)
                            row.insert(0, "Review")
                        elif crrnt_cust_nbr not in placeholder_customers and rqstd_cust_nbr in placeholder_customers:
                            for num in green_field_nums:
                                create_a_list.append(row[int(num)])
                            create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                            row.append(f"Billing account is over 90 days old, there is revenue attached, and the movement is cross-BU to a Placeholder customer. Nah. ")
                            row.append(f"Feedback Required: The movement is Cross-BU - from {crrnt_cust_bu_or_ext_sales_channel} to a Placeholder customer in ({rqstd_cust_bu_or_ext_sales_channel}). Please work with Sales, Sales Ops, or your peers to find or create the desired customer.")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(financeAcctNbr)
                            row.insert(0, "Review")
                        elif crrnt_cust_nbr in placeholder_customers and rqstd_cust_nbr in placeholder_customers:
                            for num in green_field_nums:
                                create_a_list.append(row[int(num)])
                            create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                            row.append(f"Billing account is over 90 days old, there is revenue attached, and the movement is cross-BU from a Placeholder to a Placeholder customer. Nah. ")
                            row.append(f"Feedback Required: The movement is Cross-BU - from an {crrnt_cust_bu_or_ext_sales_channel} Placeholder to a Placeholder in ({rqstd_cust_bu_or_ext_sales_channel}). Please work with Sales, Sales Ops, or your peers to find or create the desired customer.")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(financeAcctNbr)
                            row.insert(0, "Review")
                        else:
                            for num in green_field_nums:
                                create_a_list.append(row[int(num)])
                            create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                            row.append(f"Not sure why this request made it here. Dave needs to review")
                            row.append("REVIEW == Mapped | Feedback Required ")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(financeAcctNbr)
                            row.insert(0, "Review")
                    elif int(ban_3_month_avg) < 2000:
                        print("at 9 in reds " , financeAcctNbr)
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is over 90 days old, there is revenue attached, and the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel} - but the revenue is less than $2k. Warrants a look.")
                        row.append("REVIEW == Mapped | Feedback Required: The movement is cross-BU, the billing account is over 90 days old and there is revenue attached. Why is this movement desired?")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Review")
                    elif int(ban_3_month_avg) > 2000:
                        print("at 10 in reds " , financeAcctNbr)
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is over 90 days old, there is revenue attached, the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel}, and the revenue is over $2k. Maybe not okay?")
                        row.append("REVIEW == Mapped | Feedback Required: The movement is cross-BU, and the amount of revenue attached is over $2k, we need more information before we can proceed with this request")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Review")
                elif (crrnt_cust_lob not in SBG) and (rqstd_cust_lob in SBG) and (ban_3_month_avg == 'Yes'):
                    if row[1] == '00096184': #00302684 - matty t # jacquie 00096184
                        print("at 11 in reds" , financeAcctNbr)
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is over 90 days old, there is revenue attached, and the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel} - but the request is from Jackie, so it is likely okay")
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                    elif int(ban_3_month_avg) < 2000:
                        print("at 12 in reds" , financeAcctNbr)
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is over 90 days old, there is revenue attached, and the movement is from {crrnt_cust_bu_or_ext_sales_channel} to {rqstd_cust_bu_or_ext_sales_channel} - but the revenue is less than $2k. Warrants a look.")
                        row.append("REVIEW == Mapped | Feedback Required: the movement is cross-BU, we need more information before we can proceed with this request")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Review")
                    elif int(ban_3_month_avg) > 2000:
                        print("at 13 in reds" , financeAcctNbr)
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Billing account is over 90 days old, there is revenue attached, the movement is from {rqstd_cust_bu_or_ext_sales_channel} to SBG, and the revenue is over $2k. Maybe not okay?")
                        row.append("REVIEW == Mapped | Feedback Required: the movement is cross-BU and the amount of revenue that would change channels is over $2k - why is this movement requested?")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Review")
                elif (crrnt_cust_lob in SBG) and (rqstd_cust_lob in SBG):
                    print("both in SBG check in red" , financeAcctNbr)
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                    row.append(f"Billing account is only moving internally, within SBG, so hey, no biggie here - let's do it.")
                    row.append("Mapped")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(financeAcctNbr)
                    row.insert(0, "Map it!")
                elif badOESystemCheck == 'Y' and empID in bac_team_empIDs:
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the Current and Requested customer LOBs are outside SBG, to-customer is Approved')
                    row.append(f"This billing account is associated to an Order Entry (OE) system that cannot be disassociated from its current customer - But the request comes from the BAC team, so hopefully they are already working on it. 'sokay")
                    row.append("Mapped - you are on our exception list, you lucky duck, you.")
                    green_can_map_list.append(create_a_list)
                    can_map_bans_list.append(financeAcctNbr)
                    row.insert(0, "Map it!")
                print("\nmade it OUT of RED Over 90 selection: ", financeAcctNbr)
            
            elif (currentHouseAcctIdentifier== 1 and rqstd_cust_house_accnt_cd == 1): # current and requested customer are house accounts
                print("in the rest placeholder check")
                for num in green_field_nums:
                    create_a_list.append(row[int(num)])
                create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                row.append(f"Current customer and Requested customer are both House acounts - could be okay, but let's review the movement")
                row.append("REVIEW == Mapped | Feedback Required: please work with Sales, Sales Ops, or your peers to either find or create a customer for this billing account.")
                green_can_map_list.append(create_a_list)
                can_map_bans_list.append(financeAcctNbr)
                row.insert(0, "Review")
                #print("leaving the red biller section to errors")

            ###########################    
            #### --> error sheet for Red billers ############################################################
            ###########################
            else:
                print("\n\nin red error file", row[14])
                create_a_list = []
                error_book = []
                # append to row and change to elif statements to change back (if it is fucked)
                ## --> and this " cannot_map_list.append(create_a_list) "
                if rqstd_cust_status != 'Approved' or rqstd_cust_status == '':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"Customer # {rqstd_cust_nbr} is {rqstd_cust_status} so {financeAcctNbr} can't be mapped to it" )
                    row.append(f"Requested customer is not Approved")
                    row.append("Feedback Required: the Requested customer is not Approved and we cannot map billing accounts to customers in any other status but 'Approved'. Please work with Sales, Sales Ops, your peers (or leadership) to identify or create the desired customer for this billing account. Then update the ticket. ")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif rqstd_cust_nbr == crrnt_cust_nbr: # if currnt customer == rqstd customer
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The desired customer's Secure Company Number is not 1 (Lumen)... so... there's that")
                    row.append(f"Current customer and Requested customer are the same")
                    row.append("Validated - No Action Required: The billing account is already mapped as desired")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Mark ticket: Closed - Resolved, Validated...")
                elif rqstd_cust_nbr in placeholder_customers:
                    if crrnt_cust_nbr == '' or crrnt_cust_nbr is None:
                        for num in error_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"Requested customer is a Placeholder" )
                        row.append(f"Requested customer is a Placeholder account, but the customer is currently unmapped")
                        row.append("Mapped")
                        cannot_map_list.append(create_a_list)
                        row.insert(0, "Map it!")
                    else:
                        for num in error_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"Requested customer is a Placeholder" )
                        row.append(f"Requested customer is a Placeholder account")
                        row.append("Feedback Required: We cannot map billing accounts to Placeholder customers. Please work with Sales, Sales Ops, your peers, or leadership to identify or create the desired customer for this billing account. Then update the ticket.")
                        cannot_map_list.append(create_a_list)
                        row.insert(0, "Review")
                    ''' # removing Cross-BU problem
                if (row[31] in SBG and row[63] not in SBG) or (row[31] not in SBG and row[63] in SBG) and row[16] == 'Yes':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    #print(f"{row[14]} done made it to the SBG identfier in Error file")
                    create_a_list.append(f"BAN and desired customer have cross-BU LOB's" )
                    row.append(f"Movement is cross-BU")
                    cannot_map_list.append(create_a_list)
                    '''
                elif billingSystem == 'ORAFIN':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"{financeAcctNbr} is an Orafin build. It's a fallout build and can't be mapped for that reason" )
                    row.append(f"BAN is an ORAFIN billing account - we do not govern ORAFIN billing accounts.")
                    row.append("Feedback Required: Customer Hierarchy does not govern ORAFIN billing accounts. Did you intend a different billing account than the one provided?")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif billingSystem == 'BTN':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"""{financeAcctNbr} is a Fallout billing account. It should have a CRIS brother somewhere..
                                            It isn't a real billing account and cannot be mapped for that reason""" )
                    row.append(f"BAN is a BTN billing account - we do not govern BTN billing accounts.")
                    row.append(f"Feedback Required: Customer Hierarchy does not govern BTN billing accounts. Please visit this link for more information: {howToFindCRISBans}, then update this ticket once you've identified the correct, valid CRIS Finance Account Number. ")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif billingSystem == 'NFCN':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"""{financeAcctNbr} is a Fallout billing account. It should have a CRIS brother somewhere..
                                            It isn't a real billing account and cannot be mapped for that reason""" )
                    row.append(f"BAN is an NFCN billing account - we do not govern NFCN billing accounts.")
                    row.append(f"Feedback Required: Customer Hierarchy does not govern NFCN billing accounts. Please visit this link for more information: {howToFindCRISBans}, then update this ticket once you've identified the correct, valid CRIS Finance Account Number.")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif billingSystem == 'ZUORA':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"""{financeAcctNbr} is a Fallout billing account. It should have a CRIS brother somewhere..
                                            It isn't a real billing account and cannot be mapped for that reason""" )
                    row.append(f"BAN is a ZUORA billing account -- we don't govern these. Traci Graham might")
                    row.append("Feedback Required: Customer Hierarchy does not govern ZUORA billing accounts. Did you intend a different billing account?")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                    '''
                elif row[37] != 1:
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The current customer's DW Secure Company Number is not 1 (Lumen)... so... there's that")
                    row.append("Current customer DW Secure Company Number is not 1")
                    '''
                elif (banDWSecure != 1 and banDWSecure is not None) or requestedCustDWSecure != 1: # --> DW Secure company code is not Lumen for either the BAN or requested Customer
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"{row[14]}'s DW Secure Company Number is not 1 (Lumen)... so... there's that")
                    if banDWSecure!= 1 and requestedCustDWSecure == 1:
                        if banDWSecure == 4:
                            row.append("Billing account is Brightspeed, Requested customer is Lumen")
                            row.append("Denied - We do not govern Brightspeed billing accounts")
                        elif banDWSecure == 2:
                            row.append("Billing account is Cirion, Requested customer is Lumen")
                            row.append("Denied - We do not govern Cirion billing accounts")
                        elif banDWSecure == 6:
                            row.append("Billing account is Colt, Requested customer is Lumen")
                            row.append("Denied - We do not govern Colt billing accounts")
                    elif banDWSecure == 1 and requestedCustDWSecure != 1:
                        if requestedCustDWSecure == 4:
                            row.append("Requested customer is Brightspeed, billing account is Lumen")
                            row.append("Denied - We do not map Lumen billing accounts to Brightspeed customers")
                        elif requestedCustDWSecure == 2:
                            row.append("Requested customer is Cirion, billing account is Lumen")
                            row.append("Denied - We do not map Lumen billing accounts to Cirion customers")
                        elif requestedCustDWSecure == 6:
                            row.append("Requested Customer is Colt, billing account is Lumen")
                            row.append("Denied - We do not map Lumen billing accounts to Colt customers")
                    elif banDWSecure != 1 and requestedCustDWSecure != 1:
                        if banDWSecure == 4:
                            if requestedCustDWSecure == 4:
                                row.append("Billing account and Requested customer are Brightspeed")
                                row.append("Denied - we do not govern Brightspeed accounts")
                            elif requestedCustDWSecure == 2:
                                row.append("Billing account is Brightspeed, Requested customer is Cirion")
                                row.append("Denied - We do not govern Brightspeed billing accounts")
                            elif requestedCustDWSecure == 6:
                                row.append("Billing account is Brightspeed, Requested customer is Colt")
                                row.append("Denied - We do not govern Brightspeed billing accounts")
                        elif banDWSecure == 2:
                            if requestedCustDWSecure == 2:
                                row.append("Billing account and Requested customer are Cirion")
                                row.append("Denied - we do not govern Cirion accounts")
                            elif requestedCustDWSecure == 4:
                                row.append("Billing account is Cirion, Requested customer is Brightspeed")
                                row.append("Denied - We do not govern Cirion billing accounts")
                            elif requestedCustDWSecure == 6:
                                row.append("Billing account is Cirion, Requested customer is Colt")
                                row.append("Denied - We do not govern Cirion billing accounts")
                        elif banDWSecure == 6:
                            if requestedCustDWSecure == 2:
                                row.append("Billing account is Colt and Requested customer is Cirion")
                                row.append("Denied - we do not govern Colt billing accounts")
                            elif requestedCustDWSecure == 4:
                                row.append("Billing account is Colt, Requested customer is Brightspeed")
                                row.append("Denied - We do not govern Colt billing accounts")
                            elif requestedCustDWSecure == 6:
                                row.append("Billing account and Requested customer are Colt")
                                row.append("Denied - We do not govern Colt accounts")
                    else:
                        row.append("Either the BAN or the customer is not Lumen -- needs to be Reviewed to fix")
                        row.append("Customer Hierarchy does not govern Cirion, Brightspeed, or Colt billing accounts (or customers). Please work with the Cirion, Brightspeed, or Colt teams instead. ")
                    cannot_map_list.append(row[14])
                    row.insert(0, "Denied")
                elif rqstd_cust_house_accnt_cd == 1 and crrnt_cust_house_accnt_cd == 0:
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The desired customer's Secure Company Number is not 1 (Lumen)... so... there's that")
                    row.append(f"Current customer is a valid, Enterprise customer. Requested customer is a House or Placeholder")
                    row.append("Feedback Required: We do not map billing accounts from valid Enterprise customers to Placeholders or House accounts. Please work with Sales, Sales Ops, your peers, or leadership to find or create the desired customer, then update this ticket.")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif requestedCust == cpniIssuesCust:
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The desired customer's Secure Company Number is not 1 (Lumen)... so... there's that")
                    row.append(f"Requested customer is the CPNI Issues customer.  We have to find the correct customer.")
                    row.append("We are reviewing this request to find the correct customer")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif (crrnt_cust_lob in SBG and rqstd_cust_lob in indirect):
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The current cust is in SBG and moving to Indirect - raises eyebrow")
                    row.append(f"Current customer in SBG and Requested customer is Indirect - please review the movement")
                    row.append("REVIEW == Mapped | Feedback Required: Please provide more evidence supporting this request. How do you know that this billing belongs on this specific customer in Indirect?")
                    research_list.append(row[14])
                    row.insert(0, "Review")
                elif (billAcctLevelType == 'FALLOUT'):
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The current cust is in SBG and moving to Indirect - raises eyebrow")
                    row.append(f"Billing account is listed as 'Fallout' which means it is no longer being pulled into the Source Billing account table; i.e., it is no longer a valid billing account")
                    row.append("The billing account's BILL_ACCOUNT_LEVEL_TYP = 'FALLOUT' which means the billing account is no longer a valid billing account. We cannot govern this request." )
                    research_list.append(row[14])
                    row.insert(0, "Denied")
                elif (billingSystem in ['IDC_KENAN', 'KENANFX'] and billCycleCode == 'K86'):
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The current cust is in SBG and moving to Indirect - raises eyebrow")
                    if re.search(voice_services, billAcctName):
                        row.append(f"This billing account may have Voice Services attached (according to its suffix) and its Bill Cycle Code is K86. If this don't have Voice Services, I don't know what done do.")
                        row.append("Feedback Required: This billing account very likely has Voice products attached. For that reason, it cannot be moved from its current customer. Please reach out to Ashley Ellis to confirm the Voice products on this billing account.")
                        cannot_map_list.append(create_a_list)
                        row.insert(0, "Nope")
                    elif not re.search(voice_services, billAcctName):
                        row.append(f"This is an IDC KENAN acct w/ a K86 Bill Cycle - it is possible this has voice services attached, but the name doesn't feature any of the known Voice Suffixes - probably fine")
                        row.append("Mapped")
                        research_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                elif re.search(voice_services, billAcctName):
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"This billing account may have Voice Services on it. It can't be moved")
                    row.append(f"This billing account may have Voice Services attached (according to its suffix) don't touch it.")
                    row.append("Feedback Required: This billing account may have Voice Products attached. It cannot be moved from its current customer for that reason. Please reach out to Ashley Ellis for conformation of the Voice products attached to this billing account.")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif billStatusIndForKenan == 'N':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"This billing account is Inactive")
                    row.append(f"This billing account is Inactive in AM, so we are unable to map it.")
                    row.append("Feedback Required: This billing account is Inactive in AM, so we are unable to map it. Please work with the KIP team (Kenan.InvoiceProduction@lumen.com) to Activated the billing account. Then notify us here once it has been activated.")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif kenanActvSrvcsAndRevenueCheck == 'Y':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"This billing account is Inactive")
                    row.append(f"This is a Kenan biller with no Revenue and Active Services -- there's a good chance it has Voice services, Ashley may need to look at it")
                    row.append("REVIEW == Mapped | Feedback Required: This is a Kenan billing account with no revenue and active services, it needs approval from Ashley Ellis before it can move ")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif badOESystemCheck == 'Y':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"This billing account may have Voice Services on it. It can't be moved")
                    row.append(f"This billing account is associated to an Order Entry (OE) system that cannot be disassociated from its current customer.")
                    row.append("Feedback Required: This billing account has products which have ties to the billing account's current customer at an individual service level. Please visit this link for more information: https://bit.ly/4eT7NdH")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                else:
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"not sure 'bout this one -- somethin' ain't right" )
                    row.append(f"We must decide whether requests of this type should be mapped or not - Green loop")
                    row.append("REVIEW")
                    research_list.append(row[14])
                    row.insert(0, "Review")
                    
        # ---- ===========================================
        #---->  begin error sheet prep
        # ---- ===========================================
        else:
            # print("\n\nYASOG SHAGOTH -- at the error sheet prep", financeAcctNbr, billAcctLevelType, "\n")
            create_a_list = []
            if re.search(voice_services, billAcctName):
                for num in error_field_nums:
                    create_a_list.append(row[int(num)])
                create_a_list.append(f"This billing account may have Voice Services on it. It can't be moved")
                row.append(f"This billing account may have Voice Services attached (according to its name) -- don't touch it.")
                row.append("Feedback Required: This billing account may have Voice products attached. It needs approval from Ashley Ellis before it can be moved.")
                cannot_map_list.append(create_a_list)
                row.insert(0, "Review")
            else:
                #print("\nnot in voice services check in error sheet, so moving on to the rest", financeAcctNbr, "\n")
                if rqstd_cust_status != 'Approved' or rqstd_cust_status == '':
                    if requestedCust == cpniIssuesCust:
                        for num in error_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"The desired customer's Secure Company Number is not 1 (Lumen)... so... there's that")
                        row.append(f"Requested customer is the CPNI Issues customer. We have to find the correct customer.")
                        row.append("We are reviewing this request to find the correct customer")
                        cannot_map_list.append(create_a_list)
                        row.insert(0, "Review")
                    #print("at 1 error" , row[14])
                    else:
                        for num in error_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"Customer # {rqstd_cust_nbr} is {rqstd_cust_status} so {financeAcctNbr} can't be mapped to it" )
                        row.append(f"Requested Customer is not Approved")
                        row.append("Feedback Required: the Requested customer is not Approved. Please work with Sales, Sales Ops, your peers, or leadership to find or create the desired customer, then update this ticket.")
                        cannot_map_list.append(create_a_list)
                        row.insert(0, "Review")
                    '''
                elif (row[31] in SBG and row[63] not in SBG) or (row[31] not in SBG and row[63] in SBG) and row[16] == 'Yes':
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    #print(f"{row[14]} done made it to the SBG identfier in Error file")
                    create_a_list.append(f"BAN and desired customer have cross-BU LOB's" )
                    row.append(f"Movement is cross-BU")
                    cannot_map_list.append(create_a_list)
                    '''
                elif (rqstd_cust_nbr == damages and (crrnt_cust_nbr != damages or crrnt_cust_nbr is None or crrnt_cust_nbr == '')):
                    #print("at 2 error" , financeAcctNbr)
                    for num in green_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f'{financeAcctNbr} is over 90 days old, LOB of current customer and requested is {rqstd_cust_lob}, to-customer is Approved')
                    if crrnt_cust_nbr is not None and crrnt_cust_nbr != damages:
                        row.append(f"The billing account is moving from a legitimate customer to Damages -- why?") 
                        row.append("REVIEW == Mapped | Feedback Required: the billing account is mapped to a customer that appears accurate - why does it need to move to Damages?")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Review")
                    elif crrnt_cust_nbr is None or crrnt_cust_nbr == '':
                        row.append(f"The billing account is unmapped and heading to Damages -- not necessarily an issue. Let's do it.") 
                        row.append("Mapped")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                elif crrnt_cust_nbr == damages: #row11 == rqstd row29 == current
                    #print("at 3 error" , financeAcctNbr)
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The desired customer's Secure Company Number is not 1 (Lumen)... so... there's that")
                    row.append(f"Current customer is Damages -- why do they want to move it out?")
                    row.append("REVIEW == Mapped | Feedback Required: This billing account is moving from Damages, if it doesn't belong there, please explain why.")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif (rqstd_cust_nbr == consumer100 or crrnt_cust_nbr == consumer100): ## if the billing account is moving to or from Consumer 100 (regardless of whether the current customer is a placeholder/house)
                    #print("at 4 error" , financeAcctNbr)
                    if crrnt_cust_nbr == consumer100:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Current customer is Consumer 100 -- why is the account moving out of Consumer 100? What is -- as they say -- 'The dilly-o?' Do they say that still? Did they ever?")
                        row.append("REVIEW == Mapped | Feedback Required: Please provide more information. What evidence do you have that this billing account is not a Consumer account?")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Review")
                    elif rqstd_cust_nbr == consumer100:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        if empID == '00096221':
                            if re.search('^.*(-CRISE|-CRISW|-CRISC)$', financeAcctNbr): # steve isaac
                                row.append(f"Requested customer is Consumer 100 -- but the billing account is a CRIS account, with an MCN ID that identifies a Consumer account. And the request comes from Steve Isaac. This one should be in Consumer.")
                                row.append("Mapped")
                                green_can_map_list.append(create_a_list)
                                can_map_bans_list.append(financeAcctNbr)
                                row.insert(0, "Map it!")
                            else:
                                row.append(f"Requested customer is Consumer 100, and the billing account is not a CRIS account, but the request comes from Steve Isaac. Let's take a peek, but it is likely okay.")
                                row.append("REVIEW == Mapped | Feedback Required: Please provide evidence that supports moving this billing account to Consumer 100") 
                                green_can_map_list.append(create_a_list)
                                can_map_bans_list.append(financeAcctNbr)
                                row.insert(0, "Review")
                        else:
                            row.append(f"Requested customer is Consumer 100 and the current")
                            row.append("REVIEW == Mapped | Feedback Required: Please provide more information. What evidence do you have that this billing account is a Consumer account?")
                            green_can_map_list.append(create_a_list)
                            can_map_bans_list.append(financeAcctNbr)
                            row.insert(0, "Review")
                    elif crrnt_cust_nbr == consumer100 and rqstd_cust_nbr == consumer100:
                        for num in green_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f'{financeAcctNbr} is over 90 days old, both the current LOB and the to-customer LOB are outside SBG, to-customer is Approved')
                        row.append(f"Requested customer is Consumer 100 and the current customer is Consumer 100 -- the billing account is mapped as desired")
                        row.append("Validated - No Action Required: The billing account is mapped as desired")
                        green_can_map_list.append(create_a_list)
                        can_map_bans_list.append(financeAcctNbr)
                        row.insert(0, "Review")
                    else:
                        pass
                
                elif billingSystem == 'ORAFIN':
                    #print("at orafin error" , financeAcctNbr)
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"{financeAcctNbr} is an Orafin build. It's a fallout build and can't be mapped for that reason" )
                    row.append(f"BAN is an Orafin billing account - we do not govern Orafin billing accounts.")
                    row.append("Feedback Required: We do not govern ORAFIN billing accounts. Did you intend to provide a different billing account?")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif billingSystem == 'BTN':
                    #print("at btn error" , financeAcctNbr)
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"""{financeAcctNbr} is a Fallout billing account. It should have a CRIS brother somewhere..
                                                It isn't a real billing account and cannot be mapped for that reason""" )
                    row.append(f"BAN is a BTN billing account - we do not govern BTN billing accounts.")
                    row.append(f"Feedback Required: We do not govern BTN billing accounts. Please visit this link for more information: {howToFindCRISBans}, then update this ticket once you've identified the correct, valid CRIS Finance Account Number")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif billingSystem == 'NFCN':
                    #print("at nfcn error" , financeAcctNbr)
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"""{financeAcctNbr} is a Fallout billing account. It should have a CRIS brother somewhere..
                                            It isn't a real billing account and cannot be mapped for that reason""" )
                    row.append(f"BAN is an NFCN billing account - we do not govern NFCN billing accounts.")
                    row.append(f"Feedback Required: We do not govern NFCN billing accounts. Please visit this link for more information: {howToFindCRISBans}, then update this ticket once you've identified the correct, valid CRIS Finance Account Number")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif rqstd_cust_nbr == crrnt_cust_nbr: # if currnt customer == rqstd customer
                    #print("at 5 error" , financeAcctNbr)
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The desired customer's Secure Company Number is not 1 (Lumen)... so... there's that")
                    row.append(f"Current customer and Requested customer are the same")
                    row.append("Validated - No Action Required: The billing account is already mapped as desired")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Mark ticket: Closed - Resolved, Validated...")
                elif rqstd_cust_nbr in placeholder_customers and (banDWSecure == 1 and requestedCustDWSecure == 1): #56 and 77 are the DW code indicators for currnt and rqstd customer respectively
                    #print("at 6 error" , financeAcctNbr)
                    if crrnt_cust_nbr == '' or crrnt_cust_nbr is None:
                        for num in error_field_nums:
                            create_a_list.append(row[int(num)])
                        create_a_list.append(f"The request is to map {financeAcctNbr} to a House Customer. Which cannot happen!")
                        row.append(f"There is no customer assigned in CODS.BILLING_ACCOUNT, but the requested customer is a Placeholder, which we should have resolved during our cleanup")
                        row.append("Feedback Required: Due to ERP there are only so many BANs that any customer, including Placeholders, can have mapped to them. We have already completed clean-up projects to map what we can to an approved customer or to a placeholder. Please work with Sales or Sales Ops to find or create the desired customer, then update this ticket.")
                        cannot_map_list.append(create_a_list)
                        row.insert(0, "Review")               
                    else:
                        if crrnt_cust_lob in SBG and rqstd_cust_lob in SBG and empID == '00096184' :
                            for num in error_field_nums:
                                create_a_list.append(row[int(num)])
                            create_a_list.append(f"The request is to map {financeAcctNbr} to a House Customer. Which cannot happen!")
                            row.append(f"The movement is internal to SBG and the request is from Jacquie. Let's gooooooooo!!")
                            row.append("Mapped")
                            cannot_map_list.append(create_a_list)
                            row.insert(0, "Map it!")
                        elif crrnt_cust_lob in SBG and rqstd_cust_lob in SBG:
                            for num in error_field_nums:
                                create_a_list.append(row[int(num)])
                            create_a_list.append(f"The request is to map {financeAcctNbr} to a House Customer. Which cannot happen!")
                            row.append(f"The movement is internal to SBG, so, even though it IS crossing LOBs, since it is sticking in SBG specifically, then this should be fine")
                            row.append("Mapped")
                            cannot_map_list.append(create_a_list)
                            row.insert(0, "Map it!")
                        elif (crrnt_cust_lob not in SBG and rqstd_cust_lob in SBG): # if current cust lob outside SBG and requested cust LOB inside sbg
                            for num in error_field_nums:
                                create_a_list.append(row[int(num)])
                            create_a_list.append(f"The request is to map {financeAcctNbr} to a House Customer. Which cannot happen!")
                            row.append(f"The Current customer is outside SBG, and the Requested customer is an SBG Placeholder -- so not only is the movement cross-BU, but it is from a legit customer to a Placeholder. No thank you. ")
                            row.append(f"Feedback Required: the Current customer is in the {currntCustChannel} Channel (outside SBG) and the Requested customer is a Placeholder account inside SBG. Please work with Sales, Sales Ops, your peers, or leadership to find or create the desired customer, then update this ticket.")
                            cannot_map_list.append(create_a_list)
                            row.insert(0, "Review")
                        else:
                            for num in error_field_nums:
                                create_a_list.append(row[int(num)])
                            create_a_list.append(f"The request is to map {financeAcctNbr} to a House Customer. Which cannot happen!")
                            row.append(f"Requested customer is a Placeholder account")
                            row.append(f"Feedback Required: We do not map billing accounts from a valid, Enterprise customer to a House or Placeholder customer. Please work with Sales, Sales Ops, your peers, or leadership to find or create the customer for this billing account")
                            cannot_map_list.append(create_a_list)
                            row.insert(0, "Review")
                    '''
                elif row[37] != 1:
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The current customer's DW Secure Company Number is not 1 (Lumen)... so... there's that")
                    row.append("Current customer DW Secure Company Number is not 1")
                    '''
                elif (banDWSecure != 1 and banDWSecure is not None) or requestedCustDWSecure != 1: # --> DW Secure company code is not Lumen for either the BAN or Customer
                    #print("at 7 error" , financeAcctNbr)
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"{financeAcctNbr}'s DW Secure Company Number is not 1 (Lumen)... so... there's that")
                    if banDWSecure != 1 and requestedCustDWSecure == 1:
                        if banDWSecure == 4:
                            row.append("Billing account is Brightspeed, Requested customer is Lumen")
                            row.append("Denied - We do not govern Brightspeed billing accounts")
                        elif banDWSecure == 2:
                            row.append("Billing account is Cirion, Requested customer is Lumen")
                            row.append("Denied - We do not govern Cirion billing accounts")
                        elif banDWSecure == 6:
                            row.append("Billing account is Colt, Requested customer is Lumen")
                            row.append("Denied - We do not govern Colt billing accounts")
                    elif banDWSecure == 1 and requestedCustDWSecure != 1:
                        if requestedCustDWSecure == 4:
                            row.append("Requested customer is Brightspeed, billing account is Lumen")
                            row.append("Denied - We do not map Lumen billing accounts to Brightspeed customers")
                        elif requestedCustDWSecure == 2:
                            row.append("Requested customer is Cirion, billing account is Lumen")
                            row.append("Denied - We do not map Lumen billing accounts to Cirion customers")
                        elif requestedCustDWSecure == 6:
                            row.append("Requested Customer is Colt, billing account is Lumen")
                            row.append("Denied - We do not map Lumen billing accounts to Colt customers")
                    elif banDWSecure != 1 and requestedCustDWSecure != 1:
                        if banDWSecure == 4:
                            if requestedCustDWSecure == 4:
                                row.append("Billing account and Requested customer are Brightspeed")
                                row.append("Denied - we do not govern Brightspeed accounts")
                            elif requestedCustDWSecure == 2:
                                row.append("Billing account is Brightspeed, Requested customer is Cirion")
                                row.append("Denied - We do not govern Brightspeed billing accounts")
                            elif requestedCustDWSecure == 6:
                                row.append("Billing account is Brightspeed, Requested customer is Colt")
                                row.append("Denied - We do not govern Brightspeed billing accounts")
                        elif banDWSecure == 2:
                            if requestedCustDWSecure == 2:
                                row.append("Billing account and Requested customer are Cirion")
                                row.append("Denied - we do not govern Cirion accounts")
                            elif requestedCustDWSecure == 4:
                                row.append("Billing account is Cirion, Requested customer is Brightspeed")
                                row.append("Denied - We do not govern Cirion billing accounts")
                            elif requestedCustDWSecure == 6:
                                row.append("Billing account is Cirion, Requested customer is Colt")
                                row.append("Denied - We do not govern Cirion billing accounts")
                        elif banDWSecure == 6:
                            if requestedCustDWSecure == 2:
                                row.append("Billing account is Colt and Requested customer is Cirion")
                                row.append("Denied - we do not govern Colt billing accounts")
                            elif requestedCustDWSecure == 4:
                                row.append("Billing account is Colt, Requested customer is Brightspeed")
                                row.append("Denied - We do not govern Colt billing accounts")
                            elif requestedCustDWSecure == 6:
                                row.append("Billing account and Requested customer are Colt")
                                row.append("Denied - We do not govern Colt accounts")
                    else:
                        row.append("Either the BAN or the customer is not Lumen -- needs to be Reviewed to fix")
                        row.append("Customer Hierarchy does not govern Cirion, Brightspeed, or Colt billing accounts (or customers). Please work with the Cirion, Brightspeed, or Colt teams instead. ")
                    cannot_map_list.append(financeAcctNbr)
                    row.insert(0, "Denied")
                elif rqstd_cust_house_accnt_cd == 1 and crrnt_cust_house_accnt_cd == 0:
                    #print("at 8 error" , financeAcctNbr)
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The desired customer's Secure Company Number is not 1 (Lumen)... so... there's that")
                    row.append(f"Current customer is a valid, Enterprise customer. Requested customer is a House or Placeholder")
                    row.append("Feedback Required: We do not map billing accounts from a valid, Enterprise customer to a House or Placeholder customer. Please work with Sales, Sales Ops, your peers, or leadership to find or create the customer for this billing account")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif (crrnt_cust_lob in SBG and rqstd_cust_lob in indirect):
                    #print("at 9 error" , financeAcctNbr)
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The current cust is in SBG and moving to Indirect - raises eyebrow")
                    row.append(f"Current customer in SBG and Requested customer is Indirect")
                    row.append("REVIEW  == Mapped | Feedback Required: Please provide more information. What evidence do you have that this billing account belongs on this particular Indirect customer?")
                    research_list.append(financeAcctNbr)
                    row.insert(0, "Review")
                elif (billAcctLevelType == 'FALLOUT'):
                    #print("at billacctleveltype error" , financeAcctNbr, billAcctLevelType)
                    #print("\nfound it in billaccount level type check in error section", financeAcctNbr, billAcctLevelType)
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The current cust is in SBG and moving to Indirect - raises eyebrow")
                    row.append(f"Billing account is listed as 'Fallout' which means it is no longer being pulled into the Source Billing account table; i.e., it is no longer a valid billing account")
                    row.append("The billing account's BILL_ACCOUNT_LEVEL_TYP = 'FALLOUT' which means the billing account is no longer a valid billing account. We cannot govern this request.")
                    research_list.append(financeAcctNbr)
                    row.insert(0, "Denied")
                elif badOESystemCheck == 'Y':
                    #print("at 10 error" , financeAcctNbr)
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"This billing account may have Voice Services on it. It can't be moved")
                    row.append(f"This billing account is associated to an Order Entry (OE) system that cannot be disassociated from its current customer.")
                    row.append("Feedback Required: This billing account has products which have ties to the billing account's current customer at an individual service level. Please visit this link for more information: https://bit.ly/4eT7NdH")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif billStatusIndForKenan == 'N':
                    #print("at 11 error" , financeAcctNbr)
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"This billing account is Inactive")
                    row.append(f"This billing account is Inactive in AM, so we are unable to map it.")
                    row.append("Feedback Required: The billing account is Inactive in AM, so we are unable to map it, please work with the KIP team (Kenan.InvoiceProduction@lumen.com) to re-activate the billing account. Then notify us here once it has been activated. ")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Review")
                elif kenanActvSrvcsAndRevenueCheck == 'Y':
                    #print("at 12 error" , financeAcctNbr)
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"This billing account is Inactive")
                    row.append(f"This is a Kenan biller with no Revenue and Active Services -- there's a (slim) chance it may have Voice services, but it passed the voice suffix check and no K86 bill cycle neither. ")
                    row.append("Mapped")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Map it!")
                elif billAcctActvtyStatusCd != 'Active':
                    #print("at 12 error" , financeAcctNbr)
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"This billing account is Inactive")
                    row.append(f"This billing account's BILL_ACCOUNT_STATUS_CD is {billAcctActvtyStatusCd} so it is not in MDM -- and there is nothing that we can do about it. It must be Denied. ")
                    row.append("The billing account's BILL_ACCOUNT_STATUS_CD is not 'Active' which means the billing account is no longer a valid billing account. We cannot govern this request.   ")
                    cannot_map_list.append(create_a_list)
                    row.insert(0, "Denied")
                elif (billingSystem in ['IDC_KENAN', 'KENANFX'] and billCycleCode == 'K86'):
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"The current cust is in SBG and moving to Indirect - raises eyebrow")
                    if re.search(voice_services, billAcctName):
                        row.append(f"This billing account may have Voice Services attached (according to its suffix) and its Bill Cycle Code is K86. If this don't have Voice Services, I don't know what done do.")
                        row.append("Feedback Required: This billing account very likely has Voice products attached. For that reason, it cannot be moved from its current customer. Please reach out to Ashley Ellis to confirm the Voice products on this billing account.")
                        cannot_map_list.append(create_a_list)
                        row.insert(0, "Nope")
                    elif not re.search(voice_services, billAcctName):
                        row.append(f"This is a(n) {billingSystem} acct w/ a K86 Bill Cycle - it is possible this has voice services attached, but the name doesn't feature any of the known Voice Suffixes - probably fine")
                        row.append("Mapped")
                        research_list.append(financeAcctNbr)
                        row.insert(0, "Map it!")
                else:
                    #print("at 13 error" , financeAcctNbr)
                    for num in error_field_nums:
                        create_a_list.append(row[int(num)])
                    create_a_list.append(f"not sure 'bout this one -- somethin' ain't right" )
                    row.append(f"We must decide whether requests of this type should be mapped or not - Error loop")
                    row.append("REVIEW")
                    research_list.append(financeAcctNbr)
                    row.insert(0, "Review")
            #print("welp, guess we couldn't find one", financeAcctNbr, billAcctLevelType)
    for ods in oracle_ods_list:
        oracleBillingAccountsODSids.append(str(ods))
    oracleBillingAccountsODSidsString = getFANsString(oracleBillingAccountsODSids)
    #print("\n\n", oracleBillingAccountsODSidsString, " --> here it is, the Oracle ods ids \n\n")

    
    # if len(red_can_map_list) > 0:
    #     print("\n\nthis is the red_can_map_list: \n", red_can_map_list)
    # if len(green_can_map_list) > 0:
    #     print("\n\nthis is the green_can_map_list: \n", green_can_map_list)
    # if len(cannot_map_list) > 0:
    #     print("\n\nthis is the error file: \n", cannot_map_list)
    # if len(oracleBillingAccounts) > 0:
    #     print("\n\nand here is the oracle billing account list:" , oracleBillingAccounts)
    # if len(kenanBillingAccounts) > 0:
    #     print("\n\nand here are the Kenan accounts:", kenanBillingAccounts)
    # else:
    #     print("something done gone wrong")
            
    

    
def getARdata(allFans):
    global arDataResultsList

    stringified_Fans = getFANsString(allFans)
    
    query = f"""
                SELECT 
                    DISTINCT acctNum 
                FROM
                    [dbo].[Modified Segments_ExtendedBuckets_CFS_PMT]
                WHERE
                    acctNum in ({stringified_Fans}) 
            """
    
    conn = dgc.connectSSMS_0213_AR_TABLE()
    c = conn.cursor()
    results = c.execute(query).fetchall()
    c.close()

    arDataResultsList = [fan for row in results for fan in row]
    # print("here is the ar DataResults List:", arDataResultsList)





def getFANProductTier2Data(fan_groups_dict):
    global fanTier2ProductDict


    conn = dgc.connectCDW()
    c = conn.cursor()

    for key, value in fan_groups_dict.items():
        # print("\n\nhere is the value: ", value)

        #just use string slicing to cut out the interior (instead of .replace() twice to remove square brackets)
        stringified_value = str(value)[1:-1]

        query = f"""SELECT
                    FINANCE_ACCOUNT_NBR
                    ,prod.product_tier2_cd
                FROM 
                    DSL_FINANCE.F_REVENUE_DETAIL_ALL FRDA
                    INNER JOIN CODS_FINANCE.GL_ACCOUNT GA ON FRDA.GL_ACCOUNT_ODS_ID = GA.GL_ACCOUNT_ODS_ID
                    left join udl_fi.ss_dsv_d_product_co prod on frda.gl_seg4_cd = prod.product_cd
                WHERE 1=1                         
                    AND FINANCE_ACCOUNT_NBR IN ({stringified_value})
                    AND prod.PRODUCT_TIER2_CD IS NOT NULL
                    and journal_source_cd in ('BR','BA')
                GROUP BY FINANCE_ACCOUNT_NBR , prod.PRODUCT_TIER2_CD 
            """
        
        # print(query)
        results = c.execute(query).fetchall()
        
        for row in results:
            fan = row[0]
            product_tier2 = row[1]

            if fan in fanTier2ProductDict:

                # if the product_tier2 is not already in the list for the fan, add it
                if product_tier2 not in fanTier2ProductDict[fan]:
                    fanTier2ProductDict[fan].append(product_tier2)
                else:
                    pass
            else:
                fanTier2ProductDict[fan] = [product_tier2]

    c.close()
    # print("\nhere are the fan tier2Products: ", fanTier2ProductDict)

    return fanTier2ProductDict



def getLatestTeamComment(tickets):
    global latestTeamCommentDict
    stringified_tix = getFANsString(tickets)

    query = f"""
select 
    distinct ic.issue_id
	,comment
	--,ic.employee_number
from 
	salesissuetracker.dbo.issue_m2 m2
	join salesissuetracker.dbo.issue_comments ic 
		on m2.issue_id = ic.issue_id
where 1=1
	and m2.status_id in ('26', '22') -- 26 = rqstr comments added, 22 = feedback required
	and issue_category_id = '9' -- 9 = b2c requests
	--and ic.employee_number in ('00302684', '00310272', '00318599','00336006', '00326822') 
	and ic.issue_comments_id in 
		(	
			select 
				max(issue_comments_id) 
			from 
				salesissuetracker.dbo.issue_comments ic
			where 
				comment_type_id = 1
				and raw_json is not null
				and issue_id in ({stringified_tix})
				and ic.employee_number in ('00302684', '00310272', '00318599','00336006', '00326822')
			group by issue_id
		)
	and ic.issue_id in 
		(	
			select
				issue_id
			from
				salesissuetracker.dbo.issue_comments
			where
				comment in 
					(
						'status changed from New to Feedback Required' 
						,'status changed from Feedback Required to Requestor Comments Added'
					)
		)
group by ic.issue_id, ic.comment, issue_comments_id --, employee_number;
    """
    # print(query)
    conn = dgc.connectSSMS()
    c = conn.cursor()
    results = c.execute(query).fetchall()
    c.close()
    # print("\n\nhere are the quarterly restuls: ", results)

    if results:
        latestTeamCommentDict = {row[0]: row[1] for row in results}



def getLatestRqstrComment(tickets):

    global latestRqstrCommentDict
    stringified_tix = getFANsString(tickets)

    query = f"""
    select 
        distinct ic.issue_id
		,comment
	from 
		salesissuetracker.dbo.issue_m2 m2
		join salesissuetracker.dbo.issue_comments ic 
			on m2.issue_id = ic.issue_id
	where 
		m2.status_id = '26' 
		and issue_category_id = '9'
		and modify_employee_nbr not in ('00302684', '00310272', '00318599','00336006', '00326822') 
		and ic.issue_comments_id in 
			(	
				select 
					max(issue_comments_id) 
				from 
					salesissuetracker.dbo.issue_comments 
				where 
					comment_type_id <> 3 
					and issue_id in ({stringified_tix})
				group by issue_id
			)
		and ic.issue_id in 
			(	
				select
					issue_id
				from
					salesissuetracker.dbo.issue_comments
				where
					comment in 
						(
							'status changed from New to Feedback Required' 
							,'status changed from Feedback Required to Requestor Comments Added'
						)
			)
	group by ic.issue_id, ic.comment, issue_comments_id;
    """

    # print(query)
    conn = dgc.connectSSMS()
    c = conn.cursor()
    results = c.execute(query).fetchall()
    c.close()

    # print("\n\nhere are the quarterly restuls: ", results)
    # b2cQuarterlyDataResultsList = [list(row) for row in results]
    if results:
        latestRqstrCommentDict = {row[0]: row[1] for row in results}





def getB2CQuarterlyData(fans_list: list) -> None:
    global b2cQuarterlyDataResultsList

    stringified_Fans = getFANsString(fans_list)
    
    query = f"""
                SELECT 
                    FINANCE_ACCOUNT_NBR
                    ,BAN_REASSIGNMENT_STATUS
                    ,RQSTD_CUST_NBR
                FROM 
                    DBO.QUARTERLY_B2C_REQUESTS
                WHERE
                    BAN_REASSIGNMENT_STATUS = 'Pending'
                    and FINANCE_ACCOUNT_NBR IN ({stringified_Fans}) 
            """
    
    conn = dgc.connectSSMS_QUARTERLY_B2C_REQUESTS_TABLE()
    c = conn.cursor()
    results = c.execute(query).fetchall()
    c.close()

    # print("\n\nhere are the quarterly restuls: ", results)
    b2cQuarterlyDataResultsList = [list(row) for row in results]
    # print("here is the b2c quartlery results List:", b2cQuarterlyDataResultsList, "\n\n")




def getCRIS_BAN_fromENS_BAN(fan_groups_dict: dict) -> None:
    global cris_to_ens_dict


    conn = dgc.connectIDG01P()
    c = conn.cursor()

    for key, value in fan_groups_dict.items():
        # print("\n\nhere is the value: ", value)

        #just use string slicing to cut out the interior (instead of .replace() twice to remove square brackets)
        stringified_value = str(value)[1:-1]

        query = f"""SELECT
                        CONCAT(ACCOUNT_NUMBER, '-ENS') as ENS_FINANCE_ACCOUNT_NBR
                        ,CONCAT(LEGACY_IDENTIFIER, CONCAT('-', SOURCE_SYSTEM)) as CRIS_FINANCE_ACCOUNT_NBR
                    FROM 
                        MDMRULE.LKP_BAN_XREF_ENS
                    WHERE 
                        ACCOUNT_NUMBER in({stringified_value})
            """
        
        # print(query)
        results = c.execute(query).fetchall()
        
        for row in results:
            cris_to_ens_dict[row[0]] = row[1]

    c.close()
    # print("\nhere are the fan tier2Products: ", fanTier2ProductDict)

    return None




def find_CRIS_to_ENS_BAN_data(fans_list: list) -> None:

    completeFANDict = get1kBlocksFromList(fans_list, 999)

    getCRIS_BAN_fromENS_BAN(completeFANDict)
    return None



def enterpriseVsMassMarkets_ProductCheck(fans_list, completeinfo):

    completeFANDict = get1kBlocksFromList(fans_list, 999)
    getFANProductTier2Data(completeFANDict)

    return None



def findWhereToMap(sourceSystem: str, fan: str, CBP: str) -> str:

    if (sourceSystem in salesforce_billing_systems or (sourceSystem in ['IDC_KENAN', 'KENANFX'] and re.match('.*\-LATIS$', fan))):
        if fan in mdm_FANs_dict:
            mapIn = "MDM / SF"
        else:
            mapIn = "SF ONLY / NOT IN MDM"
    elif (sourceSystem in am_source_systems) and (CBP is not None):
        if fan in mdm_FANs_dict:
            mapIn = "MDM / AM"
        else:
            mapIn = "AM ONLY / NOT IN MDM"
    elif (sourceSystem in am_source_systems) and (CBP is None):
        if re.search('^.*-GC$', fan):
            mapIn = "CPM"
        else:
            mapIn = "NO CBP / CANNOT MAP"
    elif sourceSystem == 'ZUORA':
        if fan in mdm_FANs_dict:
            mapIn = "MDM"
        else:
            mapIn = "CANNOT MAP / NOT IN MDM"
    return mapIn



def fanProductsListValueSetter(fan):

    # original: str(fanTier2ProductDict[fan].sort())[1:-1] if fan in fanTier2ProductDict else 'None'
    # can't run all those fixes at once and cannot sort the list while assigning it to a variable
    
    if fan in fanTier2ProductDict:
        product_list = fanTier2ProductDict[fan]
        sorted_list = sorted(product_list)
        #print("\n\nsorted list: ", sorted_list)
        removed_square_brackets = str(sorted_list)[1:-1]
        #print("no square brackets: ", removed_square_brackets)
        removed_apostrophes = removed_square_brackets.replace("'", "")
        #print("no apostrophes: ", removed_apostrophes)
        change_commas_to_pipes = removed_apostrophes.replace(", ", "|")
        return change_commas_to_pipes




def colorizeQuarterlyReviewSheet(csv_path, excel_path):
    df = pd.read_csv(csv_path)

    num_rows = df.shape[0]

    # prepare the final row so that data validation only runs to the bottom of the available data
    final_row = "B" + str(num_rows + 3)

    # flip csv to excel
    df.to_excel(excel_path, index=False)

    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook.active

    dv = DataValidation(type="list", formula1='"APPROVED, DENIED, NO ACTION, FEEDBACK REQUIRED, UNDER REVIEW, PENDING LEADERSHIP APPROVAL"')

    # Add the data validation object to the worksheet
    worksheet.add_data_validation(dv)

    # Apply the data validation to column A
    dv.add(f'B4:{final_row}')


    dark_green_cells = ['A1']
    light_blue_cells = ['B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1']
    darker_blue_cells = ['M1', 'N1', 'O1', 'P1', 'Q1', 'R1']
    light_yeller = [ 'S1', 'T1', 'U1', 'V1']
    light_green = ['W1', 'X1', 'Y1', 'Z1']
    red = ['AA1', 'AB1', 'AC1', 'AD1', 'AE1', 'AF1']
    yeller = ['AG1', 'AH1', 'AI1', 'AJ1', 'AK1', 'AL1', 'AM1', 'AN1', 'AO1']
    sf_pink = ['AP1', 'AS1', 'AT1', 'AU1', 'AV1', 'AW1', 'AX1', 'AY1', 'AZ1', 'BA1', 'BB1', 'BC1', 'BD1', 'BE1', 'BF1']
    am_orange = ['AQ1', 'AR1']

    for cell in dark_green_cells:
        worksheet[cell].fill = PatternFill(start_color="00B050", end_color="00B050", fill_type = "solid")

    for cell in light_blue_cells:
        worksheet[cell].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type = "solid")
    
    for cell in darker_blue_cells:
        worksheet[cell].fill = PatternFill(start_color="9bc2e6", end_color="9bc2e6", fill_type = "solid")

    for cell in light_yeller:
        worksheet[cell].fill = PatternFill(start_color="ffffcc", end_color="ffffcc", fill_type = "solid")

    for cell in light_green:
        worksheet[cell].fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type = "solid")
    
    for cell in yeller:
        worksheet[cell].fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type = "solid")

    for cell in red:
        worksheet[cell].fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type = "solid")

    for cell in sf_pink:
        worksheet[cell].fill = PatternFill(start_color="ff66cc", end_color="ff66cc", fill_type = "solid")
    
    for cell in am_orange:
        worksheet[cell].fill = PatternFill(start_color="ff822d", end_color="ff822d", fill_type = "solid")
    
    workbook.save(excel_path)


    #insert 2 rows starting at row 1
    worksheet.insert_rows(1, 3)
    worksheet.insert_cols(0, 1)

    # Merge cells
    worksheet.merge_cells('C3:M3') # --> Ticket information
    worksheet.merge_cells('N3:S3') # --> Billing Account information
    worksheet.merge_cells('T3:W3') # --> Current Customer information
    worksheet.merge_cells('X3:AA3') # --> Requested Customer information
    worksheet.merge_cells('AB3:AG3') # --> Denial Columns
    worksheet.merge_cells('AH3:AP3') # --> Review Columns
    worksheet.merge_cells('AR3:AS3') # --> AM DATA

    merged_first_cells_list = ['C3', 'N3', 'T3', 'X3', 'AB3', 'AH3', 'AR3']
    merged_final_cells_list = ['M3', 'S3', 'W3', 'AA3', 'AG3', 'AP3', 'AS3']
    merged_middle_cells_list = ['D3', 'E3', 'F3', 'G3', 'H3', 'I3', 'J3', 'K3', 
                                'L3', 'O3', 'P3', 'Q3', 'R3', 
                                'U3', 'V3',  'Y3', 'Z3',
                                'AC3', 'AD3', 'AE3', 'AF3',
                                'AI3', 'AJ3', 'AK3', 'AL3', 'AM3', 'AN3', 'AO3']
    # Add a value to the merged cell
    merged_font = Font(bold=True, size=14) 
    merged_alignment = Alignment(horizontal='center', vertical='center')
    medium_border = Side(border_style="medium", color="000000")  # Creates a thin black border
    cell_border_left = Border(top=medium_border, bottom=None, left=medium_border, right=None)
    cell_border_right = Border(top=medium_border, bottom=None, left=None, right=medium_border)
    cell_border_middle = Border(top=medium_border, bottom=None, left=None, right=None)

    for cell in merged_first_cells_list:
        worksheet[cell].font = merged_font
        worksheet[cell].alignment = merged_alignment
        worksheet[cell].border = cell_border_left

    for cell in merged_final_cells_list:
        worksheet[cell].border = cell_border_right

    for cell in merged_middle_cells_list:
        worksheet[cell].border = cell_border_middle


    worksheet['C3'] = "Ticket Information"
    worksheet['N3'] = "Billing Account Information"
    worksheet['T3'] = "Current Customer Information"
    worksheet['X3'] = "Requested Customer Information"
    worksheet['AB3'] = "Denial Columns"
    worksheet['AH3'] = "Review Columns"
    worksheet['AR3'] = "AM DATA"

    worksheet['C3'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type = "solid")
    worksheet['N3'].fill = PatternFill(start_color="9bc2e6", end_color="9bc2e6", fill_type = "solid")
    worksheet['T3'].fill = PatternFill(start_color="ffffcc", end_color="ffffcc", fill_type = "solid")
    worksheet['X3'].fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type = "solid")
    worksheet['AB3'].fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type = "solid")
    worksheet['AH3'].fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type = "solid")
    worksheet['AR3'].fill = PatternFill(start_color="ff822d", end_color="ff822d", fill_type = "solid")
    # worksheet['AQ2'].font = merged_font
    # worksheet['AQ2'].alignment = merged_alignment
    # worksheet['AQ2'].fill = PatternFill(start_color="ff822d", end_color="ff822d", fill_type = "solid")
    # worksheet['AQ2'].border = cell_border_left
    # worksheet['AR2'].border = cell_border_right



    workbook.save(excel_path)
    workbook.close()




def findQuarterlyB2CRequestStatus(fan: str) -> str:
    global b2cQuarterlyDataResultsList

    result = {}

    for row in b2cQuarterlyDataResultsList:
        results_fan = row[0]
        reassignment_status = row[1]
        rqstd_cust = row[2]
        if fan == results_fan:
            result[fan] = [reassignment_status, rqstd_cust]

    return result
        


def getRootCauseStr(root_cause: str) -> str:

    str_root_cause = ''

    if root_cause == 1:
        str_root_cause = "Incorrect Initial Assignment"
    elif root_cause == 2:
        str_root_cause = "Security / CPNI Issue"
    elif root_cause == 3:
        str_root_cause = "Order Entry / Service Mapping Misalignments"
    elif root_cause == 4:
        str_root_cause = "Corporate action (M&A or Divestiture)"
    else:
        str_root_cause = "Indirect v. Direct"

    return str_root_cause




def getCrossChannelStatus(rev: float, assgnEval: str, buEval: str) -> str:
    if rev != 0:
        # split into assignment group and BU eval steps 1/23/2025
        if assgnEval == 'Pass' and buEval == 'Pass':
            return 'No Change to Assgnmnt Grp or BU'
        elif assgnEval == 'Fail' and buEval == 'Pass':
            return 'Quarterly Review'
        elif assgnEval == 'Fail' and buEval == 'Fail':
            return 'Annual Review'
        else:
            return 'Error Status'
    else:
        return 'No Revenue Impact'




def createActionStr(conditions: list) -> str:
    #print("\n\nhere is the length of the conditions list: ", len(conditions))
    
    global evalStr
    global dg_comment

    evalStr = ''
    actionStr = ''
    dg_comment = ''

    oe_link = "<a href='https://bit.ly/4eT7NdH'>this link</a>"


    ticket_id = conditions[1]
    fan = conditions[1]
    # ban_name = row[27]
    ban_dw_secure_nbr = conditions[3]
    ban_order_entry = conditions[4]
    ban_source_system = conditions[5]
    fan_products_list = conditions[6]
    three_month_revenue = conditions[7]
    # bill_acct_cust_typ = row[67]
    # bill_acct_cust_sub_typ = row[68]
    # CBP = row[61]
    bill_acct_actvty_status_cd = conditions[11]
    # billing_account_id_c = conditions[11]
    # billing_unique_external_id_c = row[50]
    # billing_system_c = row[51]
    # billing_address_line_1_c = row[52]
    # billing_address_line_2_c = row[53]
    # billing_city_c = row[54]
    # billing_state_c = row[55]
    # billing_postal_code_c = row[56]
    # billing_country_c = row[57]
    # billing_instance_c = row[60]
    # record_type_id = row[59]
    # customer_account_c = row[92]




    # map_location = findWhereToMap(ban_source_system, fan, CBP)

    crrnt_cust_nbr = conditions[25]
    # crrnt_cust_name = row[38]
    crrnt_cust_dw_secure_nbr = conditions[27]
    # cods_currnt_cust_status = row[40]
    crrnt_cust_assngmnt_grp = conditions[29]
    crrnt_cust_trdng_prtnr = conditions[30] 
    crrnt_cust_extrnl_rprting_BU = conditions[31]

    rqstd_cust_nbr = conditions[32]
    # rqstd_cust_name = row[82]
    rqstd_cust_dw_secure_nbr = conditions[34]  
    rqstd_cust_status = conditions[35]
    rqstd_cust_assngmnt_grp = conditions[36]
    rqstd_cust_trdng_prtnr = conditions[37]
    rqstd_cust_extrnl_rprting_BU = conditions[38]

    # business_unit_check= row[104]
    # sales_channel_check = row[103]

    # suggested_action = row[0]
    # evaluation = row[108]
    # comment_4_ticket = row[109]

    # # == the ones below still need to be added to the row
    # ticket_type = row[9]   
    root_cause = conditions[45]
    # crrnt_cust_lob = row[33]
    # rqstd_cust_lob = row[78]

    # rqstr_job_title = row[98]
    # rqstr_name = row[3] + ' ' + row[4]
    # rqstr_comment = row[6]
    rqstr_emp_nbr = conditions[51]
    quarterly_b2c_request_status = conditions[52]

    # print("here is the conditions data: ", ban_dw_secure_nbr, crrnt_cust_dw_secure_nbr, rqstd_cust_dw_secure_nbr, rqstd_cust_status, root_cause, rqstr_emp_nbr)

    if root_cause == 4 and rqstr_emp_nbr in bac_team_empIDs:
        actionStr = 'Auto-Map'
        evalStr = 'Root Cause is Corporate Action and Requestor is in BAC Team'
        dg_comment = 'Mapped'

    if three_month_revenue == 0 and fan not in arDataResultsList and rqstd_cust_nbr not in placeholder_customers:
        actionStr = 'Auto-Map'
        if evalStr:
            evalStr += ', No Rev Impact, FAN not in AR Data, RQSTD_CUST not Placeholder'
        else:
            evalStr = 'No Rev Impact, FAN not in AR Data, RQSTD_CUST not Placeholder'
        dg_comment = 'Mapped'
    
    if crrnt_cust_nbr in placeholder_customers or (crrnt_cust_nbr is None or crrnt_cust_nbr == ''):
        if rqstd_cust_nbr not in placeholder_customers:
            actionStr = 'Auto-Map'
            if evalStr:
                evalStr += ', CRRNT_CUST is Placeholder or Blank, RQSTD_CUST not Placeholder' if 'RQSTD_CUST not Placeholder' not in evalStr else ', CRRNT_CUST is Placeholder or Blank'
            else:
                evalStr = 'CRRNT_CUST is Placeholder or Blank, RQSTD_CUST not Placeholder'
            dg_comment = 'Mapped'
        if three_month_revenue != 0 and fan in arDataResultsList:
            if (crrnt_cust_trdng_prtnr is None or crrnt_cust_trdng_prtnr == '') and (rqstd_cust_trdng_prtnr is None or rqstd_cust_trdng_prtnr == ''):
                if (crrnt_cust_assngmnt_grp == rqstd_cust_assngmnt_grp):
                    actionStr = 'Auto-Map'
                    if evalStr:
                        evalStr += ', CRRNT_CUST is Placeholder or Blank, No Rev Impact, FAN not in AR Data, CRRNT_CUST and RQSTD_CUST have same AG' if \
                                      'CRRNT_CUST is Placeholder or Blank, FAN not in AR Data' not in evalStr else ', CRRNT_CUST and RQSTD_CUST have same AG'
                    else:
                        evalStr = 'CRRNT_CUST is Placeholder or Blank, FAN not in AR Data, CRRNT_CUST and RQSTD_CUST have same AG'
                    dg_comment = 'Mapped'

    # is this right? If Order entry flag exists and there are simply voice products IN the product list then it should be feedback required?
    if ban_order_entry == 'Y':
        if fan_products_list is not None:
            if 'Voice' in fan_products_list and rqstr_emp_nbr != big_dawg_for_voice_products and ban_source_system == 'KENANFX':
                actionStr = 'Feedback Required'
                if evalStr:
                    evalStr += f', OE problems exist, "Voice" in Product List, Requestor not {big_dawg_for_voice_products_name}'
                else:
                    evalStr = f'OE problems exist, "Voice" in Product List, Requestor not {big_dawg_for_voice_products_name}'
                # leave room to change Evaluation here (that's why there are two different evaluations)
                dg_comment= f"""Feedback Required: This billing account has products which have ties to the billing account's current customer at an individual service level.
                              Also, this billing account may have Voice Services attached, so please reach out to Ashley Ellis to confirm, before moving forward with a BAC request to move services.
                              \n\nPlease visit this link for more information: https://bit.ly/4eT7NdH"""       
        if rqstr_emp_nbr not in bac_team_empIDs:
            actionStr = 'Feedback Required'
            if evalStr:
                evalStr += ', OE problems exist, Requestor not BAC Team' if 'OE problems exist' not in evalStr else ', Requestor not BAC Team' 
            else:
                evalStr = 'OE problems exist, Requestor not BAC Team'
            dg_comment =f"Feedback Required: This billing account has products which have ties to the billing account's current customer at an individual service level.\n\nPlease visit this link for more information: https://bit.ly/4eT7NdH"


    if fan_products_list is not None:
        if fan_products_list == 'Voice' and rqstr_emp_nbr != big_dawg_for_voice_products and ban_source_system == 'KENANFX' \
            and not re.match('^.*\-LATIS$', fan): 
            actionStr = 'Feedback Required'
            if evalStr:
                evalStr += ', "Voice" is only Product on KENANFX billing account, Requestor not Ashley Ellis'
            else:
                evalStr = '"Voice" is only Product on KENANFX billing account, Requestor not Ashley Ellis'
            # leave room to change Evaluation here (that's why there are two different evaluations)
            dg_comment = f"Feedback Required: This billing account has Voice products attached. For that reason, it cannot be moved from its current customer.\n\nPlease reach out to Ashley Ellis to confirm the Voice products on this billing account."
    

    if ban_source_system in ['ZUORA', 'PPP'] and crrnt_cust_extrnl_rprting_BU == 'MASS MARKETS':
        if rqstd_cust_extrnl_rprting_BU != 'MASS MARKETS':  
            actionStr = 'Auto-Deny'
            if evalStr:
                evalStr += ', BAN is ZUORA or PPP, CRRNT_CUST is MM and RQSTD_CUST is not'
            else:
                evalStr = 'BAN is ZUORA or PPP, CRRNT_CUST is MM and RQSTD_CUST is not'
            dg_comment = f"Denied: This billing account is {ban_source_system}, which means it cannot be moved from its current, MASS MARKETS customer." 

    if rqstd_cust_status != 'Approved':
        actionStr = 'Auto-Deny'
        if evalStr:
            evalStr += ', RQSTD_CUST is not Approved'
        else:
            evalStr = 'RQSTD_CUST is not Approved'
        dg_comment = f"Denied: The Requested Customer ({rqstd_cust_nbr}) is not Approved. Please work with Sales, Sales Ops, your peers, or leadership to find or create the desired customer, then create a new ticket for the request. "  

    # does't matter if the current customer isn't LUMEN
    if ban_dw_secure_nbr == 1 and ((rqstd_cust_dw_secure_nbr != '' and rqstd_cust_dw_secure_nbr is not None) and rqstd_cust_dw_secure_nbr != 1):
        actionStr = 'Auto-Deny'
        if evalStr:
            evalStr += ', BAN is Lumen, RQSTD_CUST is not'
        else:
            evalStr = 'BAN is Lumen, RQSTD_CUST is not'
        dg_comment = f"Denied: The Requested Customer ({rqstd_cust_nbr}) is not a Lumen customer. Please work with Sales, Sales Ops, your peers, or leadership to find or create the desired customer, then create a new ticket for the request. "


    if quarterly_b2c_request_status == 'Pending':
        actionStr = 'Auto-Deny'
        if evalStr: 
            evalStr += ', BAN is Pending in Quarterly table'
        else:
            evalStr = 'BAN is Pending in Quarterly table'
        dg_comment = f"Denied: the billing account already has a requested mapping in the Quarterly moves table and cannot be mapped to a different customer until the first request is completed."


    if rqstd_cust_nbr == crrnt_cust_nbr:
        if fan in mdm_FANs_dict:
            mdm_cust = mdm_FANs_dict[fan][0]
            if rqstd_cust_nbr == mdm_cust:
                # print("\n\n rqstd cust == mdm cust: ", rqstd_cust_nbr, mdm_cust)
                actionStr = 'Auto-NoAction'
                if evalStr:
                    evalStr += ', CRRNT_CUST == RQSTD_CUST in MDM & CODS'
                else:
                    evalStr = 'CRRNT_CUST == RQSTD_CUST in MDM & CODS'          
        # elif fan not in mdm_FANs_dict:
        #     actionStr = 'Review-FAN not in MDM'
        else:
            actionStr = 'Auto-NoAction'
            if evalStr:
                evalStr += ', CRRNT_CUST == RQSTD_CUST in CODS only'
            else:
                evalStr = 'CRRNT_CUST == RQSTD_CUST in CODS only'
        dg_comment = 'Validated - No Action Required: The billing account is already mapped to the desired customer'
    else:
        if fan in mdm_FANs_dict:
            mdm_cust = mdm_FANs_dict[fan][0]
            if rqstd_cust_nbr == mdm_cust:
                actionStr = 'Auto-NoAction'
                if evalStr:
                    evalStr += ', RQSTD_CUST == MDM_CUST'
                else:
                    evalStr = 'RQSTD_CUST == MDM_CUST'
                dg_comment = 'Validated - No Action Required: The billing account is already mapped to the desired customer'


    if crrnt_cust_nbr not in placeholder_customers and (crrnt_cust_nbr is not None and crrnt_cust_nbr != ''):
        if rqstd_cust_nbr in placeholder_customers:
            actionStr = 'Review'
            if evalStr:
                evalStr = 'Review Placeholder Movement ' + evalStr
                evalStr += ', CRRNT_CUST is Real Cust and RQSTD_CUST is Placeholder'
            else:
                evalStr = 'CRRNT_CUST is Real Cust and RQSTD_CUST is Placeholder'
            dg_comment = 'UNSURE WHAT TO WRITE FOR THIS COMMENT'


    if ban_source_system not in ERP_source_systems:
        actionStr = 'Auto-Deny'
        if evalStr:
            evalStr += f', {ban_source_system} billing accounts not in ERP'
        else:
            evalStr = f'{ban_source_system} billing accounts not in ERP'
        dg_comment = f"Denied: {ban_source_system} billing accounts are not in ERP; therefore, their movement is not governed."


    if ticket_id in latestRqstrCommentDict:
        actionStr = 'Review Previous'


    if fan not in mdm_FANs_dict:
        actionStr = 'Review'
        if bill_acct_actvty_status_cd is None or bill_acct_actvty_status_cd == '':
            if evalStr:
                evalStr = 'FAN not in MDM: BILL_ACCT_ACTVTY_STATUS_CD is Blank, ' + evalStr
            else:
                evalStr = 'FAN not in MDM: BILL_ACCT_ACTVTY_STATUS_CD is Blank'
        else:
            if evalStr:
                evalStr = 'FAN not in MDM: not sure why, ' + evalStr
            else:
                evalStr = 'FAN not in MDM: not sure why'
        dg_comment = 'The billing account does not exist in MDM and therefore cannot be moved from its current customer.'


    if three_month_revenue != 0:
        if (crrnt_cust_assngmnt_grp != rqstd_cust_assngmnt_grp) and (crrnt_cust_extrnl_rprting_BU == rqstd_cust_extrnl_rprting_BU):
            if evalStr:
                evalStr = 'Revenue impact, crossing Assignment Groups only, ' + evalStr
            else:
                evalStr = 'Revenue impact, crossing Assignment Groups only'
            actionStr = 'Hold for Quarterly Review'
            dg_comment = 'This request must be held for Quarterly Review.'
        elif (crrnt_cust_assngmnt_grp != rqstd_cust_assngmnt_grp) and (crrnt_cust_extrnl_rprting_BU != rqstd_cust_extrnl_rprting_BU):
            if evalStr:
                evalStr = 'Revenue impact, crossing Assignment Groups and Business Units, ' + evalStr
            else:
                evalStr = 'Revenue impact, crossing Assignment Groups and Business Units'
            actionStr = 'Hold for Annual Review'
            dg_comment = 'This request must be held for Annual Review.'

    
    print("\n\nhere is the dg_comment: ", dg_comment)
    return actionStr
    

    

    



def buildQuarterlyReviewSheet(complete_info: list) -> None:

    global placeholder_customers
    global ticket_analysis_dict


    with open(B2C_Quarterly_and_Annual_Rev_csv_path, 'a+', newline='', encoding="utf-8") as out:
        csv_out = csv.writer(out)
        csv_out.writerow([
            'DECISION',
            'SUGGESTED_ACTION',
            'EVALUATION',
            'COMMENT_4_TICKET',
            'CROSS_CHANNEL_TYPE',
            'MDM_SF_OR_AM',
            'TICKET_ID', 
            'TICKET_TYPE',
            'ROOT_CAUSE',
            'REQUESTOR_NAME',
            'REQUESTOR_JOB_TITLE',
            'REQUESTOR_COMMENT',
            'FINANCE_ACCOUNT_NBR',
            'BILL_ACCOUNT_NAME',
            'BILL_SOURCE_SYSTEM',
            'CRIS_TO_ENS_FAN',
            'FAN_PRODUCTS_LIST',
            'THREE_MONTH_AVG_REVENUE',
            'CURRENT_CUST_NBR',
            'CRRNT_CUST_NAME',
            'CRRNT_CUST_ASSGNMNT_GRP',
            'CRRNT_CUST_EXTRNL_RPRTNG_BU',
            'RQSTD_CUST_NBR',
            'RQSTD_CUST_NAME',
            'RQSTD_CUST_ASSGNMNT_GRP',
            'RQSTD_CUST_EXTRNL_RPRTNG_BU',
            'RQSTD_CUSTOMER_STATUS',
            'RQSTD_CUST_EQUALS_CRRNT_CUST',
            'SECURE_COMPANY_NBR',
            'VOICE_SERVICES',
            'ORDER_ENTRY_CHECK',
            'QUARTERLY_MOVES_CHECK',
            'MM_ENT_MOVEMENT_CHECK',
            'TTM_REVENUE',
            'COLLECTIONS_AR',
            'CRRNT_CUST_IS_PLCHLDR_OR_BLANK',
            'MOVING_TO_PLACEHOLDER',
            'TRADING_PARTNER_CHECK',
            'ASSIGNMENT_GROUP',
            'BUSINESS_UNIT',
            'CHANNEL',
            'FAN', ## Begin SALESFORCE upload sheet #changed from "FINANACE_ACCOUNT_NBR" to comply with Table on 2/5/2025
            'BANID',
            'NEWCUSTOMERID',
            'BILL_ACCT_NAME',
            'BILLING_ACCOUNT_ID__C',
            'BILLING_UNIQUE_EXTERNAL_ID__C',
            'BILLING_SYSTEM__C',
            'BILLING_ADDRESS_LINE_1__C',
            'BILLING_ADDRESS_LINE_2__C',   
            'BILLING_CITY__C',
            'BILLING_STATE__C',
            'BILLING_POSTAL_CODE__C',
            'BILLING_COUNTRY__C',
            'BAN_DW_SECURE_COMPANY_NBR',
            'BILLING_INSTANCE__C',
            'RECORDTYPEID',
            'CUSTOMER_ACCOUNT__C',
                ]) # headers
        
        book2 = []

        
        for row in complete_info:
            # print("\nhere is the row before further index adjustments: ", row, "\n")
            # print("\nhere is the len of the row: ", len(row))
            book = []

            ticket_id = row[1]
            fan = row[15]
            ban_name = row[27]
            ban_dw_secure_nbr = row[58]
            ban_order_entry = row[107]
            ban_source_system = row[28]
            fan_products_list = fanProductsListValueSetter(fan)
            three_month_revenue = round(row[20], 2)
            bill_acct_cust_typ = row[67]
            bill_acct_cust_sub_typ = row[68]
            CBP = row[61]
            bill_acct_actvty_status_cd = row[48]
            billing_account_id_c = row[49]
            billing_unique_external_id_c = row[50]
            billing_system_c = row[51]
            billing_address_line_1_c = row[52]
            billing_address_line_2_c = row[53]
            billing_city_c = row[54]
            billing_state_c = row[55]
            billing_postal_code_c = row[56]
            billing_country_c = row[57]
            billing_instance_c = row[60]
            record_type_id = row[59]
            customer_account_c = row[92]




            map_location = findWhereToMap(ban_source_system, fan, CBP)

            crrnt_cust_nbr = row[31]
            crrnt_cust_name = row[38]
            crrnt_cust_dw_secure_nbr = row[39]
            cods_currnt_cust_status = row[40]
            crrnt_cust_assngmnt_grp = row[65]
            crrnt_cust_trdng_prtnr = row[66]
            crrnt_cust_extrnl_rprting_BU = row[35]

            rqstd_cust_nbr = row[69]
            rqstd_cust_name = row[82]
            rqstd_cust_dw_secure_nbr = row[83]  
            rqstd_cust_status = row[84]
            rqstd_cust_assngmnt_grp = row[94]
            rqstd_cust_trdng_prtnr = row[95]
            rqstd_cust_extrnl_rprting_BU = row[79]

            business_unit_check= row[104]
            sales_channel_check = row[103]

            suggested_action = row[0]
            evaluation = row[108]
            comment_4_ticket = row[109]

            # == the ones below still need to be added to the row
            ticket_type = row[9]   
            root_cause = row[17]
            crrnt_cust_lob = row[33]
            rqstd_cust_lob = row[78]

            rqstr_job_title = row[98]
            rqstr_name = row[3] + ' ' + row[4]
            rqstr_comment = row[6]
            rqstr_emp_nbr = row[2]

            quarterly_b2c_request_status = findQuarterlyB2CRequestStatus(fan)

            conditions_list = [ticket_id, fan, ban_name, ban_dw_secure_nbr, ban_order_entry, 
                               ban_source_system, fan_products_list, three_month_revenue, bill_acct_cust_typ, 
                               bill_acct_cust_sub_typ, CBP, bill_acct_actvty_status_cd, billing_account_id_c, billing_unique_external_id_c, 
                               billing_system_c, billing_address_line_1_c, billing_address_line_2_c, 
                               billing_city_c, billing_state_c, billing_postal_code_c, billing_country_c, 
                               billing_instance_c, record_type_id, customer_account_c, map_location, crrnt_cust_nbr, 
                               crrnt_cust_name, crrnt_cust_dw_secure_nbr, cods_currnt_cust_status,
                               crrnt_cust_assngmnt_grp, crrnt_cust_trdng_prtnr, crrnt_cust_extrnl_rprting_BU,
                               rqstd_cust_nbr, rqstd_cust_name, rqstd_cust_dw_secure_nbr, rqstd_cust_status,
                               rqstd_cust_assngmnt_grp, rqstd_cust_trdng_prtnr, rqstd_cust_extrnl_rprting_BU,
                               business_unit_check, sales_channel_check, suggested_action, evaluation, comment_4_ticket,
                               ticket_type, root_cause, crrnt_cust_lob, rqstd_cust_lob, rqstr_job_title, rqstr_name,
                               rqstr_comment, rqstr_emp_nbr, quarterly_b2c_request_status]



            book.append('')
            book.append(createActionStr(conditions_list))#used to be suggested_action -- this will be replaced as we figure things out 1/21/2025
            book.append(evalStr)
            book.append(dg_comment) #global that handles the NEW process (instead of comment_4_ticket)
            book.append(getCrossChannelStatus(three_month_revenue, ('Pass' if rqstd_cust_assngmnt_grp == crrnt_cust_assngmnt_grp else 'Fail'), ('Pass' if crrnt_cust_extrnl_rprting_BU == rqstd_cust_extrnl_rprting_BU else 'Fail')))
            book.append(map_location)
            book.append(ticket_id)
            book.append(ticket_type)
            book.append(getRootCauseStr(root_cause))
            book.append(rqstr_name)
            book.append(rqstr_job_title)
            book.append(latestRqstrCommentDict[ticket_id] if ticket_id in latestRqstrCommentDict else rqstr_comment)
            book.append(fan)
            book.append(ban_name)
            book.append(ban_source_system)
            book.append(cris_to_ens_dict[fan] if fan in cris_to_ens_dict else '')
            book.append(fan_products_list)
            book.append(three_month_revenue)
            book.append(crrnt_cust_nbr)
            book.append(crrnt_cust_name)
            book.append(crrnt_cust_assngmnt_grp)
            book.append(crrnt_cust_extrnl_rprting_BU)
            book.append(rqstd_cust_nbr)
            book.append(rqstd_cust_name)
            book.append(rqstd_cust_assngmnt_grp)
            book.append(rqstd_cust_extrnl_rprting_BU)

            ticket_analysis_dict_key = fan + str(ticket_id)
            ticket_analysis_dict[ticket_analysis_dict_key] = []


            ## mm vs ent
            ## and re.match('.*\-LATIS$', row[15])
            if rqstd_cust_status == 'Approved':
                book.append('Pass')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Pass')
            else:
                book.append('Fail')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Fail')
                # decisionDict[0] = 'Denied'
                # decisionDict[1] = 'The requested customer is not Approved'
                # decisionDict[2] = 'Request Denied: The requested customer is not in an Approved status; therefore, the requested mapping cannot be complated'

            
            if rqstd_cust_nbr == crrnt_cust_nbr:
                book.append('Fail')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Fail')
            else:
                book.append('Pass')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Pass')



            if (ban_dw_secure_nbr == 1 and rqstd_cust_dw_secure_nbr == 1 and crrnt_cust_dw_secure_nbr == 1) \
                or (ban_dw_secure_nbr == 1 and rqstd_cust_dw_secure_nbr == 1 and (crrnt_cust_dw_secure_nbr is None and crrnt_cust_nbr is None)):
                book.append('Pass')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Pass')
            else: 
                book.append('Fail')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Fail')


            if ban_source_system in ('KENANFX') and fan in fanTier2ProductDict and not re.match('^.*\-LATIS$', fan): 
                if 'Voice' in fanTier2ProductDict[fan]:
                    book.append('Fail')
                    ticket_analysis_dict[ticket_analysis_dict_key].append('Fail')
                # this one runs the ban through the Voice Services Regex as a last-ditch to 
                # identify potential Voice Services accounts    
                elif re.search(voice_services, ban_name):
                    book.append('Fail')
                    ticket_analysis_dict[ticket_analysis_dict_key].append('Fail')
                else:
                    book.append('Pass')
                    ticket_analysis_dict[ticket_analysis_dict_key].append('Pass')
            else:
                book.append('Pass')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Pass')
            

            # ====================================== ORDER ENTRY ==========================================
            # =============================================================================================
            # - add the BAC team to this check 10/31/2024
            # - added check for Root Cause "Corporate Actions M&A" which will be handled as exception
            # - added ENS check for Blue Marble 1/7/2025
            # ---> if BlueMarble ENS biller is on MM and moving to ENT then fail, if on ENT and moving to ENT then fail
            if ban_order_entry == 'Y':
                if ticket_id == 20037616:
                    print("found the ticket on PASS for OE system check ", ticket_id)
                if rqstr_emp_nbr in bac_team_empIDs or root_cause == 4:
                    book.append('Pass')
                    ticket_analysis_dict[ticket_analysis_dict_key].append('Pass')
                elif ban_source_system in ('ENS'):
                    if fan in FANandOrderEntryDict:
                        if 'BLUEMARBLE' in FANandOrderEntryDict[fan]:
                            if crrnt_cust_extrnl_rprting_BU == 'MASS MARKETS' and rqstd_cust_extrnl_rprting_BU != 'MASS MARKETS':
                                book.append('Fail')
                                ticket_analysis_dict[ticket_analysis_dict_key].append('Fail')
                            elif crrnt_cust_extrnl_rprting_BU != 'MASS MARKETS' and rqstd_cust_extrnl_rprting_BU != 'MASS MARKETS':
                                book.append('Fail')
                                ticket_analysis_dict[ticket_analysis_dict_key].append('Fail')
                            else:
                                book.append('Pass')
                                ticket_analysis_dict[ticket_analysis_dict_key].append('Pass')
                else:
                    book.append('Fail')
                    ticket_analysis_dict[ticket_analysis_dict_key].append('Fail')
            else:
                if ticket_id == 20037616:
                    print("found the ticket on fail for OE system check ", ticket_id)
                book.append('Pass')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Pass')



            if fan in quarterly_b2c_request_status:
                if quarterly_b2c_request_status[fan][0] == 'Pending':
                    book.append(quarterly_b2c_request_status[fan][1])
                else:
                    book.append(quarterly_b2c_request_status[0])
            else:
                book.append('')



            if ban_source_system in ['PPP', 'ZUORA'] and rqstd_cust_extrnl_rprting_BU != 'MASS MARKETS':
                book.append('Fail')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Fail')
            elif ban_source_system in ['KENANFX', 'LATIS', 'SAP', 'NIBS', 'DASSIAN', 'CABS'] and rqstd_cust_extrnl_rprting_BU == 'MASS MARKETS':
                book.append('Fail')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Fail')
            elif ban_source_system in ('ENS', 'MBS'):
                if ban_source_system == 'ENS' and bill_acct_cust_typ == 'Residential' and rqstd_cust_extrnl_rprting_BU != 'MASS MARKETS':
                    decision = 'Fail'
                elif fan in fanTier2ProductDict:
                    # bloody hell. If any of the products attached to the FAN are in the Mass Markets products list
                    ### and none of its products are in the Enterprise or Both products lists 
                    ### AND the requested customer is NOT Mass Markets, then the requested movement should be denied (Fail)
                    ### !! ADDED 1/7/2025 - and the current customer is NOT already in Enterprise (because if it is, then it can stay there)
                    ### This is because the only products identified belong solely to MM, so the FAN should belong to a MM Cust
                    if (any(product in fanTier2ProductDict[fan] for product in mass_markets_products) \
                        and not any(product in fanTier2ProductDict[fan] for product in enterprise_products) \
                        and not any(product in fanTier2ProductDict[fan] for product in mass_markets_and_enterprise_products)) \
                        and rqstd_cust_extrnl_rprting_BU != 'MASS MARKETS' \
                        and not (crrnt_cust_extrnl_rprting_BU == 'MASS MARKETS'):

                        decision = 'Fail'
                    # on the other hand, if any of the products attached to the FAN are in the Enterprise products list
                    ### and none of its products are in the MM or Both products lists
                    ### AND the requested customer IS Mass Markets, then the requested movement should be denied (Fail)
                    ### !! ADDED 1/7/2025 - and the current customer is NOT already in MASS MARKETS (because if it is, then it can stay there)
                    ### because all the products identified belong solely to Enterprise, so the FAN should belong to an Enterprise Cust
                    elif any(product in fanTier2ProductDict[fan] for product in enterprise_products) \
                        and not any(product in fanTier2ProductDict[fan] for product in mass_markets_products) \
                        and not any(product in fanTier2ProductDict[fan] for product in mass_markets_and_enterprise_products) \
                        and (rqstd_cust_extrnl_rprting_BU == 'MASS MARKETS' and crrnt_cust_extrnl_rprting_BU != 'MASS MARKETS'):

                        decision = 'Fail'
                    else:
                        decision = 'Pass'               
                # try:
                #     decision
                #     print("Decision has a value: ", decision)   
                # except NameError:
                #     print("\n\nDecision Error: ", row)

                book.append(decision)
                ticket_analysis_dict[ticket_analysis_dict_key].append(decision)
            else:
                book.append('Pass')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Pass')



            #  =========================================== REVENUE =======================================
            # changed to != 0 from > 0 to be sure that any amount that isn't 0 (including credits) will be caught
            if three_month_revenue != 0:
                if three_month_revenue <= 2000:
                    book.append('$')
                    ticket_analysis_dict[ticket_analysis_dict_key].append('$')
                elif three_month_revenue > 2000 and three_month_revenue <= 5000:
                    book.append('$$')
                    ticket_analysis_dict[ticket_analysis_dict_key].append('$$')
                elif three_month_revenue > 5000 and three_month_revenue <= 10000:
                    book.append('$$$')
                    ticket_analysis_dict[ticket_analysis_dict_key].append('$$$')
                elif three_month_revenue > 10000 and three_month_revenue <= 25000:
                    book.append('$$$$')
                    ticket_analysis_dict[ticket_analysis_dict_key].append('$$$$')
                elif three_month_revenue > 25000 and three_month_revenue <= 50000:
                    book.append('$$$$$')
                    ticket_analysis_dict[ticket_analysis_dict_key].append('$$$$$')
                elif three_month_revenue > 50000 and three_month_revenue <= 100000:
                    book.append('$$$$$$')
                    ticket_analysis_dict[ticket_analysis_dict_key].append('$$$$$$')
                elif three_month_revenue > 100000:
                    book.append('!!!!!!!')
                    ticket_analysis_dict[ticket_analysis_dict_key].append('!!!!!!!')
            else:
                book.append('Zero')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Zero')

            if fan in arDataResultsList:
                book.append('Fail') 
                ticket_analysis_dict[ticket_analysis_dict_key].append('Fail')
            else:
                book.append('Pass')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Pass')
    

            #ticket_analysis_dict[ticket_analysis_dict_key].append('UNK')

            if crrnt_cust_nbr in placeholder_customers or (crrnt_cust_nbr is None or crrnt_cust_nbr == ''):
                book.append('Y')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Y')
            else:
                book.append('N')
                ticket_analysis_dict[ticket_analysis_dict_key].append('N')

            # speak to team about adding Consumer 100 to the placeholder customers list
            if rqstd_cust_nbr in placeholder_customers:
                book.append('Fail')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Fail')
            else:
                book.append('Pass')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Pass')


            if (crrnt_cust_trdng_prtnr is None or crrnt_cust_trdng_prtnr == '') and (rqstd_cust_trdng_prtnr is None or rqstd_cust_trdng_prtnr == ''):
                book.append('Pass')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Pass')
            else:
                book.append('Fail')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Fail')
            
            if rqstd_cust_assngmnt_grp == crrnt_cust_assngmnt_grp:
                book.append('Pass')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Pass')
            else:
                book.append('Fail')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Fail')

            if business_unit_check == 'Y':
                book.append('Fail')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Fail')
            else:
                book.append('Pass')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Pass')
            
            if sales_channel_check == 'Y':
                book.append('Fail')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Fail')
            else:
                book.append('Pass')
                ticket_analysis_dict[ticket_analysis_dict_key].append('Pass')

            book.append(fan)
            book.append(CBP)
            book.append(rqstd_cust_nbr)
            book.append(ban_name)
            book.append(billing_account_id_c)
            book.append(billing_unique_external_id_c)
            book.append(billing_system_c)
            book.append(billing_address_line_1_c)
            book.append(billing_address_line_2_c)
            book.append(billing_city_c)
            book.append(billing_state_c)
            book.append(billing_postal_code_c)
            book.append(billing_country_c)
            book.append(ban_dw_secure_nbr)
            book.append(billing_instance_c)
            book.append(record_type_id)
            book.append(customer_account_c)



            

            book2.append(book)

        for entry in book2:
            csv_out.writerow(entry)
    
    out.close()

    colorizeQuarterlyReviewSheet(B2C_Quarterly_and_Annual_Rev_csv_path, B2C_Quarterly_and_Annual_Rev_xlsx_path)






def addDataToCompleteInfoList(complete_info):
    ## adds the new flag fields to the existing complete info

    # if '87201298-LATIS20036872' in ticket_analysis_dict:
    #     print("found ittttt!!")

    for row in complete_info:
        #print("\n\nBefore row addition: ", len(row))
        worksheet_key_compare = row[15] + str(row[1])

        #if worksheet_key_compare in ticket_analysis_dict:
            #print("hot diggity damn")
        
        row.extend(ticket_analysis_dict[worksheet_key_compare])
        #print("\n and after: ", len(row))

        #print("\n\nhere's the new row: ", row, "and length: ", len(row), "\n")
    
    # print("\n\n\n all the info here!! : ", complete_info)

    return complete_info

        



def addDataValidationToExcelSheet(path):
        # Load the Excel file
    df = pd.read_excel(path)

    # Insert a new column at column A (index 0) ### 
    df.insert(0, "DECISION", '')

    num_rows = df.shape[0]

    # prepare the final row so that data validation only runs to the bottom of the available data
    final_row = "A" + str(num_rows + 1)

    # Save the modified DataFrame back to the Excel file
    df.to_excel(path, index=False)


    
        #this workbook doesn't exist yet - it gets written in main()
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active

    
    dv = DataValidation(type="list", formula1='"APPROVED, DENIED, NO ACTION, FEEDBACK REQUIRED, UNDER REVIEW"')

    # Add the data validation object to the worksheet
    worksheet.add_data_validation(dv)

    # Apply the data validation to column A
    dv.add(f'A1:{final_row}')

    # Save the workbook

    workbook.save(path)
    workbook.close() 



            




    '''
    for index in b2c_quarterly_review_nums:
        #print("\n\n\nhere is the row length: ", len(row))
        #print("\nand the index being stored to book before written to file: " , index, " : ", row[index], "\n")
        #print("\nand the corresponding row index: \n", row[index])
        book.append(row[index])
    '''






def main():
    global can_map_list
    global cannot_map_list
    global current_date
    global can_map_bans_list
    global cannot_map_bans_list
    global oracle_query
    global oracleBillingAccountsODSidsString
    global mdm_query_results

    global analyzed_bulk_requests_dict
    global bulk_ticket_num_list

    global mdm_cust_results
    global mdm_fan_results

    global mdm_Cust_list
    global mdm_FANs_list

    global mdm_fan_results_list
    global mdm_cust_results_list

    global voice_services

    global billAcctActvtyStatusCd
    # 14= fan 11 and 55 are reqstd cust -- this is all we need for AM
    # add one (since the complete list will have the suggestion added to the front
    # add two to the normal COMPLETE file index, since we are adding two fields during evaluation below (before writing to file)

    global containsMSG

    #analysis_sheet_nums = [0, 78, 1, 15, 25, 54, 28, 34, 35, 36, 56, 61, 62, 63, 29, 30, 31, 32, 33, 57, 58, 59, 60, 22, 73, 74, 75, 9, 10, 14, 1, 15, 56, 25, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 71]
    analysis_sheet_nums = [0, 86, 1, 2, 16, 26, 27, 58, 29, 31, 32, 38, 39, 40, 65, 66, 69, 70, 71, 33, 35, 36, 37, 65, 66, 67, 68, 23, 81, 82, 83, 10, 11, 15, 2, 16, 61, 63, 27, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 60, 59, 79]
    #analysis_sheet_nums = [0, 1, 14, 24, 53, 27, 33, 34, 35, 55, 60, 61, 62, 28, 30, 31, 32, 56, 57, 58, 59, 21, 0, 14, 55, 24, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 70]
    #ticket_analysis_sheet_nums = [0, 76, 1, 15, 25, 54, 28, 34, 35, 36, 56, 61, 62, 63, 29, 31, 32, 33, 57, 58, 59, 60, 22, 73, 74, 75, 6, 1, 15, 56, 25, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 71]
    ## --> 0 and 1 are computed fields created just before writing to csv, so everything moves up by 3 - including ticket # (which was at 0 and now at 3)
    ticket_analysis_sheet_nums = [0, 110, 111, 1, 11, 2, 105, 3, 18, 28, 68, 65, 29, 60, 33, 34, 40, 35, 37, 38, 67, 41, 42, 71, 72, 84, 80, 81, 82, 96, 85, 86, 22, 104, 105, 106, 102, 100, 126, 8, 3, 18, 63, 71, 29, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 62, 61, 94, 127, 112, 113, 114, 115, 116, 117, 118, 119, 120, 121, 122, 123, 124, 125]
              #MDM               [0, 79, 1, 15, 25, 26, 56, 29, 30, 36, 37, 38, 58, 64, 65, 66, 31, 33, 34, 35, 60, 61, 62, 63, 22, 76, 77, 78, 9, 10, 14, 1, 15, 58, 26, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 74]
              #index nums         1,  2, 3,  4,  5,  6,  7,  8,  9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26,27, 28, 29,30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46   

    getLast13Months(current_date)
    lDriveTest()
    ssms_query_results = querySSMS(query)

    #---> removed MDM results for now because the set was over 100k -- this was where the OG MDM query was -- as it was only intended
    ## ---> to bring in all the recently mapped accounts in MDM so we could compare the results of that engine with
    ### ---> what the automation would actually show. IN other words, was the mapping that MDM made accurate? 
    ## mdm_query_results = queryMDM()

    #have to call main() since the file has "if __name__ == "__main__"" which means that if the file is called 
    #as a module then it won't just queue up and run everything inside. It will only run if called directly
    # as in python bmp.py (essentially) 
    
    bmp_results = bmp.main() 

    containsMSG = bmp.containsMSG #does the ticket contain a .msg file?

    full_query_results = ssms_query_results  + bmp_results #+ mdm_query_results ## add in the bulk tickets data here


    print(f"\n\nand now we get revenue for each ban")
    
    mdm_data_done = getMDMdata(full_query_results)
    
    # print("\n\nhere is the mdm_data_done: ", mdm_data_done) 


    revenue_results = getBanRevenue(full_query_results)



    
    #for x in revenue_results:
     #   print("\n\nso, here is the length of x: ", len(x))

        
    rev_and_ban_lob_results = getBanAgeAndLOB(revenue_results)
    #for entry in rev_and_ban_lob_results:
        #print("\nhere is the entry from rev_and_ban_lob_results: " , entry)
    #print("here are the rev_and_ban_lob_results: " , rev_and_ban_lob_results)


    
    all_LOBinfo_and_revenue = getToBusOrgLOBInfo(rev_and_ban_lob_results)
    #for entry in all_LOBinfo_and_revenue:
        #print("\nhere is the entry from all_LOBinfo_and_revenue:", entry)

    all_previous_data_and_requester_info = getRequesterInfo(all_LOBinfo_and_revenue)


    #for item in all_LOBinfo_and_revenue:
        #print("here is the item in allLOBinfo: ", item)
        
    complete_info = compareLOBs(all_LOBinfo_and_revenue)
    complete_info = compareDWSecure(complete_info)

    
    
    #queryOracle_PLSQL()
    getKenanSrvcStatusAndOrderSystemsAndVoiceNameCheck(complete_info)
    
    createErrorAndMappingCSVs(complete_info)
    #print("\n\nhere is the green_can_map_list in main(): ", green_can_map_list)
    #print(rev_and_ban_lob_results)
    #getRevFANsFromSSMSList(revenue_results, results)

    # print("here's all fans: ", allFans)
    getARdata(allFans)
    getB2CQuarterlyData(allFans)
    getLatestRqstrComment(allTickets)
    getLatestTeamComment(allTickets)
    enterpriseVsMassMarkets_ProductCheck(allFans, complete_info)
    find_CRIS_to_ENS_BAN_data(ens_BANs)
    buildQuarterlyReviewSheet(complete_info)

    complete_info = addDataToCompleteInfoList(complete_info)



    ## ----> dw secure company code  / 1= lumen 2= latam, 4 is brightspeed
    # DW_SECURE_COMPANY_NBR


    
    '''
    14 = FAN
    11 = RQST_CUST
    '''
    if len(green_can_map_list) > 0:
        if os.path.isfile(green_BAN_to_DL_csv_path):
            with open(green_BAN_to_DL_csv_path, 'a+', newline='') as out:
                csv_out = csv.writer(out)
                for result in green_can_map_list:
                    csv_out.writerow(result)
                print("wrote out the CAN MAP results to a new csv")
        else:
            with open(green_BAN_to_DL_csv_path, 'a+', newline='') as out:
                csv_out = csv.writer(out)
                csv_out.writerow([
                                'CHIT_TICKET_NUM', 'FINANCE_ACCOUNT_NBR', 'RQSTD_CUST_NBR', 'NAME', 'BILLING_ACCOUNT_ID__C',
                                'BILLING_UNIQUE_EXTERNAL_ID__C', 'BILLING_SYSTEM__C', 'BILLING_ADDRESS_LINE_1__C',
                                'BILLING_ADDRESS_LINE_2__C', 'BILLING_CITY__C', 'BILLING_STATE__C',
                                'BILLING_POSTAL_CODE__C', 'BILLING_COUNTRY__C', 'BAN_DW_SECURE_COMPANY_NBR',
                                'RECORDTYPEID', 'CUSTOMER_ACCOUNT__C', 'EXPLANATION'
                                ])
                for result in green_can_map_list:
                    csv_out.writerow(result)
                print("wrote out the CAN MAP results to a new csv")

    if len(red_can_map_list) > 0:
        print("\n\nred list is filled yo\n\n")
        if os.path.isfile(red_BAN_csv_path):
            with open(red_BAN_csv_path, 'a+', newline='') as out:
                csv_out = csv.writer(out)
                for result in red_can_map_list:
                    csv_out.writerow(result)
                print("wrote out the RED CAN MAP results to a new csv")
        else:
            with open(red_BAN_csv_path, 'a+', newline='') as out:
                csv_out = csv.writer(out)
                csv_out.writerow([
                                'CHIT_TICKET_NUM', 'FINANCE_ACCOUNT_NBR', 'RQSTD_CUST_NBR', 'EXPLANATION'
                                ])
                for result in red_can_map_list:
                    csv_out.writerow(result)
                print("wrote out the RED CAN MAP results to a new csv")
            
    if len(complete_info) > 0:   
        if os.path.isfile(complete_data_csv_path):
            with open(complete_data_csv_path,'a+', newline='') as out: ## w+ to read and write (just in case) a+ to append (read and write)
                csv_out = csv.writer(out)
                ##for result in results:
                for result in complete_info:
                    csv_out.writerow(result)
                print("wrote out the complete_data list to an existing csv")
                
            
        else:
            with open(complete_data_csv_path,'a+', newline='', encoding="utf-8") as out: ## w+ to read and write (just in case) a+ to append (read and write)
                csv_out = csv.writer(out)
                print("ban LIST filled - did not find a csv exists, so going to print to csv")
                csv_out.writerow([
                                'SUGGESTION',
                                'TICKET_ID', 'RQSTR_EMP_NBR', 'RQSTR_FIRSTNAME', 'RQSTR_LASTNAME',
                                'ISSUE_CLASS_ID', 'ISSUE_CLASS_NAME', 'ISSUE_DESCRIPTION', 'STATUS_ID', 'STATUS_NAME',
                                'SOURCE_CUST_NBR', 'BILL_ACCOUNT_NBR', 'RQSTD_CUST_NBR', 'RQSTD_CUST_NAME', 'RQSTD_CUST_SALES_CHAN',
                                'FINANCE_ACCOUNT_NBR', 'CRRNT_SALES_CHAN', 'ROOT_CAUSE', 'FAN_HAS_REVENUE',
                                'FAN_CRRNT_MONTH_REV',
                                'FAN_AVG_LAST_3_MONTHS_REV',
                                'FAN_AVG_LAST_6_MONTHS_REV',
                                'FAN_AVG_LAST_9_MONTHS_REV',
                                'FAN_AVG_LAST_12_MONTHS_REV',
                                'FAN_AVG_LAST_13_MONTHS_REV',
                                'FAN_FROM_CODSBA',
                                'BAN_ODS_ID',
                                'BILL_ACCOUNT_NAME',
                                'BA_SOURCE_SYSTEM',
                                'DW_CREATE_DT',
                                'BAN_AGE', 'CRRNT_CUST_NBR',
                                'CRRNT_CUST_ODS_ID',
                                'CRRNT_CUST_LOB',
                                'CURRNT_CUST_LOB_GRP',
                                'CRRNT_CUST_EXTRNL_RPRT_SALES_CHANNEL',
                                'CRRNT_CUST_SALES_CHAN',
                                'CRRNT_CUST_SALES_SUB_CHAN',
                                'CRRNT_CUST_NAME',
                                'CRRNT_CUST_DW_SECURE_COMPANY_NBR',
                                'CRRNT_CUST_STATUS',
                                'CRRNT_CUST_HOUSE_ACCOUNT_CD',
                                'CRRNT_CUST_OWNER_ID',
                                'CRRNT_CUST_OWNER_EMAIL',
                                'CRRNT_CUST_OWNER_NAME',
                                'CRRNT_CUST_OWNER_COMPANY',
                                'CRRNT_CUST_OWNER_DEPARTMENT',
                                'CRRNT_CUST_OWNER_TITLE',
                                'BILL_ACCT_ACTVTY_STATUS_CD', 'BILLING_ACCOUNT_ID__C',
                                'BILLING_UNIQUE_EXTERNAL_ID__C', 'BILLING_SYSTEM__C',
                                'BILLING_ADDRESS_LINE_1__C', 'BILLING_ADDRESS_LINE_2__C', 'BILLING_CITY__C', 'BILLING_STATE__C',
                                'BILLING_POSTAL_CODE__C', 'BILLING_COUNTRY__C', 'BAN_DW_SECURE_COMPANY_NBR',
                                'RECORDTYPEID',
                                'BILLING_INSTANCE__C',
                                'CBP',
                                'BILL_ACCOUNT_LEVEL_TYP',
                                'BA.BILL_CYCLE_CD',
                                'BA.BILL_ACCT_ACTIVE_STATUS_IND',
                                'CURRNT_CUST_ASSIGNMENT_GROUP__C', 'CRRNT_CUST_TRDNG_PRTNR',
                                'BILL_ACCT_CUST_TYP', 'BILL_ACCT_CUST_SUB_TYP',
                                'RQSTD_CUST_NBR_CRPL',
                                'RQSTD_CUST_ODS_ID',
                                'RQSTD_CUST_DUNS',
                                'RQSTD_CUST_NAICS_CD',
                                'RQSTD_CUST_ADDRESS_LINE_1',
                                'RQSTD_CUST_CITY',
                                'RQSTD_CUST_STATE',
                                'RQSTD_CUST_POSTAL_CD',
                                'RQSTD_CUST_COUNTRY',
                                'RQSTD_CUST_LOB',
                                'RQSTD_CUST_EXTRNL_RPRT_SALES_CHANNEL',
                                'RQSTD_CUST_SALES_CHAN',
                                'RQSTD_CUST_SALES_SUB_CHAN',
                                'RQSTD_CUST_NAME',
                                'RQSTD_CUST_DW_SECURE_COMPANY_NBR',
                                'RQSTD_CUST_STATUS',
                                'RQSTD_CUST_HOUSE_ACCOUNT_CD',
                                'RQSTD_CUST_OWNER_ID',
                                'RQSTD_CUST_OWNER_EMAIL',
                                'RQSTD_CUST_OWNER_NAME',
                                'RQSTD_CUST_OWNER_COMPANY',
                                'RQSTD_CUST_OWNER_DEPARTMENT',
                                'RQSTD_CUST_OWNER_TITLE',
                                'RQSTD_CUST_ID', 'RQSTD_CUST_LOB_GROUP', 'RQSTD_CUST_ASSIGNMENT_GROUP__C', 'RQSTD_CUST_TRDNG_PRTNR',
                                'RQSTR_EMPLOYEE_NBR', 'RQSTR_JOB_FAMILY', 'RQSTR_JOB_TITLE', 'RQSTR_MANAGER_ID', 'RQSTR_MANAGER_NAME', 'RQSTR_MANAGER_JOB_TITLE',
                                'CROSS_LOB', 'CROSS_SALES_CHANNEL', 'CROSS_BU',
                                'SECURE_CD_MATCH',
                                'KENAN_BAN_ACTIVE_SRVCS_NO_BILLING',
                                'ASSOCIATED_TO_BAD_OE_SYS',
                                'FINAL_EVALUATION',
                                'TICKET_COMMENT',
                                'RQSTD_CUSTOMER_STATUS', 'RQSTD_EQUALS_CRRNT_CUST', 'MM_ENT_MOVEMENT_CHECK', 'SECURE_COMPANY_NBR', 'VOICE_SERVICES', 'ORDER_ENTRY_CHECK', 'MOVING_FROM_PLACEHOLDER_FLAG', 'TTM_REVENUE', 'COLLECTIONS_AR',
                                'MOVING_TO_PLACEHOLDER', 'TRADING_PARTNER_CHECK', 'ASSIGNMENT_GROUP', 'BUSINESS_UNIT', 'CHANNEL'
                                  ]) # headers
                for result in complete_info:
                    csv_out.writerow(result)
        ## --> create tickets analysis template
        if os.path.isfile(tickets_analysis_template_path):
            with open(tickets_analysis_template_path, 'a+', newline='') as out:
                csv_out = csv.writer(out)
                book = []
                for index in analysis_sheet_nums:
                    for row in complete_info:
                        if row[0] != 'MDM MATCHING':
                            if row[14] in can_map_bans_list:
                                print("\n\n\nhere is the row \nthe row that might have an issue: ", row)
                                book.append('Map it!')
                                #print(f"\n\nhere's the index we're attempting to add to {row[14]} ", index)
                                #print(f"and here's the row: ", row)
                                #print(f"and the row count: ", len(row))
                                book.append(row[index])
                            #print("\n\n\nhere's the index!!: ", index , ' ', row[index])
                        csv_out.writerow(book)
                df = pd.read_csv(tickets_analysis_template_path, encoding = "ISO-8859-1", sep=',', names=['ZERO', 'ONE', 'TWO', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16',
                                                                     '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32',
                                                                     '33', '34', '35', '36'])
            for ci_row in complete_info:
                if ci_row[14] in can_map_bans_list:
                    df.insert(0, 'SUGGESTION', 'Map it!')
        else:
            for row in complete_info:
                #print("\n\nwrapping things up - here's the row before we write to file", len(row), row)
                fan = row[15]
                rqstd_cust = row[69]
                action = row[0]
                comment = row[109]
                kept_book = []
                ticketType = row[9]
                kept_book.append(fan)
                kept_book.append(rqstd_cust)
                kept_book.append(comment)
                #kept_book.append(book)
                if row[9] == 'BULK':
                    #print("\n\n bulk ban request here \n\n")
                    bulk_ticket_num = row[1]
                    if bulk_ticket_num not in bulk_ticket_num_list:
                        bulk_ticket_num_list.append(bulk_ticket_num)
                    #print("the bulk ticket num = ", bulk_ticket_num)
                    if bulk_ticket_num not in analyzed_bulk_requests_dict.keys():
                        analyzed_bulk_requests_dict[bulk_ticket_num] = [] ## if the ticket (as a key) doesn't exist yet, then create that key with a list value (so we can make a list of lists, babay!)
                        analyzed_bulk_requests_dict[bulk_ticket_num].append(kept_book)
                    else:
                        analyzed_bulk_requests_dict[bulk_ticket_num].append(kept_book) ## else, if it already exists, then append the latest list (row) to that existing list (of lists, babay!)
                    
                    
            if os.path.exists(LDrive_BULK_templates_path):
                for filename in os.listdir(LDrive_BULK_templates_path):
                    bulk_request_attachment_file_path = LDrive_BULK_templates_path + r"\\" + filename
                    completed_bulk_request_attachment_file_path = LDrive_BULK_templates_path + r"\\" + "ANALYZED__" + filename
                    ticketNUM_from_filename = int(filename[0:8]) # change to int to actually find the damn key
                    # print("here is the dict", analyzed_bulk_requests_dict)
                    # print("here is the dict list at ticketnum", analyzed_bulk_requests_dict[ticketNUM_from_filename])
                    # print("\n\n here is the ticket num from file name:" , ticketNUM_from_filename)
                    if ticketNUM_from_filename in analyzed_bulk_requests_dict.keys(): # == str(bulk_ticket_num):
                        print("\n\n bulk ban ticket found here \n\n")
                        """
                        with open(bulk_request_attachment_file_path, 'r', newline='', encoding='utf-8-sig') as opened_file:
                            read_file = csv.reader(opened_file)
                            next(read_file, None) ## skip the headers
                            for filerow in read_file:
                                fan_from_filerow = filerow[0]
                                #print("here is the filerow fan", fan_from_filerow)
                                #print("\nand here is the filerow:" , filerow)
                                if fan == fan_from_filerow:
                                    book = []
                                    book.append(filerow[0])
                                    book.append(filerow[1])
                                    book.append(comment)
                                    kept_book.append(book)
                                    #filerow.insert(2, f'{action}')
                        opened_file.close() 
                        """

                        working_list = analyzed_bulk_requests_dict[ticketNUM_from_filename]
                        # print("here is the working list", working_list)
                        with open(completed_bulk_request_attachment_file_path, 'w', newline='', encoding='utf-8-sig') as outfile: 
                            assert os.path.isfile(completed_bulk_request_attachment_file_path)
                            os.chmod(completed_bulk_request_attachment_file_path, stat.S_IRUSR | stat.S_IRGRP | stat.S_IROTH | stat.S_IWUSR | stat.S_IWGRP | stat.S_IWOTH) #to make sure we can access the file to write to
                            write_file = csv.writer(outfile, delimiter=',')
                            

                            for list in working_list:
                                write_file.writerow(list)
                                # print("wrote this row", list)
                                

                        outfile.close()

            with open(tickets_analysis_template_path, 'a+', newline='', encoding="utf-8") as out:
                csv_out = csv.writer(out)
                book2 = []
                for row in complete_info:
                    #print("\nhere is the row before further index adjustments: ", row, "\n")

                    fan = row[15]
                    rqstd_cust = row[69]
                    cods_currnt_cust = row[31]
                    cods_currnt_cust_status = row[40]
                    rqstd_cust_status = row[84]
                    ticket = row[1]

                    billAcctActvtyStatusCd = row[48]

                    action = row[0]
                    book = []
                    ticketType = row[9]
                    #print("\n\n here is the tickettype", ticketType)
                    # print("\n\nalmost there!!\nfound the fan and rqstd_cust: ", fan, rqstd_cust)
                    
                                
                                    

                    if row[0] != 'MDM MATCHING':
                        #print("here's the book right now: ", book)
                        #print("here's the row: ",  row)
                        #print("here's the fan in the row: ", row[14])
                        ''' # moved inside the decision waterfall
                        if row[14] in can_map_bans_list:
                            row.insert(0, "Map it!")
                        elif row[14] in research_list:
                            row.insert(0, "Review")
                        elif row[14] in definitely_map_these_mamma_jammas_list:
                            row.insert(0, "Automap")
                        else:
                            row.insert(0, "Nope")
                        #print("n\n\nhere is the completed row including the suggestion column: ", row)
                        '''

                        ## --> concat and append the requester name
                        #print(f"Here are row3 and row4 and row5 values for {row[15]}: {row[3]}, {row[4]}, {row[5]}")
                        # print(len(row))
                        row.append(row[3] + ' ' + str(row[4]))
                        

                        this_fan = row[15]  
                        # AGAIN, 1 more index past the original in COMPLETED file because we insert the mapping suggestion at [0] just before this  
                        # print("\n\nhere is muhfuggin source system and fuggin cbp, muhfugga, healzyeah!", row[27], row[60], "and fan", row[15])  
                        sourceSystem = row[28]
                        CBP = row[61]
                        if (sourceSystem in salesforce_billing_systems or (sourceSystem in ['IDC_KENAN', 'KENANFX'] and re.match('.*\-LATIS$', this_fan))):
                            '''
                            if row[27] in ['IDC_KENAN', 'KENANFX'] and re.match('.*\-LATIS$', row[15]):
                                print("found ban in salesforce systems", this_fan)
                                row.insert(1, "SF/AM")
                            else:
                                print("found ban in salesforce systems", this_fan)
                                row.insert(1, "SF")
                            '''
                            #print("found ban in salesforce systems", this_fan)
                            row.insert(1, "MDM")
                        elif (sourceSystem in am_source_systems) and (CBP is not None):
                            #print("found ban in am systems", this_fan)
                            row.insert(1, "MDM / AM")
                        elif (sourceSystem in am_source_systems) and (CBP is None):
                            #print("found ban in am systems but row 60 is none", this_fan)
                            if re.search('^.*-GC$', fan):
                                #print("found ban in am systems as gc ban not in am", this_fan)
                                row.insert(1, "CPM")
                            else:
                                #print("found ban in am systems but no cbp", this_fan)
                                row.insert(1, "NO CBP")
                                row[110] = "The billing account may exist in AM, but its CBP number hasn't flowed to CODS. " + row[110]
                                row[111] = "This billing account's CBP number is NULL in CODS, so we have to map it manually in AM. " + row[111]
                                # row[0] = "Denied"
                        else:
                            #print("found ban nowhere", this_fan)
                            row.insert(1, "UNK")


                        #print("row 15:", row[15])
                        #print("row 14:", row[14])

                        #insert data at index 2 to save place for ORACLE hierarchy Parent ODS ID data
                        row.insert(2, "")

                        '''
                        for index in ticket_analysis_sheet_nums:
                            #print("\n\n\nhere is the row length: ", len(row), "\nand the index being printed: \n" , index, " -- " , "\nand the corresponding row index: \n", row[index], "\nand the row itself: \n", row)
                            book.append(row[index])

                        #print("\n\n\nhere's the book right now (at the bottom): ", book)
                        book2.append(book)
                        '''


                            
                        
                        #print("\n\n\nand here's book 2 -- this should get stacked full of lists: ", book2)
                            #print("\n\n\nhere's the index!!: ", index , ' ', row[index])
                    else:
                        continue
                    # print("\n\nand awwaaaaaaay we go! here's fand and the source system: ", fan,  sourceSystem)
                    # print("\n\nhere is the MDM_FANs_list and MDM_CUST_list: ", mdm_FANs_list, mdm_Cust_list)  

                    possible_actions = {
                        1 : "the Requested customer is not Approved in MDM",
                        2 : "something"
                    }
                    #print("rikkitikkitavi: here is the fan we are hunting and the mdmd fans list: ", fan, mdm_fan_results_list)
                    if fan in mdm_FANs_list and rqstd_cust in mdm_Cust_list and (sourceSystem in salesforce_billing_systems) or (sourceSystem in ['IDC_KENAN', 'KENANFX'] and re.match('.*\-LATIS$', fan)):
                        # print(f"\n\nFOUND THE FAN!!! in the MDM list\n: ", fan )


                        # fan = row[15]
                        # rqstd_cust = row[64]
                        # cods_currnt_cust = row[30]
                        new_text = ''
                        fan_dw = ''
                        cust_dw = ''
                        cust_stat = ''
                        final_text = ''
                        can_map = ''
                        for fan_row in mdm_fan_results_list:
                            fan_secure_cd = fan_row[2]
                            mdm_fan = fan_row[0]
                            mdm_cust = fan_row[1]
                            # print("heres the fan we're hunting", fan_row[0])
                            if fan == mdm_fan:
                                # print("found fan in mdm fan_row data and secure", fan,fan_row)
                                if fan_secure_cd == '1':
                                    new_text = "FAN in MDM, DWSecure == 1"
                                    fan_dw = 1
                                else:
                                    new_text = f"FAN in MDM, DWSecure == {fan_secure_cd}"
                                    can_map = 'no'
                                if mdm_cust == rqstd_cust and rqstd_cust != cods_currnt_cust:
                                    cust_text = f"The MDM and Requested customer match"
                                    decision = f"no action"
                                else:
                                    decision = ''
                        for cust_row in mdm_cust_results_list:
                            if rqstd_cust == cust_row[0]:
                                # print("found cust in mdm cust_row data", rqstd_cust, cust_row)
                                if rqstd_cust == cpniIssuesCust:
                                    pass
                                else:
                                    if cust_row[2] == '1':
                                        if cust_row[1] == 'Approved':
                                            new_text = new_text + f" | Customer in MDM, DWSecure == 1, Approved"
                                            cust_dw = 1
                                        elif cust_row[1] != 'Approved':
                                            new_text = new_text + f" | Customer in MDM, DWSecure == 1, not Approved"
                                            can_map = 'no'
                                    elif cust_row[2] != '1':
                                        if cust_row[1] == 'Approved':
                                            new_text = new_text + f" | Customer in MDM, DWSecure == {cust_row[2]}, Approved"
                                            can_map = 'no'
                                        elif cust_row[1] != 'Approved':
                                            new_text = new_text + f" | Customer in MDM, DWSecure == {cust_row[2]}, not Approved"
                                            can_map = 'no'
                        if new_text == '':
                            # print("\n\nand awwaaaaaaay we go! here's fan and the source system: ", fan,  sourceSystem)
                            row.append('Map in MDM')
                        else:
                            row.append(new_text)

                        if can_map == 'no' and decision == '':
                            row[110] = "The requested mapping cannot be completed in MDM " + row[110]
                            row[0] = "Denied"
                        elif can_map == 'no' and decision == 'no action':
                            row[110] = "The billing account is already mapped as desired in MDM" + row[110]
                            row[111] = f"This billing account is already mapped to {mdm_cust} in MDM. If the mapping isn't showing in other systems, please reach out to dl_mdm@lumen.com to find out what the dataflow issue might be "
                            row[0] = "Denied"
    
                    #BILL_ACCT_ACTVTY_STATUS_CD
                    elif fan not in mdm_FANs_list:
                        if sourceSystem in salesforce_billing_systems:
                            if rqstd_cust in mdm_Cust_list:
                                row.append('FAN not in MDM')
                                if billAcctActvtyStatusCd != 'Active':
                                    row[110] = f"This FAN does not exist in MDM because its BILL_ACCT_ACTVTY_STATUS_CD is {billAcctActvtyStatusCd} instead of 'Active'. We cannot fix it." + row[110]
                                    row[111] = f"This FAN does not exist in MDM; therefore the Data Governance team is unable to remap it. Please reach out to dl_mdm@lumen.com for help with this issue."
                                    row[0] = "Denied"
                                else:
                                    row[110] = "The requested mapping cannot be completed in MDM - the FAN does not exist there - we do not know WHY it doesn't exist " + row[110]
                                    row[111] = f"This FAN does not exist in MDM, so it can't be mapped."
                                    row[0] = "Review"
                            elif rqstd_cust not in mdm_Cust_list:
                                row.append('FAN and Requested cust not in MDM')
                                if billAcctActvtyStatusCd != 'Active':
                                    row[110] = f"The Requested cust and the FAN do not exist in MDM. The {fan} does not exist in MDM because its BILL_ACCT_ACTVTY_STATUS_CD is {billAcctActvtyStatusCd} instead of 'Active'. We cannot fix it." + row[110]
                                    row[111] = f"Denied. The Requested customer and the FAN do not exist in MDM where they would need to exist in order for us to map them. We cannot govern this request."
                                    row[0] = "Denied"
                                else:
                                    row[110] = "The requested mapping cannot be completed in MDM - neither the FAN nor the Requested customer exist there. " + row[110]
                                    row[0] = "Denied"
                        else:
                            row.append("AM billing account - map in MDM and in AM")
                    elif fan in mdm_FANs_list and (rqstd_cust not in mdm_Cust_list and sourceSystem in salesforce_billing_systems) \
                        or (sourceSystem in ['IDC_KENAN', 'KENANFX'] and re.match('.*\-LATIS$', row[15])):
                        # print("\n\nDone found 'er! the fan and source system for mdm: ", sourceSystem, fan, "\n\n")
                        # print("\n\nTARDIGRAVE: here is the Cods currnt cust, the currnt cust status and rqstd cust: ", cods_currnt_cust, rqstd_cust_status, rqstd_cust, "\n\n")
                        # print("/nTARDIGRAVE - the fan and the requested cust: ", fan, rqstd_cust)
                        if rqstd_cust_status != 'Approved':
                            # print("/nTARDIGRAVE - the fan and the requested cust: ", fan, rqstd_cust)
                            row.append('Requested cust not in MDM')
                            row[110] = "This mapping cannot be completed in MDM - the Requested customer does not exist there. " + row[110]
                            ##### --> not including any data already appended to 104 because this resolution takes precedence -- might need to change to consider all potential outcomes in waterfall above
                            row[111] = "Feedback Required: The Requested customer does not exist in MDM because it is not Approved; therefore, the billing account cannot be mapped to it. Please update this ticket once the customer is Approved." 
                            row[0] = "Feedback Required"
                            
                        else:
                            print("\n\nOSCELOT BRIGME YOUNG MCMASTER - it's here.")
                            row.append('Requested cust not in MDM')
                            row[110] = "This mapping cannot be completed in MDM - the Requested customer does not exist there. Since we don't know WHY, we need to find out. " + row[110]
                            row[111] = "Feedback Required: The Requested customer does not exist in MDM because it is not Approved; therefore, the billing account cannot be mapped to it. Please update this ticket once the customer is Approved." 
                            row[0] = "Review"
                    else:
                        row.append('AM billing account - map in MDM and in AM')

                    
                    if ticket in containsMSG:
                        row[0] = "Email / " + row[0]

                    
                    #if fan in arDataResultsList:
                        #row.append('Falk')
                    #else:
                        #row.append('Palk')


                    for index in ticket_analysis_sheet_nums:
                        # print("offending row: ", row[3])
                        # print("\n\n\nhere is the row length: ", len(row))
                        # print("\nand the next index is: ", index)
                        # print("\nand the index being stored to book before written to file: " , index, " : ", row[index], "\n")
                        # print("\nand the corresponding row index: \n", row[index])
                        book.append(row[index])

                    #print("\n\n\nhere's the book (row) right now -- after being appended according to the ticket analysis index #s: ", book)
                    book2.append(book)


                #print("\n\n\n\nhere's the book2: ", book2)

                """        
                for book_row in book:
                print(book_row)
                book2.append([book_row])
                """
                csv_out.writerow(['SUGGESTED_ACTION', 'EVALUATION', 'COMMENT_4_TICKET',
                                  'SF_OR_AM', 'TICKET_TYPE',
                                  'ORACLE_FAMILY_ODS_ID',
                                  'SECURE_CD_MATCH',
                                  'TICKET_NUM', 'FINANCE_ACCOUNT_NBR',
                                  'BAN_ODS_ID',
                                  'BILL_ACCT_ACTIVE_STATUS_IND',
                                  'BILL_CYCLE_CD',
                                  'BILL_ACCOUNT_NAME', 'BAN_DW_SECURE_COMPANY_NBR',
                                  'CRRNT_CUST_NBR',
                                  'CRRNT_CUST_ODS_ID',
                                  'CRRNT_CUST_NAME',
                                  'CRRNT_CUST_LOB', 'CRRNT_CUST_EXTRNL_RPRT_SALES_CHANNEL', 'CRRNT_CUST_SALES_CHAN', 'CRRNT_CUST_ASSGNMNT_GRP',
                                  'CRRNT_CUST_DW_SECURE_COMPANY_NBR', 'CRRNT_CUST_STATUS',
                                  'RQSTD_CUST_NBR',
                                  'RQSTD_CUST_ODS_ID',
                                  'RQSTD_CUST_NAME',
                                  'RQSTD_CUST_LOB', 'RQSTD_CUST_EXTRNL_RPRT_SALES_CHANNEL', 'RQSTD_CUST_SALES_CHAN', 'RQSTD_CUST_ASSGNMNT_GRP',
                                  'RQSTD_CUST_DW_SECURE_COMPANY_NBR', 'RQST_CUST_STATUS',
                                  'BAN_AVG_LAST_3_MONTHS_REV', 'CROSS_LOB', 'CROSS_SALES_CHANNEL', 'CROSS_BU',
                                  'RQSTR_MANAGER_NAME', 'RQSTR_JOB_TITLE', 'RQSTR_NAME',
                                  'COMMENTS',
                                  'TICKET_ID', 'FINANCE_ACCOUNT_NBR',
                                  'CBP',
                                  'CUST_NBR', 'NAME',
                                  'BILLING_ACCOUNT_ID__C', 'BILLING_UNIQUE_EXTERNAL_ID__C', 'BILLING_SYSTEM__C', 'BILLING_ADDRESS_LINE_1__C',
                                  'BILLING_ADDRESS_LINE_2__C', 'BILLING_CITY__C', 'BILLING_STATE__C', 'BILLING_POSTAL_CODE__C', 'BILLING_COUNTRY__C',
                                  'BAN_DW_SECURE_COMPANY_NBR',
                                  'BILLING_INSTANCE__C',
                                  'RECORDTYPEID', 'CUSTOMER_ACCOUNT__C',
                                  'MDM_DATA', 
                                  'RQSTD_CUSTOMER_STATUS', 'RQSTD_EQUALS_CRRNT_CUST', 'MM_ENT_MOVEMENT_CHECK', 'SECURE_COMPANY_NBR', 'VOICE_SERVICES', 'ORDER_ENTRY_CHECK', 'MOVING_FROM_PLCHLDR_OR_BLANK', 'TTM_REVENUE', 'COLLECTIONS_AR',
                                  'MOVING_TO_PLACEHOLDER', 'TRADING_PARTNER_CHECK', 'ASSIGNMENT_GROUP', 'BUSINESS_UNIT', 'CHANNEL'])
                for book_item in book2:
                    #print("\n\n and here's the book (row) after all changes and just before being written to file: ", book_item, "\n\n")
                    csv_out.writerow(book_item)
                out.close() ## --> file needs to be closed to begin work on changing it with pandas (below)

            #convert csv to xlsx doc
            filepath_in = tickets_analysis_template_path
            filepath_out = tickets_analysis_template_xlsx_path
            #this_file = pd.read_csv(filepath_in, delimiter=",", encoding="latin1")

            # change CSV to XLSX
            xlsxFile = pd.read_csv(filepath_in, delimiter=",", on_bad_lines="error")
            xlsxFile.to_excel(filepath_out, index=False)

            #create new worksheet
            workbook = openpyxl.load_workbook(tickets_analysis_template_xlsx_path)
            workbook.create_sheet('ORACLE')
                        
            #save the file to make sure the changes are... ahem... saved
            workbook.save(tickets_analysis_template_xlsx_path)

            if oracleBillingAccountsODSidsString != '':
                
                dsn_tns = cx_Oracle.makedsn(r'RACORAP16-SCAN.IDC1.LEVEL3.COM', '1521', service_name='CDW01P_USERS')
                conn = cx_Oracle.connect(user='AC79386', password='FlimFlamNon2241~$', dsn=dsn_tns)

                oracle_query = f"""
                                WITH TABLE1 AS (
                                                SELECT
                                                    BA.PARENT_BILL_ACCOUNT_ODS_ID
                                                    ,(CASE WHEN BA.PARENT_BILL_ACCOUNT_ODS_ID IS NOT NULL THEN BA.PARENT_BILL_ACCOUNT_ODS_ID
                                                          WHEN BA.PARENT_BILL_ACCOUNT_ODS_ID IS NULL THEN BA.BILL_ACCOUNT_ODS_ID END) AS PARENT_BILL_ACCOUNT_ODS_ID2
                                                    ,BA.BILL_ACCOUNT_ODS_ID
                                                    ,BA.BILL_ACCOUNT_NBR
                                                    ,BA.FINANCE_ACCOUNT_NBR
                                                    ,BA.DW_SOURCE_SYSTEM_CD
                                                    ,BA.BILL_ACCOUNT_NAME
                                                    ,BA.CUST_NBR
                                                    ,C.CUST_NAME
                                                    ,CA.LOB_ID             
                                                    ,CA.EXTERNAL_RPT_SALES_CHAN_NAME
                                                    ,CA.SALES_CHAN_NAME
                                                    ,REPLACE( REPLACE(ba.bill_account_nbr ,'-',''),' ','') AS BILLING_ACCOUNT_ID__C
                                                    ,(CASE
                                                        WHEN (ba.dw_source_system_cd = 'ORACLE2E' OR ba.dw_source_system_cd = 'ORACLE2E_FED')
                                                            THEN
                                                                (
                                                                    CASE
                                                                        WHEN ba.finance_account_nbr LIKE '%-P-%'
                                                                            THEN 'QIA' || bill_account_nbr || '-P'
                                                                    ELSE 'QIA' || bill_account_nbr
                                                                    END
                                                                )
                                                        END) AS BILLING_UNIQUE_EXTERNAL_ID__C
                                                    ,ba.dw_source_system_cd AS BILLING_SYSTEM__C
                                                    ,BA.BILL_LINE1_ADDR
                                                    ,BA.BILL_LINE2_ADDR
                                                    ,BA.BILL_CITY_NAME
                                                    ,BA.BILL_STATE_CD
                                                    ,BA.BILL_POSTAL_CD
                                                    ,BILL_COUNTRY_ISO_3_CD
                                                    ,BA.DW_SECURE_COMPANY_NBR
                                                    ,CASE
                                                        WHEN BA.FINANCE_ACCOUNT_NBR  LIKE '%-ORACLE%'
                                                            THEN CASE
                                                                    WHEN BA.FINANCE_ACCOUNT_NBR LIKE '%-P-ORACLE%' THEN '012F0000000yFVmIAM'
                                                                    ELSE '012F0000000yEHtIAM'
                                                                END
                                                        ELSE NULL
                                                    END AS RECORDTYPEID
                                                    ,CASE
                                                        WHEN ba.dw_source_system_cd = 'ORACLE2E' THEN 'QIA'
                                                        ELSE NULL
                                                    END AS BILLING_INSTANCE__C
                                                    FROM CODS.BILLING_ACCOUNT BA
                                                        LEFT JOIN CODS.CUSTOMER C
                                                            ON C.CUST_ODS_ID = BA.CUST_ODS_ID
                                                        LEFT JOIN CODS.CUSTOMER_ATTRIBUTION CA
                                                            ON CA.CUST_ODS_ID = BA.CUST_ODS_ID
                                                    WHERE BA.DW_SOURCE_SYSTEM_CD = 'ORACLE2E'
                                                        AND BA.BILL_ACCT_ACTVTY_STAT_CALC_TYP != 'No Longer in Source'
                                                        AND BA.DW_SECURE_COMPANY_NBR NOT IN ('2','4')
                                                )
----------------------------------------------------------------------------------DO NOT RECOMMEND LIST
                                                ,DNR AS
                                                    (SELECT DISTINCT
                                                        C.CUST_NBR
                                                    FROM CODS.CUSTOMER C
                                                        LEFT JOIN CODS.CUSTOMER_ATTRIBUTION CA
                                                            ON CA.CUST_ODS_ID = C.CUST_ODS_ID
                                                        LEFT JOIN CODS_FINANCE.GL_CUSTOMER_SEGMENT CS
                                                            ON CS.GL_LOB_CD = CA.LOB_ID
                                                    WHERE 1=1
                                                        AND (CASE
                                                                WHEN UPPER(c.cust_name) like '%HOUSE ACCOUNT%' THEN 'KEEP'
                                                                WHEN UPPER(c.cust_name) like '%DO NOT USE%' THEN 'KEEP'
                                                                WHEN UPPER(c.cust_name) like 'CENTURYLINK%' THEN 'KEEP'
                                                                WHEN UPPER(c.cust_name) like 'LEVEL3%' THEN 'KEEP'
                                                                WHEN UPPER(c.cust_name) like 'LEVEL 3%' THEN 'KEEP'
                                                                WHEN UPPER(c.cust_name) like '%TRAINING ACCOUNT%' THEN 'KEEP'
                                                                WHEN UPPER(c.cust_name) like 'QWEST%' THEN 'KEEP'
                                                                WHEN UPPER(c.cust_name) like '%DUPLICATE%' THEN 'KEEP'
                                                                WHEN c.cust_name = 'NO VALID NAME' THEN 'KEEP'
                                                                WHEN UPPER(c.cust_line1_addr) LIKE '100 CENTURYLINK%' THEN 'KEEP'
                                                                WHEN sales_chan_name IN ('COLO','INTERNAL','CORPORATE') THEN 'KEEP'
                                                                WHEN c.ultimate_cust_nbr IN ('1-AEU-307-UC','1-H1PU17-UC') THEN 'KEEP'
                                                                WHEN c.cust_name = 'Small Business: No Revenue' THEN 'KEEP'
                                                                WHEN c.cust_name like 'ARCHIVE%' THEN 'KEEP'
                                                                WHEN c.cust_name like '%INACTIVE%' THEN 'KEEP'
                                                                WHEN c.cust_name LIKE '%SBG Unassigned%' THEN 'KEEP'
                                                                WHEN c.cust_name LIKE '%SBG Customer%' THEN 'KEEP'
                                                                WHEN c.cust_name LIKE '%SBG LATIS%' THEN 'KEEP'
                                                                WHEN c.cust_postal_cd = '71211-4065' THEN 'KEEP'
                                                                --or UPPER(cc.cust_line1_addr) LIKE '1025 ELDORADO BLVD%'
                                                                --or cc.cust_line1_addr LIKE '1025 Eldorado Boulevard%'
                                                                WHEN c.cust_class_typ IN('Competitor', 'Customer - CLEC', 'Iaas Target', 'Company Official', 'Customer - National', 'Partner', 'Payment Office') THEN 'KEEP'
                                                                WHEN c.cust_status_cd IN('Under Review', 'On Hold', 'Rejected', 'Inactive','New') THEN 'KEEP'
                                                                WHEN (UPPER(cust_line1_addr) like '%PO BOX 5003%' AND CUST_CITY_NAME = 'CAROL STREAM') THEN 'KEEP'
                                                                WHEN (UPPER(cust_line1_addr) like '1805 SHEA CENTER%' AND CUST_CITY_NAME = 'HIGHLANDS RANCH') THEN 'KEEP'
                                                                WHEN (UPPER(cust_line1_addr) like '9800 S MERIDIAN BLVD%' AND CUST_CITY_NAME = 'ENGLEWOOD') THEN 'KEEP'
                                                                WHEN cust_phone_nbr = '7208881000' THEN 'KEEP'
                                                                WHEN UPPER(C.CUST_NAME) LIKE 'DSR%' THEN 'KEEP'
                                                                WHEN UPPER(C.CUST_NAME) LIKE '%DSR%' THEN 'KEEP'
                                                                WHEN UPPER(C.CUST_NAME) LIKE 'DIVERSIFIED SYSTEMS RESOURCES' THEN 'KEEP'
                                                                WHEN c.cust_nbr in ('3-TVXNBMKCGN', '3-XHYWBJW7Q7', '3-4QJKVCWQPM', '3-A85231', '3-A85628', '3-849928') THEN 'KEEP'
                                                                WHEN C.CUST_NBR = '3-00003144' THEN 'KEEP'
                                                                WHEN gl_cust_type_cd = 'R' THEN 'KEEP'
                                                            ELSE 'OMIT' END
                                                            ) = 'KEEP')
                                                        --------------------------------------------------------------------------------
                                                        --------------------------------------------------------------------------------
                                                SELECT
                                                    --COUNT(T1.BILL_ACCOUNT_ODS_ID) AS BANS_TO_MODIFY
                                                    T1.PARENT_BILL_ACCOUNT_ODS_ID
                                                    ,T1.PARENT_BILL_ACCOUNT_ODS_ID2
                                                    ,T1.BILL_ACCOUNT_ODS_ID
                                                    ,T1.BILL_ACCOUNT_NBR
                                                    ,T1.FINANCE_ACCOUNT_NBR
                                                    ,T1.DW_SOURCE_SYSTEM_CD
                                                    ,T1.BILL_ACCOUNT_NAME
                                                    ,T1.CUST_NBR
                                                    ,T1.CUST_NAME
                                                    ,T1.LOB_ID             
                                                    ,T1.EXTERNAL_RPT_SALES_CHAN_NAME
                                                    ,T1.SALES_CHAN_NAME
                                                    ,T3.REC_CUST_LIST
                                                    ,T3.CNT_CUST_NBR
                                                    ,(CASE WHEN T1.CUST_NBR = T3.REC_CUST_LIST THEN 'CORRECT' ELSE 'MODIFY' END) AS REVIEW_FLAG
                                                    ,T1.BILL_ACCOUNT_NAME AS NAME
                                                    ,T1.BILLING_ACCOUNT_ID__C
                                                    ,T1.BILLING_UNIQUE_EXTERNAL_ID__C
                                                    ,T1.BILLING_SYSTEM__C
                                                    ,T1.BILL_LINE1_ADDR
                                                    ,T1.BILL_LINE2_ADDR
                                                    ,T1.BILL_CITY_NAME
                                                    ,T1.BILL_STATE_CD
                                                    ,T1.BILL_POSTAL_CD
                                                    ,T1.BILL_COUNTRY_ISO_3_CD
                                                    ,'QIA' AS BILLING_INSTANCE__C
                                                    ,T1.RECORDTYPEID
                                                FROM TABLE1 T1
                                                    LEFT JOIN
                                                        (SELECT
                                                            T2.PARENT_BILL_ACCOUNT_ODS_ID2
                                                            ,LISTAGG(T2.CUST_NBR, ',') AS REC_CUST_LIST
                                                            ,COUNT(T2.CUST_NBR) AS CNT_CUST_NBR
                                                        FROM
                                                            (SELECT DISTINCT
                                                                T1.PARENT_BILL_ACCOUNT_ODS_ID2
                                                                ,T1.CUST_NBR
                                                            FROM TABLE1 T1
                                                            WHERE 1=1
                                                                AND T1.CUST_NBR IS NOT NULL
                                                                AND T1.CUST_NBR NOT IN (
                                                                                        SELECT
                                                                                            D.CUST_NBR
                                                                                        FROM DNR D)
                                                                                        ) T2
                                                        GROUP BY T2.PARENT_BILL_ACCOUNT_ODS_ID2) T3
                                                            ON T3.PARENT_BILL_ACCOUNT_ODS_ID2 = T1.PARENT_BILL_ACCOUNT_ODS_ID2
                                                            WHERE 1=1
                                                        --AND T3.CNT_CUST_NBR IS NOT NULL
                                                        --AND T3.CNT_CUST_NBR = '1'
                                                        --AND T1.CUST_NBR != T3.REC_CUST_LIST
                                                        --AND T1.PARENT_BILL_ACCOUNT_ODS_ID2 = '94781826'
                                                        --AND (CASE WHEN T1.CUST_NBR = T3.REC_CUST_LIST THEN 'CORRECT' ELSE 'MODIFY' END) = 'MODIFY'
                                                        AND T1.PARENT_BILL_ACCOUNT_ODS_ID2 IN (
                                                                                                SELECT T1.PARENT_BILL_ACCOUNT_ODS_ID2 
                                                                                                    FROM TABLE1 T1 
                                                                                                WHERE T1.BILL_ACCOUNT_ODS_ID IN ({oracleBillingAccountsODSidsString})
                                                                                                -- WHERE T1.FINANCE_ACCOUNT_NBR IN ()
                                                                                               )
                                                ORDER BY T3.CNT_CUST_NBR DESC, T1.PARENT_BILL_ACCOUNT_ODS_ID2
                                                """

                df = pd.read_sql(oracle_query, conn)


                #write to the Oracle worksheet
                writer = pd.ExcelWriter(tickets_analysis_template_xlsx_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
                df.to_excel(writer, sheet_name = 'ORACLE', index=False) ### --> index=False so that the rows aren't numbered. Cuz that's annoying... and unnecessary
                                                                        ### --> At least, in this case since they are going to an excel doc anyhow
                writer.close()
                print(f"here are the oracleBillingAccounts, {oracleBillingAccounts}")
                #print("here is the specific oracle row: ", oracleBillingAccounts[x])
                list_int = 0
                all_parent_ods_ids_in_oracle_file = []
                for x in oracleBillingAccounts:
                    # print("here is x", x, "and length", len(x))
                    fan = x[0]
                    ods = x[1]
                    for i in range(len(df)):
                        all_parent_ods_ids_in_oracle_file.append(df.iloc[i, 1])
                        #print("inside the for loop for OrcleBillingaccounts, here is the fan and ods:", fan, ods)
                        #print("and the iloc:", df.iloc[i,1], df.iloc[i,2], df.iloc[i,3])

                        ### --> if the ods id matches the ods id in column 3 of the oracle dataframe
                        ### --> (as in, if the ods id for the Orcle ban in the oraclebillaaccounts list matches the BILL_ACCOUNT_ODS_ID in the dataframe
                        ### --> then we know what the parent_bill_account_ods_id is for the entire group. We attache that value
                        ### --> to the oracleBillingAccounts row where the match is, so that now we have the parent_bill_account_ods_id
                        ### --> available for us to evaluate later
                        if df.iloc[i, 2] == ods:
                            oracleBillingAccounts[list_int].append(df.iloc[i, 1])
                            #print(oracleBillingAccounts, " = there it is, the oracleBillingAccounts ")

                    ## --> this last if then is a fix for the Oracle query that Fabian wrote above it. I'm not sure why... yet... but Oracle accounts
                    ## -----> continually fall out of it. I'm truly not sure what it is looking for
                    ## --------- FIX FIX FIX 12/15/2023
                    if ods not in all_parent_ods_ids_in_oracle_file:
                        oracleBillingAccounts[list_int].append('None')

                    list_int+=1
            else:
                pass
            

            print("\n\n made it here \n ========++++++ \n aaaaa")
            """
            if df:
                for x in range(len(oracleBillingAccounts)):
                    for fan, ods in oracleBillingAccounts:
                        for i in range(len(df)):
                            print(fan, ods)
                            if df.iloc[i, 2] == ods:
                                oracleBillingAccounts[x].append(df.iloc[i, 1])
                                print(oracleBillingAccounts, " = there it is ")
            """

                            
            workbook = openpyxl.load_workbook(tickets_analysis_template_xlsx_path)
            worksheet = workbook.active

            ## tuples are produced below in the "row" assignment -- so we can only change the value with this code, or we need to convert the tuple
            #### --> and this is AFTER the sheet has already been written to - so the index values correspond to the columns on the Tickets sheet
            #### --> not the index values at the top that govern writing the particular index from the complete list that goes into the file
            for x in oracleBillingAccounts:
                print("looking at dis billing account now ", x)
                fan = x[0]
                ods = x[1]
                prnt_ods = x[2]
                for row in worksheet.rows:
                    billAcctName = row[12].value
                    #print("scr cd, fan, fan_ods_id, and fan we're hunting: ", row[6].value, row[8].value, row[9].value, fan)
                    #print(len(row))
                    if row[8].value == fan and not re.search(voice_services, billAcctName):
                        print("\n\n found that sumbitch:", row[8].value, "\n\n")
                        row[5].value = prnt_ods
                        row[1].value = "Find the ORACLE Family! Use the Parent ODS ID in column F to filter column B on the Oracle sheet" + row[1].value
                        row[0].value = "Review"
                    elif row[8].value == fan and re.search(voice_services, billAcctName):
                        row[5].value = prnt_ods
                        row[1].value = "This is an ORACLE Family Account, but it may also have voice services attached - " + row[1].value
                        row[0].value = "Review" # row[0].value + "/ORACLE Family"
            for row in worksheet.rows:
                if re.match('^.*\-A$', str(row[8].value)): # find a match on Kenan accounts by matching the "-A" at the end of the Finance Account Number
                    billAcctName = row[12].value
                    print("here is row11 and row19: ", billAcctName, row[21].value, row[16].value)
                    
                    if re.search(voice_services, billAcctName): 
                        row[1].value = "Kenan account w/ Voice products! - " + row[1].value
                    else:
                        pass
                        '''
                        row[0].value = "Review/Send to Ashley"
                        row[1].value = "Kenan Billing Account! - " + row[1].value
                        '''
                else:
                    continue
            

            
            assert os.path.isfile(tickets_analysis_template_xlsx_path)
            os.chmod(tickets_analysis_template_xlsx_path, stat.S_IRUSR | stat.S_IRGRP | stat.S_IROTH | stat.S_IWUSR | stat.S_IWGRP | stat.S_IWOTH) #to make sure we can access the file to write to
            workbook.save(tickets_analysis_template_xlsx_path)
                        
            #writer.close()

                
                
        # -->
        ### ----> prep MDM ANALYSIS sheet
        if os.path.isfile(mdm_analysis_template_path):
            with open(mdm_analysis_template_path, 'a+', newline='') as out:
                csv_out = csv.writer(out)
                book = []
                for index in analysis_sheet_nums:
                    for row in complete_info:
                        if row[0] == 'MDM MATCHING':
                            if row[14] in can_map_bans_list:
                                book.append('Map it!')
                                #print(f"\n\nhere's the index we're attempting to add to {row[14]} ", index)
                                #print(f"and here's the row: ", row)
                                #print(f"and the row count: ", len(row))
                                book.append(row[index])
                            #print("\n\n\nhere's the index!!: ", index , ' ', row[index])
                        csv_out.writerow(book)
                df = pd.read_csv(mdm_analysis_template_path, encoding = "ISO-8859-1", sep=',', names=['ZERO', 'ONE', 'TWO', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16',
                                                                     '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32',
                                                                     '33', '34', '35', '36'])
            for ci_row in complete_info:
                if ci_row[14] in can_map_bans_list:
                    df.insert(0, 'SUGGESTION', 'Map it!')
        else:       
            with open(mdm_analysis_template_path, 'a+', newline='', encoding="utf-8") as out:
                csv_out = csv.writer(out)
                book2 = []
                for row in complete_info:
                    book = []
                    #print("here's row[0]: " , row[0])
                    if row[0] == 'MDM MATCHING':
                        #print("here's the book right now: ", book)
                        #print("here's the row: ",  row)
                        #print("here's the fan in the row: ", row[14])
                        if row[14] in can_map_bans_list:
                            row.insert(0, "Map it!")
                        else:
                            row.insert(0, "Nope")
                        #print("n\n\nhere is the completed row including the suggestion column: ", row)


                        for index in analysis_sheet_nums:
                            #print("\n\n\nhere is the index being printed: " , index, " -- " , "\nand the corresponding row: ", row[index])
                            book.append(row[index])

                        #print("\n\n\nhere's the book right now (at the bottom): ", book)
                        book2.append(book)
                        #print("\n\n\nand here's book 2 -- this should get stacked full of lists: ", book2)
                            #print("\n\n\nhere's the index!!: ", index , ' ', row[index])

                #print("\n\n\n\nhere's the book2: ", book2)

                """        
                for book_row in book:
                print(book_row)
                book2.append([book_row])
                """
                csv_out.writerow(['SUGGESTION', 'EVALUATION', 'TICKET_ID', 'FINANCE_ACCOUNT_NBR',
                                  'BAN_ODS_ID',
                                  'BILL_ACCOUNT_NAME', 'BAN_DW_SECURE_COMPANY_NBR',
                                  'DW_CREATE_DT',
                                  'CRRNT_CUST_NBR',
                                  'CRRNT_CUST_ODS_ID',
                                  'CRRNT_CUST_NAME', 'CRRNT_CUST_DW_SECURE_COMPANY_NBR', 'CRRNT_CUST_STATUS',
                                  'RQST_CUST_NBR',
                                  'RQSTD_CUST_ODS_ID',
                                  'RQST_CUST_NAME', 'RQSTD_CUST_DW_SECURE_COMPANY_NBR', 'RQST_CUST_STATUS',
                                  'CRRNT_CUST_LOB', 'CRRNT_CUST_EXTRNL_RPRT_SALES_CHANNEL', 'CRRNT_CUST_SALES_CHAN', 'CRRNT_CUST_SALES_SUB_CHAN',
                                   'RQST_CUST_LOB', 'RQSTD_CUST_EXTRNL_RPRT_SALES_CHANNEL', 'RQSTD_CUST_SALES_CHAN', 'RQSTD_CUST_SALES_SUB_CHAN',
                                  'BAN_AVG_LAST_12_MONTHS_REV', 'CROSS_LOB', 'CROSS_SALES_CHANNEL', 'CROSS_BU',
                                  'CONFIDENCE_SCORE', 'RULE', 'NORMALIZED_SCORE',
                                  'TICKET_ID', 'FINANCE_ACCOUNT_NBR', 'CUST_NBR', 'NAME',
                                  'BILLING_ACCOUNT_ID__C', 'BILLING_UNIQUE_EXTERNAL_ID__C', 'BILLING_SYSTEM__C', 'BILLING_ADDRESS_LINE_1__C',
                                  'BILLING_ADDRESS_LINE_2__C', 'BILLING_CITY__C', 'BILLING_STATE__C', 'BILLING_POSTAL_CODE__C', 'BILLING_COUNTRY__C',
                                  'BAN_DW_SECURE_COMPANY_NBR',
                                  'BILLING_INSTANCE__C',
                                  'RECORDTYPEID', 'CUSTOMER_ACCOUNT__C'])
                for book_item in book2:
                    csv_out.writerow(book_item)
                out.close() ## --> file needs to be closed to begin work on changing it with pandas (below)
                    
    if len(cannot_map_list) > 0:
        if os.path.isfile(error_csv_path):
            with open(error_csv_path,'a+', newline='', encoding="utf-8") as out: ## w+ to read and write (just in case) a+ to append (read and write)
                csv_out=csv.writer(out)
                ##for result in results:
                for result in cannot_map_list:
                    csv_out.writerow(result)
                print("in the with loop to print ERRORS to csv")
        else:
            with open(error_csv_path,'a+', newline='', encoding="utf-8" ) as out: ## w+ to read and write (just in case) a+ to append (read and write)
                csv_out=csv.writer(out)
                print("did not find a csv exists, so going to print to ERRORS csv")
                csv_out.writerow([
                                'TICKET_ID', 'FINANCE_ACCOUNT_NBR', 'FAN_HAS_REVENUE', 'BAN_SOURCE_SYSTEM',
                                'BAN_AGE', 'CRRNT_CUST_LOB', 'CRRNT_CUST_DW_SECURE_COMPANY_NBR', 'BAN_DW_SECURE_COMPANY_NBR',
                                'RQSTD_CUST_NBR_CRPL',
                                'RQSTD_CUST_LOB',  'RQSTD_CUST_DW_SECURE_COMPANY_NBR', 'RQSTD_CUST_STATUS', 'RQSTD_CUST_HOUSE_ACCOUNT_CD', 'EXPLANATION'
                                  ]) # headers
                for result in cannot_map_list:
                    csv_out.writerow(result)
        
    '''        
    for result in results:
        print("this result: ", result[9])


    '''
    addDataValidationToExcelSheet(tickets_analysis_template_xlsx_path)
    print("The Billing Account to Customer Mapping automation process is complete")

    
   
if __name__ == "__main__":
    main()


